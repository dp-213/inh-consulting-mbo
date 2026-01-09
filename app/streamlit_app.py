import io
import json
import os
import subprocess
import sys
import hashlib
from datetime import datetime
import zipfile
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

NAV_OPTIONS = [
    "Overview",
    "Operating Model (P&L)",
    "Cashflow & Liquidity",
    "Balance Sheet",
    "Revenue Model",
    "Cost Model",
    "Other Assumptions",
    "Financing & Debt",
    "Equity Case",
    "Valuation & Purchase Price",
    "Model Settings",
]

_APP_DIR = os.path.dirname(__file__)
_ROOT_DIR = os.path.dirname(_APP_DIR)
if _ROOT_DIR not in sys.path:
    sys.path.insert(0, _ROOT_DIR)
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

try:
    from .data_model import InputModel, create_demo_input_model
    from .calculations.investment import _calculate_irr, calculate_investment
    from .calculations.pnl import calculate_pnl
    from .calculations.cashflow import calculate_cashflow
    from .calculations.debt import calculate_debt_schedule
    from .calculations.balance_sheet import calculate_balance_sheet
    from .run_model import run_model
    from .revenue_model import (
        render_revenue_model_assumptions,
        build_revenue_model_outputs,
    )
    from .cost_model import (
        render_cost_model_assumptions,
        build_cost_model_outputs,
    )
except (ImportError, KeyError):
    try:
        from app.data_model import InputModel, create_demo_input_model
        from app.calculations.investment import _calculate_irr, calculate_investment
        from app.calculations.pnl import calculate_pnl
        from app.calculations.cashflow import calculate_cashflow
        from app.calculations.debt import calculate_debt_schedule
        from app.calculations.balance_sheet import calculate_balance_sheet
        from app.run_model import run_model
        from app.revenue_model import (
            render_revenue_model_assumptions,
            build_revenue_model_outputs,
        )
        from app.cost_model import (
            render_cost_model_assumptions,
            build_cost_model_outputs,
        )
    except (ImportError, KeyError):
        from data_model import InputModel, create_demo_input_model
        from calculations.investment import _calculate_irr, calculate_investment
        from calculations.pnl import calculate_pnl
        from calculations.cashflow import calculate_cashflow
        from calculations.debt import calculate_debt_schedule
        from calculations.balance_sheet import calculate_balance_sheet
        from run_model import run_model
        from revenue_model import (
            render_revenue_model_assumptions,
            build_revenue_model_outputs,
        )
        from cost_model import (
            render_cost_model_assumptions,
            build_cost_model_outputs,
        )


def _pnl_dict_to_list(pnl_dict):
    # Convert P&L dict to a list with explicit year for downstream models.
    pnl_list = []
    for year_label, year_data in pnl_dict.items():
        try:
            year_index = int(str(year_label).split()[-1])
        except (ValueError, IndexError):
            year_index = year_label
        pnl_list.append({"year": year_index, **year_data})
    return pnl_list


def _format_section_title(section_key):
    return section_key.replace("_", " ").title()


def _percent_to_display(value):
    if value is None or pd.isna(value):
        return value
    return float(value) * 100


def _percent_from_display(value):
    if value is None or pd.isna(value):
        return value
    return float(value) / 100


def _format_number_display(value, decimals=0):
    if value is None or pd.isna(value):
        return ""
    try:
        number = float(str(value).replace(",", ""))
    except ValueError:
        return value
    if decimals == 0:
        return f"{number:,.0f}"
    return f"{number:,.{decimals}f}"


def _parse_number_display(value):
    if value is None or pd.isna(value) or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return value


def _apply_unit_display(df, value_col="Value", unit_col="Unit"):
    display_df = df.copy()
    if value_col in display_df.columns and unit_col in display_df.columns:
        display_df[value_col] = display_df.apply(
            lambda row: _format_number_display(
                _percent_to_display(row[value_col]), 1
            )
            if row[unit_col] == "%"
            else _format_number_display(row[value_col], 0),
            axis=1,
        )
    return display_df


def _restore_unit_values(df, value_col="Value", unit_col="Unit"):
    restored_df = df.copy()
    if value_col in restored_df.columns and unit_col in restored_df.columns:
        restored_df[value_col] = restored_df.apply(
            lambda row: _percent_from_display(_parse_number_display(row[value_col]))
            if row[unit_col] == "%" and isinstance(_parse_number_display(row[value_col]), (int, float))
            else _parse_number_display(row[value_col]),
            axis=1,
        )
    return restored_df


def _build_model_snapshot_payload(
    input_model,
    assumptions_state,
    pnl_result,
    cashflow_result,
    balance_sheet,
    debt_schedule,
    investment_result,
):
    scenario = st.session_state.get("output_scenario", "Base")
    try:
        commit = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
        ).decode("utf-8").strip()
    except Exception:
        commit = "unknown"

    revenue_state = assumptions_state.get("revenue_model", {})
    cost_state = assumptions_state.get("cost_model", {})

    payload = {
        "meta": {
            "timestamp": datetime.utcnow().isoformat(),
            "app_version": commit,
            "scenario": scenario,
            "currency": "EUR",
            "planning_horizon_years": 5,
        },
        "assumptions": {
            "revenue_model": revenue_state,
            "cost_model": cost_state,
            "financing": assumptions_state.get("financing", []),
            "equity": assumptions_state.get("equity", []),
            "cashflow": assumptions_state.get("cashflow", []),
            "balance_sheet": assumptions_state.get("balance_sheet", []),
            "valuation": assumptions_state.get("valuation", []),
        },
        "outputs": {
            "pnl": pnl_result,
            "cashflow": cashflow_result,
            "balance_sheet": balance_sheet,
            "debt_schedule": debt_schedule,
            "investment": investment_result,
        },
        "layout": {
            "Operating Model (P&L)": [
                "Full P&L statement (Years 0–4)",
                "P&L KPI block",
            ],
            "Cashflow & Liquidity": [
                "Cashflow statement (Years 0–4)",
                "Cashflow KPIs",
            ],
            "Balance Sheet": [
                "Simplified balance sheet (Years 0–4)",
                "Balance KPIs",
            ],
            "Financing & Debt": [
                "Debt schedule",
                "Debt service & covenants",
                "Financing KPIs",
            ],
            "Valuation & Purchase Price": [
                "Seller valuation range",
                "Buyer valuation (cash-based)",
                "Purchase price bridge",
                "Offer range (buyer view)",
                "Decision KPIs",
            ],
            "Equity Case": [
                "Deal structure",
                "Entry equity & ownership",
                "Headline returns (Investor / Management)",
                "Exit equity bridge",
                "Investor cashflows",
                "Management cashflows",
                "Equity KPIs",
            ],
            "Revenue Model": [
                "Derived consultant FTEs (from Cost Model)",
                "Capacity drivers (workdays, utilization, growth)",
                "Capacity allocation (group vs external)",
                "Pricing assumptions (group vs external)",
                "Guarantee inputs",
                "Revenue bridge summary",
            ],
            "Cost Model": [
                "Personnel cost inputs",
                "Overhead inputs",
                "Cost summary",
            ],
            "Other Assumptions": [
                "Financing assumptions",
                "Equity assumptions",
                "Cashflow assumptions",
                "Balance sheet assumptions",
                "Valuation assumptions",
            ],
        },
        "logic": [
            "Revenue capacity uses Consultant FTEs from the Cost Model.",
            "Capacity is split between group and external revenue; external does not add capacity.",
            "Group Revenue (after Floor Check) = max(modeled group revenue, reference × guarantee %).",
            "Final revenue = group revenue (after floor check) + modeled external revenue.",
            "P&L consumes final revenue and total operating costs (Cost Model).",
            "Operating cashflow = EBITDA – cash taxes – working capital delta.",
            "Free cashflow = operating cashflow – capex.",
            "Debt service = interest on opening debt + scheduled repayment.",
            "Equity IRR uses equity injections and exit proceeds only.",
        ],
        "notes": [
            "Distributions are pro-rata with no waterfall.",
            "Investor exits fully in the selected exit year.",
        ],
    }
    return payload


def _build_model_snapshot_zip(payload):
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr("snapshot.json", json.dumps(payload, indent=2))
        zip_file.writestr(
            "README.txt",
            "Model snapshot for AI review. Use snapshot.json as ground truth.",
        )
    buffer.seek(0)
    return buffer


def render_advanced_assumptions(input_model, show_header=True):
    def _local_clamp_pct(value):
        if value is None or pd.isna(value):
            return 0.0
        return max(0.0, min(float(value), 1.0))

    def _local_non_negative(value):
        if value is None or pd.isna(value):
            return 0.0
        return max(0.0, float(value))

    if show_header:
        st.title("Assumptions")
        st.write("Master input sheet – all model assumptions in one place")

    assumptions_state = st.session_state["assumptions"]

    st.markdown("### Financing Assumptions")
    financing_df = pd.DataFrame(assumptions_state["financing"])
    financing_display = _apply_unit_display(financing_df)
    financing_edit = st.data_editor(
        financing_display,
        hide_index=True,
        key="assumptions.financing",
        column_config={
            "Parameter": st.column_config.TextColumn(disabled=True),
            "Unit": st.column_config.TextColumn(disabled=True),
            "Notes": st.column_config.TextColumn(disabled=True),
            "Value": st.column_config.TextColumn(),
        },
        use_container_width=True,
    )
    financing_edit = _restore_unit_values(financing_edit)
    assumptions_state["financing"] = financing_edit.to_dict("records")
    for _, row in financing_edit.iterrows():
        parameter = row["Parameter"]
        if parameter == "Senior Debt Amount":
            senior_debt_amount = _local_non_negative(row["Value"])
            st.session_state["financing.senior_debt_amount"] = senior_debt_amount
            st.session_state["transaction_and_financing.senior_term_loan_start_eur"] = senior_debt_amount
        elif parameter == "Interest Rate":
            st.session_state["transaction_and_financing.senior_interest_rate_pct"] = _local_clamp_pct(row["Value"])
        elif parameter == "Amortisation Years":
            st.session_state["financing.amortization_period_years"] = int(max(1, row["Value"]))
        elif parameter == "Transaction Fees (%)":
            st.session_state["valuation.transaction_cost_pct"] = _local_clamp_pct(row["Value"])

    st.markdown("### Equity & Investor Assumptions")
    equity_df = pd.DataFrame(assumptions_state["equity"])
    equity_display = _apply_unit_display(equity_df)
    equity_edit = st.data_editor(
        equity_display,
        hide_index=True,
        key="assumptions.equity",
        column_config={
            "Parameter": st.column_config.TextColumn(disabled=True),
            "Unit": st.column_config.TextColumn(disabled=True),
            "Notes": st.column_config.TextColumn(disabled=True),
            "Value": st.column_config.TextColumn(),
        },
        use_container_width=True,
    )
    equity_edit = _restore_unit_values(equity_edit)
    assumptions_state["equity"] = equity_edit.to_dict("records")
    for _, row in equity_edit.iterrows():
        parameter = row["Parameter"]
        if parameter == "Sponsor Equity Contribution":
            st.session_state["equity.sponsor_equity_eur"] = _local_non_negative(row["Value"])
        elif parameter == "Investor Equity Contribution":
            st.session_state["equity.investor_equity_eur"] = _local_non_negative(row["Value"])
        elif parameter == "Investor Exit Year":
            try:
                exit_val = int(float(row["Value"]))
            except (TypeError, ValueError):
                exit_val = _default_equity_assumptions(input_model)["exit_year"]
            st.session_state["equity.exit_year"] = int(
                max(3, min(7, exit_val))
            )
        elif parameter == "Exit Multiple (x EBITDA)":
            st.session_state["equity.exit_multiple"] = float(row["Value"])

    st.markdown("### Cashflow Assumptions")
    cashflow_df = pd.DataFrame(assumptions_state["cashflow"])
    cashflow_display = _apply_unit_display(cashflow_df)
    cashflow_edit = st.data_editor(
        cashflow_display,
        hide_index=True,
        key="assumptions.cashflow",
        column_config={
            "Parameter": st.column_config.TextColumn(disabled=True),
            "Unit": st.column_config.TextColumn(disabled=True),
            "Notes": st.column_config.TextColumn(disabled=True),
            "Value": st.column_config.TextColumn(),
        },
        use_container_width=True,
    )
    cashflow_edit = _restore_unit_values(cashflow_edit)
    assumptions_state["cashflow"] = cashflow_edit.to_dict("records")
    for _, row in cashflow_edit.iterrows():
        parameter = row["Parameter"]
        if parameter == "Tax Cash Rate":
            st.session_state["cashflow.tax_cash_rate_pct"] = _local_clamp_pct(row["Value"])
        elif parameter == "Tax Payment Lag":
            st.session_state["cashflow.tax_payment_lag_years"] = int(max(0, min(1, row["Value"])))
        elif parameter == "Capex (% of Revenue)":
            st.session_state["cashflow.capex_pct_revenue"] = _local_clamp_pct(row["Value"])
        elif parameter == "Working Capital (% of Revenue)":
            st.session_state["cashflow.working_capital_pct_revenue"] = _local_clamp_pct(row["Value"])
        elif parameter == "Opening Cash Balance":
            st.session_state["cashflow.opening_cash_balance_eur"] = _local_non_negative(row["Value"])

    st.markdown("### Balance Sheet Assumptions")
    balance_df = pd.DataFrame(assumptions_state["balance_sheet"])
    balance_display = _apply_unit_display(balance_df)
    balance_edit = st.data_editor(
        balance_display,
        hide_index=True,
        key="assumptions.balance_sheet",
        column_config={
            "Parameter": st.column_config.TextColumn(disabled=True),
            "Unit": st.column_config.TextColumn(disabled=True),
            "Notes": st.column_config.TextColumn(disabled=True),
            "Value": st.column_config.TextColumn(),
        },
        use_container_width=True,
    )
    balance_edit = _restore_unit_values(balance_edit)
    assumptions_state["balance_sheet"] = balance_edit.to_dict("records")
    for _, row in balance_edit.iterrows():
        parameter = row["Parameter"]
        if parameter == "Opening Equity":
            st.session_state["balance_sheet.opening_equity_eur"] = _local_non_negative(row["Value"])
        elif parameter == "Depreciation Rate":
            st.session_state["balance_sheet.depreciation_rate_pct"] = _local_clamp_pct(row["Value"])
        elif parameter == "Minimum Cash Balance":
            st.session_state["balance_sheet.minimum_cash_balance_eur"] = _local_non_negative(row["Value"])

    st.markdown("### Valuation Assumptions")
    valuation_df = pd.DataFrame(assumptions_state["valuation"])
    valuation_display = _apply_unit_display(valuation_df)
    valuation_edit = st.data_editor(
        valuation_display,
        hide_index=True,
        key="assumptions.valuation",
        column_config={
            "Parameter": st.column_config.TextColumn(disabled=True),
            "Unit": st.column_config.TextColumn(disabled=True),
            "Notes": st.column_config.TextColumn(disabled=True),
            "Value": st.column_config.TextColumn(),
        },
        use_container_width=True,
    )
    valuation_edit = _restore_unit_values(valuation_edit)
    assumptions_state["valuation"] = valuation_edit.to_dict("records")
    for _, row in valuation_edit.iterrows():
        parameter = row["Parameter"]
        if parameter == "Seller EBITDA Multiple":
            st.session_state["valuation.seller_ebit_multiple"] = float(row["Value"])
        elif parameter == "Reference Year":
            st.session_state["valuation.reference_year"] = int(max(0, min(4, row["Value"])))
        elif parameter == "Discount Rate (WACC)":
            st.session_state["valuation.buyer_discount_rate"] = _local_clamp_pct(row["Value"])
        elif parameter == "Valuation Start Year":
            st.session_state["valuation.valuation_start_year"] = int(max(0, min(4, row["Value"])))
        elif parameter == "Transaction Costs (%)":
            st.session_state["valuation.transaction_cost_pct"] = _local_clamp_pct(row["Value"])

    apply_fn = globals().get("_apply_assumptions_state")
    if callable(apply_fn):
        apply_fn()

def format_currency(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, str):
        return value
    return f"{value / 1_000_000:,.2f} m EUR"


def format_pct(value):
    if value is None or pd.isna(value) or value == "":
        return ""
    return f"{value:.1%}"


def format_int(value):
    if value is None or pd.isna(value):
        return ""
    return f"{int(round(value)):,}"


def _clamp_pct(value):
    if value is None or pd.isna(value):
        return 0.0
    return max(0.0, min(float(value), 1.0))


def _non_negative(value):
    if value is None or pd.isna(value):
        return 0.0
    return max(0.0, float(value))


def _apply_assumptions_state():
    state = st.session_state["assumptions"]
    active_scenario = st.session_state.get("assumptions.scenario", "Base")
    scenario_col = active_scenario

    revenue_state = state.get("revenue_model", {})
    if revenue_state:
        st.session_state["revenue_model.reference_revenue_eur"] = _non_negative(
            revenue_state["reference_revenue_eur"].get(scenario_col, 0.0)
        )
        for year_index in range(5):
            st.session_state[
                f"revenue_model.guarantee_pct_year_{year_index}"
            ] = _clamp_pct(
                revenue_state["guarantee_pct_by_year"][scenario_col][
                    year_index
                ]
            )
            st.session_state[
                f"revenue_model.workdays_year_{year_index}"
            ] = _non_negative(
                revenue_state["workdays_per_year"][scenario_col][year_index]
            )
            st.session_state[
                f"revenue_model.utilization_rate_year_{year_index}"
            ] = _clamp_pct(
                revenue_state["utilization_rate"][scenario_col][year_index]
            )
            st.session_state[
                f"revenue_model.group_day_rate_eur_year_{year_index}"
            ] = _non_negative(
                revenue_state["group_day_rate_eur"][scenario_col][year_index]
            )
            st.session_state[
                f"revenue_model.external_day_rate_eur_year_{year_index}"
            ] = _non_negative(
                revenue_state["external_day_rate_eur"][scenario_col][year_index]
            )
            st.session_state[
                f"revenue_model.day_rate_growth_pct_year_{year_index}"
            ] = _clamp_pct(
                revenue_state["day_rate_growth_pct"][scenario_col][
                    year_index
                ]
            )
            st.session_state[
                f"revenue_model.revenue_growth_pct_year_{year_index}"
            ] = _clamp_pct(
                revenue_state["revenue_growth_pct"][scenario_col][
                    year_index
                ]
            )
            st.session_state[
                f"revenue_model.group_capacity_share_pct_year_{year_index}"
            ] = _clamp_pct(
                revenue_state["group_capacity_share_pct"][scenario_col][
                    year_index
                ]
            )
            st.session_state[
                f"revenue_model.external_capacity_share_pct_year_{year_index}"
            ] = _clamp_pct(
                revenue_state["external_capacity_share_pct"][scenario_col][
                    year_index
                ]
            )

    cost_state = state.get("cost_model", {})
    if cost_state:
        if "consultant_fte" in cost_state:
            for year_index in range(5):
                st.session_state[
                    f"cost_model.consultant_fte_year_{year_index}"
                ] = _non_negative(
                    cost_state["consultant_fte"][scenario_col][year_index]
                )
                st.session_state[
                    f"cost_model.consultant_base_cost_eur_year_{year_index}"
                ] = _non_negative(
                    cost_state["consultant_costs"][scenario_col][year_index]
                )
                st.session_state[
                    f"cost_model.backoffice_fte_year_{year_index}"
                ] = _non_negative(
                    cost_state["backoffice_fte"][scenario_col][year_index]
                )
                st.session_state[
                    f"cost_model.backoffice_base_cost_eur_year_{year_index}"
                ] = _non_negative(
                    cost_state["backoffice_costs"][scenario_col][year_index]
                )
                st.session_state[
                    f"cost_model.fixed_overhead_advisory_year_{year_index}"
                ] = _non_negative(
                    cost_state["fixed_overhead"][scenario_col][year_index][
                        "Advisory"
                    ]
                )
                st.session_state[
                    f"cost_model.fixed_overhead_legal_year_{year_index}"
                ] = _non_negative(
                    cost_state["fixed_overhead"][scenario_col][year_index][
                        "Legal"
                    ]
                )
                st.session_state[
                    f"cost_model.fixed_overhead_it_year_{year_index}"
                ] = _non_negative(
                    cost_state["fixed_overhead"][scenario_col][year_index][
                        "IT & Software"
                    ]
                )
                st.session_state[
                    f"cost_model.fixed_overhead_office_year_{year_index}"
                ] = _non_negative(
                    cost_state["fixed_overhead"][scenario_col][year_index][
                        "Office Rent"
                    ]
                )
                st.session_state[
                    f"cost_model.fixed_overhead_services_year_{year_index}"
                ] = _non_negative(
                    cost_state["fixed_overhead"][scenario_col][year_index][
                        "Services"
                    ]
                )
                st.session_state[
                    f"cost_model.variable_training_pct_year_{year_index}"
                ] = _clamp_pct(
                    cost_state["variable_costs"][scenario_col][year_index][
                        "Training"
                    ]
                )
                st.session_state[
                    f"cost_model.variable_travel_pct_year_{year_index}"
                ] = _clamp_pct(
                    cost_state["variable_costs"][scenario_col][year_index][
                        "Travel"
                    ]
                )
                st.session_state[
                    f"cost_model.variable_communication_pct_year_{year_index}"
                ] = _clamp_pct(
                    cost_state["variable_costs"][scenario_col][year_index][
                        "Communication"
                    ]
                )
        elif "personnel" in cost_state:
            if "inflation" in cost_state:
                st.session_state["cost_model.apply_inflation"] = bool(
                    cost_state["inflation"].get("apply", False)
                )
                st.session_state["cost_model.inflation_rate_pct"] = _clamp_pct(
                    cost_state["inflation"].get("rate_pct", 0.0)
                )
            for row in cost_state.get("personnel", []):
                year_index = int(row["Year"].split()[-1])
                st.session_state[
                    f"cost_model.consultant_fte_year_{year_index}"
                ] = _non_negative(row["Consultant FTE"])
                st.session_state[
                    f"cost_model.consultant_base_cost_eur_year_{year_index}"
                ] = _non_negative(row["Consultant Loaded Cost (EUR)"])
                st.session_state[
                    f"cost_model.backoffice_fte_year_{year_index}"
                ] = _non_negative(row["Backoffice FTE"])
                st.session_state[
                    f"cost_model.backoffice_base_cost_eur_year_{year_index}"
                ] = _non_negative(row["Backoffice Loaded Cost (EUR)"])
                st.session_state[
                    f"cost_model.management_cost_eur_year_{year_index}"
                ] = _non_negative(row["Management Cost (EUR)"])

            for row in cost_state.get("fixed_overhead", []):
                year_index = int(row["Year"].split()[-1])
                st.session_state[
                    f"cost_model.fixed_overhead_advisory_year_{year_index}"
                ] = _non_negative(row["Advisory"])
                st.session_state[
                    f"cost_model.fixed_overhead_legal_year_{year_index}"
                ] = _non_negative(row["Legal"])
                st.session_state[
                    f"cost_model.fixed_overhead_it_year_{year_index}"
                ] = _non_negative(row["IT & Software"])
                st.session_state[
                    f"cost_model.fixed_overhead_office_year_{year_index}"
                ] = _non_negative(row["Office Rent"])
                st.session_state[
                    f"cost_model.fixed_overhead_services_year_{year_index}"
                ] = _non_negative(row["Services"])
                st.session_state[
                    f"cost_model.fixed_overhead_other_year_{year_index}"
                ] = _non_negative(row["Other Services"])

            for row in cost_state.get("variable_costs", []):
                year_index = int(row["Year"].split()[-1])
                training_value = _non_negative(row["Training Value"])
                travel_value = _non_negative(row["Travel Value"])
                communication_value = _non_negative(row["Communication Value"])

                st.session_state[
                    f"cost_model.variable_training_pct_year_{year_index}"
                ] = training_value if row["Training Type"] == "%" else 0.0
                st.session_state[
                    f"cost_model.variable_training_eur_year_{year_index}"
                ] = training_value if row["Training Type"] == "EUR" else 0.0
                st.session_state[
                    f"cost_model.variable_travel_pct_year_{year_index}"
                ] = travel_value if row["Travel Type"] == "%" else 0.0
                st.session_state[
                    f"cost_model.variable_travel_eur_year_{year_index}"
                ] = travel_value if row["Travel Type"] == "EUR" else 0.0
                st.session_state[
                    f"cost_model.variable_communication_pct_year_{year_index}"
                ] = (
                    communication_value
                    if row["Communication Type"] == "%"
                    else 0.0
                )
                st.session_state[
                    f"cost_model.variable_communication_eur_year_{year_index}"
                ] = (
                    communication_value
                    if row["Communication Type"] == "EUR"
                    else 0.0
                )

    for row in state.get("financing", []):
        param = row["Parameter"]
        if param == "Senior Debt Amount":
            senior_debt_amount = _non_negative(row["Value"])
            st.session_state["financing.senior_debt_amount"] = senior_debt_amount
            st.session_state[
                "transaction_and_financing.senior_term_loan_start_eur"
            ] = senior_debt_amount
        elif param == "Interest Rate":
            st.session_state[
                "transaction_and_financing.senior_interest_rate_pct"
            ] = _clamp_pct(row["Value"])
        elif param == "Amortisation Years":
            st.session_state["financing.amortization_period_years"] = int(
                max(1, row["Value"])
            )
        elif param == "Transaction Fees (%)":
            st.session_state["valuation.transaction_cost_pct"] = _clamp_pct(
                row["Value"]
            )

    for row in state.get("equity", []):
        param = row["Parameter"]
        if param == "Sponsor Equity Contribution":
            st.session_state["equity.sponsor_equity_eur"] = _non_negative(
                row["Value"]
            )
        elif param == "Investor Equity Contribution":
            st.session_state["equity.investor_equity_eur"] = _non_negative(
                row["Value"]
            )
        elif param == "Investor Exit Year":
            try:
                exit_year = int(float(row["Value"]))
            except (TypeError, ValueError):
                exit_year = _default_equity_assumptions(
                    create_demo_input_model()
                )["exit_year"]
            st.session_state["equity.exit_year"] = int(
                max(3, min(7, exit_year))
            )
        elif param == "Exit Multiple (x EBITDA)":
            st.session_state["equity.exit_multiple"] = _non_negative(row["Value"])

    for row in state.get("cashflow", []):
        param = row["Parameter"]
        if param == "Tax Cash Rate":
            st.session_state["cashflow.tax_cash_rate_pct"] = _clamp_pct(
                row["Value"]
            )
        elif param == "Tax Payment Lag":
            st.session_state["cashflow.tax_payment_lag_years"] = int(
                max(0, min(1, row["Value"]))
            )
        elif param == "Capex (% of Revenue)":
            st.session_state["cashflow.capex_pct_revenue"] = _clamp_pct(
                row["Value"]
            )
        elif param == "Working Capital (% of Revenue)":
            st.session_state["cashflow.working_capital_pct_revenue"] = _clamp_pct(
                row["Value"]
            )
        elif param == "Opening Cash Balance":
            st.session_state["cashflow.opening_cash_balance_eur"] = _non_negative(
                row["Value"]
            )

    for row in state.get("balance_sheet", []):
        param = row["Parameter"]
        if param == "Opening Equity":
            st.session_state["balance_sheet.opening_equity_eur"] = _non_negative(
                row["Value"]
            )
        elif param == "Depreciation Rate":
            st.session_state["balance_sheet.depreciation_rate_pct"] = _clamp_pct(
                row["Value"]
            )
        elif param == "Minimum Cash Balance":
            st.session_state["balance_sheet.minimum_cash_balance_eur"] = _non_negative(
                row["Value"]
            )

    for row in state.get("valuation", []):
        param = row["Parameter"]
        if param == "Seller EBITDA Multiple":
            st.session_state["valuation.seller_ebit_multiple"] = _non_negative(
                row["Value"]
            )
        elif param == "Reference Year":
            st.session_state["valuation.reference_year"] = int(
                max(0, row["Value"])
            )
        elif param == "Discount Rate (WACC)":
            st.session_state["valuation.buyer_discount_rate"] = _clamp_pct(
                row["Value"]
            )
        elif param == "Valuation Start Year":
            st.session_state["valuation.valuation_start_year"] = int(
                max(0, row["Value"])
            )
        elif param == "Transaction Costs (%)":
            st.session_state["valuation.transaction_cost_pct"] = _clamp_pct(
                row["Value"]
            )

def _style_totals(df, columns_to_bold):
    def style_row(_):
        return [
            "font-weight: 600;" if col in columns_to_bold else ""
            for col in df.columns
        ]

    return df.style.apply(style_row, axis=1)


def _default_cashflow_assumptions():
    return {
        "tax_cash_rate_pct": 0.30,
        "tax_payment_lag_years": 1,
        "capex_pct_revenue": 0.01,
        "working_capital_pct_revenue": 0.02,
        "opening_cash_balance_eur": 250000,
    }


def _default_balance_sheet_assumptions(input_model):
    opening_equity = _default_cashflow_assumptions()["opening_cash_balance_eur"]
    minimum_cash = input_model.capex_and_working_capital[
        "minimum_cash_balance_eur"
    ].value
    return {
        "opening_equity_eur": opening_equity,
        "depreciation_rate_pct": 0.20,
        "minimum_cash_balance_eur": minimum_cash,
    }


def _default_financing_assumptions(input_model):
    cashflow_defaults = _default_cashflow_assumptions()
    return {
        "senior_debt_amount": input_model.transaction_and_financing[
            "senior_term_loan_start_eur"
        ].value,
        "initial_debt_eur": input_model.transaction_and_financing[
            "senior_term_loan_start_eur"
        ].value,
        "interest_rate_pct": input_model.transaction_and_financing[
            "senior_interest_rate_pct"
        ].value,
        "amortization_type": "Linear",
        "amortization_period_years": 5,
        "grace_period_years": 0,
        "special_repayment_year": None,
        "special_repayment_amount_eur": 0.0,
        "minimum_dscr": 1.3,
        "minimum_cash_balance_eur": input_model.capex_and_working_capital[
            "minimum_cash_balance_eur"
        ].value,
        "maintenance_capex_pct_revenue": cashflow_defaults[
            "capex_pct_revenue"
        ],
    }


def _default_valuation_assumptions(input_model):
    return {
        "seller_ebit_multiple": input_model.valuation_assumptions[
            "multiple_valuation"
        ]["seller_multiple"].value
        or 0.0,
        "reference_year": 1,
        "buyer_discount_rate": input_model.valuation_assumptions["dcf_valuation"][
            "discount_rate_wacc"
        ].value
        or 0.10,
        "valuation_start_year": 0,
        "debt_at_close_eur": input_model.transaction_and_financing[
            "senior_term_loan_start_eur"
        ].value,
        "transaction_cost_pct": 0.01,
        "include_terminal_value": False,
    }


def _default_equity_assumptions(input_model):
    return {
        "sponsor_equity_eur": 2750000,
        "investor_equity_eur": 5750000,
        "exit_year": 4,
        "exit_method": "Exit Multiple",
        "exit_multiple": 7.0,
    }


def _seed_session_defaults(input_model):
    def _seed_section(section_data, prefix=""):
        for key, value in section_data.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if hasattr(value, "value"):
                st.session_state.setdefault(full_key, value.value)
            elif isinstance(value, dict):
                _seed_section(value, full_key)

    for section_key, section_value in input_model.__dict__.items():
        if isinstance(section_value, dict):
            _seed_section(section_value, section_key)

    selected_scenario = st.session_state.get(
        "scenario_selection.selected_scenario",
        input_model.scenario_selection["selected_scenario"].value,
    )
    scenario_key = selected_scenario.lower()
    base_utilization = input_model.scenario_parameters["utilization_rate"][
        scenario_key
    ].value
    utilization_curve = []
    for year_index in range(5):
        utilization_curve.append(min(base_utilization + 0.005 * year_index, 0.68))
    st.session_state.setdefault("utilization_by_year", utilization_curve)
    for year_index in range(5):
        st.session_state.setdefault(
            f"utilization_by_year.{year_index}",
            st.session_state["utilization_by_year"][year_index],
        )

    for key, value in _default_cashflow_assumptions().items():
        st.session_state.setdefault(f"cashflow.{key}", value)

    for key, value in _default_balance_sheet_assumptions(input_model).items():
        st.session_state.setdefault(f"balance_sheet.{key}", value)

    for key, value in _default_financing_assumptions(input_model).items():
        st.session_state.setdefault(f"financing.{key}", value)

    for key, value in _default_valuation_assumptions(input_model).items():
        st.session_state.setdefault(f"valuation.{key}", value)

    for key, value in _default_equity_assumptions(input_model).items():
        st.session_state.setdefault(f"equity.{key}", value)


def _render_input_field(
    input_field, widget_key, scenario_tag=None, scenario_help=None
):
    label = input_field.description
    if scenario_tag:
        label = f"{label} {scenario_tag}"
    help_text = f"Excel: {input_field.excel_ref}"
    if scenario_help:
        help_text = f"{help_text} | {scenario_help}"
    value = input_field.value

    if isinstance(value, bool):
        return st.checkbox(
            label,
            value=value,
            help=help_text,
            key=widget_key,
            disabled=not input_field.editable,
        )
    if isinstance(value, (int, float)):
        step = 1.0 if isinstance(value, int) and not isinstance(value, bool) else 0.01
        return st.number_input(
            label,
            value=float(value),
            step=step,
            help=help_text,
            key=widget_key,
            disabled=not input_field.editable,
        )
    if value is None:
        text_value = st.text_input(
            label,
            value="",
            help=help_text,
            key=widget_key,
            disabled=not input_field.editable,
        )
        return None if text_value == "" else text_value

    return st.text_input(
        label,
        value=str(value),
        help=help_text,
        key=widget_key,
        disabled=not input_field.editable,
    )


def _render_section(
    section_data, section_key, selected_scenario=None, is_scenario_section=False
):
    edited_values = {}
    for key, value in section_data.items():
        field_key = f"{section_key}.{key}"
        if hasattr(value, "value") and hasattr(value, "editable"):
            edited_values[key] = _render_input_field(value, field_key)
        elif isinstance(value, dict):
            st.markdown(f"**{_format_section_title(key)}**")
            if is_scenario_section:
                edited_values[key] = {}
                for scenario_key, scenario_field in value.items():
                    scenario_tag = (
                        "(Scenario-driven)"
                        if scenario_key == selected_scenario
                        else "(Scenario parameter)"
                    )
                    scenario_help = (
                        "Active scenario input"
                        if scenario_key == selected_scenario
                        else "Not active for selected scenario"
                    )
                    scenario_field_key = (
                        f"{field_key}.{scenario_key}"
                    )
                    edited_values[key][scenario_key] = _render_input_field(
                        scenario_field,
                        scenario_field_key,
                        scenario_tag=scenario_tag,
                        scenario_help=scenario_help,
                    )
            else:
                edited_values[key] = _render_section(
                    value,
                    field_key,
                    selected_scenario=selected_scenario,
                    is_scenario_section=is_scenario_section,
                )
    return edited_values


def _apply_section_values(section_data, edited_values):
    for key, value in edited_values.items():
        if hasattr(section_data.get(key), "value"):
            section_data[key].value = value
        elif isinstance(section_data.get(key), dict):
            _apply_section_values(section_data[key], value)


def _collect_values_from_session(section_data, section_key):
    values = {}
    for key, value in section_data.items():
        field_key = f"{section_key}.{key}"
        if hasattr(value, "value"):
            if field_key in st.session_state:
                values[key] = st.session_state[field_key]
            else:
                values[key] = value.value
        elif isinstance(value, dict):
            values[key] = _collect_values_from_session(value, field_key)
    return values


def _get_field_by_path(base_model, path_parts):
    current = base_model
    for part in path_parts:
        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None
    if hasattr(current, "value"):
        return current
    return None


def _set_field_value(field_key, value):
    st.session_state[field_key] = value


def _get_current_value(field_key, fallback):
    return st.session_state.get(field_key, fallback)


def _render_inline_controls(title, controls, columns=3):
    st.subheader(title)
    cols = st.columns(columns)
    for index, control in enumerate(controls):
        col = cols[index % columns]
        with col:
            widget_key = f"inline.{title}.{control['field_key']}"
            if control["type"] == "select":
                selection = st.selectbox(
                    control["label"],
                    control["options"],
                    index=control["index"],
                    key=widget_key,
                )
                if isinstance(control["options"][0], str) and control["options"] == [
                    "Off",
                    "On",
                ]:
                    _set_field_value(
                        control["field_key"], True if selection == "On" else False
                    )
                elif isinstance(control["options"][0], str) and any(
                    option.startswith("Year ") for option in control["options"]
                ):
                    if selection == "None":
                        _set_field_value(control["field_key"], None)
                    else:
                        _set_field_value(
                            control["field_key"], int(selection.split()[-1])
                        )
                else:
                    _set_field_value(control["field_key"], selection)
            else:
                if control["type"] == "pct":
                    ui_value = st.number_input(
                        control["label"],
                        value=float(control["value"] * 100),
                        step=0.1,
                        format="%.1f",
                        key=widget_key,
                    )
                    _set_field_value(control["field_key"], ui_value / 100)
                elif control["type"] == "int":
                    value = st.number_input(
                        control["label"],
                        value=float(control["value"]),
                        step=1.0,
                        format="%.0f",
                        key=widget_key,
                    )
                    _set_field_value(control["field_key"], value)
                else:
                    value = st.number_input(
                        control["label"],
                        value=float(control["value"]),
                        step=control.get("step", 1.0),
                        format=control.get("format", "%.0f"),
                        key=widget_key,
                    )
                    _set_field_value(control["field_key"], value)


def _render_pnl_html(pnl_statement, section_rows, bold_rows):
    def escape(text):
        return (
            str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    columns = ["Line Item", "Year 0", "Year 1", "Year 2", "Year 3", "Year 4"]
    header_cells = "".join(f"<th>{escape(col)}</th>" for col in columns)

    body_rows = []
    for _, row in pnl_statement.iterrows():
        label = row["Line Item"]
        row_class = ""
        if label in section_rows:
            row_class = "section-row"
        elif label in bold_rows:
            row_class = "total-row"
        cells = []
        for col in columns:
            value = row[col]
            if col != "Line Item":
                if label in {
                    "EBITDA Margin",
                    "EBIT Margin",
                    "Personnel Cost Ratio",
                    "Guaranteed Revenue %",
                    "Net Margin",
                    "Opex Ratio",
                }:
                    value = format_pct(value)
                else:
                    value = format_currency(value)
            cell_value = "&nbsp;" if value in ("", None) else escape(value)
            cells.append(f"<td>{cell_value}</td>")
        body_rows.append(f"<tr class=\"{row_class}\">{''.join(cells)}</tr>")

    css = """
    <style>
      .pnl-table { width: 100%; border-collapse: collapse; table-layout: fixed; }
      .pnl-table col.line-item { width: 42%; }
      .pnl-table col.year { width: 11.6%; }
      .pnl-table th, .pnl-table td {
        padding: 2px 6px;
        white-space: nowrap;
        line-height: 1.0;
        border: 0;
        font-size: 0.9rem;
      }
      .pnl-table th { text-align: right; font-weight: 600; }
      .pnl-table th:first-child { text-align: left; }
      .pnl-table td { text-align: right; }
      .pnl-table td:first-child { text-align: left; }
      .pnl-table .section-row td {
        font-weight: 700;
        background: #f9fafb;
      }
      .pnl-table .total-row td {
        font-weight: 700;
        background: #f3f4f6;
        border-top: 1px solid #c7c7c7;
      }
    </style>
    """
    colgroup = (
        "<colgroup>"
        "<col class=\"line-item\"/>"
        "<col class=\"year\"/><col class=\"year\"/><col class=\"year\"/>"
        "<col class=\"year\"/><col class=\"year\"/>"
        "</colgroup>"
    )
    table_html = (
        f"{css}<table class=\"pnl-table\">{colgroup}"
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody></table>"
    )
    st.markdown(table_html, unsafe_allow_html=True)


def _render_cashflow_html(cashflow_statement, section_rows, bold_rows):
    def escape(text):
        return (
            str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    columns = ["Line Item", "Year 0", "Year 1", "Year 2", "Year 3", "Year 4"]
    header_cells = "".join(f"<th>{escape(col)}</th>" for col in columns)

    body_rows = []
    for _, row in cashflow_statement.iterrows():
        label = row["Line Item"]
        row_class = ""
        if label in section_rows:
            row_class = "section-row"
        elif label in bold_rows:
            row_class = "total-row"
        cells = []
        for col in columns:
            value = row[col]
            cell_class = ""
            if col != "Line Item":
                value = format_currency(value)
                try:
                    if float(row[col]) < 0:
                        cell_class = "negative"
                except (TypeError, ValueError):
                    cell_class = ""
            cell_value = "&nbsp;" if value in ("", None) else escape(value)
            cells.append(f"<td class=\"{cell_class}\">{cell_value}</td>")
        body_rows.append(f"<tr class=\"{row_class}\">{''.join(cells)}</tr>")

    css = """
    <style>
      .cashflow-table { width: 100%; border-collapse: collapse; table-layout: fixed; }
      .cashflow-table col.line-item { width: 42%; }
      .cashflow-table col.year { width: 11.6%; }
      .cashflow-table th, .cashflow-table td {
        padding: 2px 6px;
        white-space: nowrap;
        line-height: 1.0;
        border: 0;
        font-size: 0.9rem;
      }
      .cashflow-table th { text-align: right; font-weight: 600; }
      .cashflow-table th:first-child { text-align: left; }
      .cashflow-table td { text-align: right; }
      .cashflow-table td:first-child { text-align: left; }
      .cashflow-table .section-row td {
        font-weight: 700;
        background: #f9fafb;
      }
      .cashflow-table .total-row td {
        font-weight: 700;
        background: #f3f4f6;
        border-top: 1px solid #c7c7c7;
      }
      .cashflow-table td.negative { color: #b45309; }
    </style>
    """
    colgroup = (
        "<colgroup>"
        "<col class=\"line-item\"/>"
        "<col class=\"year\"/><col class=\"year\"/><col class=\"year\"/>"
        "<col class=\"year\"/><col class=\"year\"/>"
        "</colgroup>"
    )
    table_html = (
        f"{css}<table class=\"cashflow-table\">{colgroup}"
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody></table>"
    )
    st.markdown(table_html, unsafe_allow_html=True)


def _render_balance_sheet_html(balance_statement, section_rows, bold_rows):
    def escape(text):
        return (
            str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    columns = ["Line Item", "Year 0", "Year 1", "Year 2", "Year 3", "Year 4"]
    header_cells = "".join(f"<th>{escape(col)}</th>" for col in columns)

    body_rows = []
    for _, row in balance_statement.iterrows():
        label = row["Line Item"]
        row_class = ""
        if label in section_rows:
            row_class = "section-row"
        elif label in bold_rows:
            row_class = "total-row"
        cells = []
        for col in columns:
            value = row[col]
            cell_class = ""
            if col != "Line Item":
                value = format_currency(value)
                try:
                    if float(row[col]) < 0:
                        cell_class = "negative"
                except (TypeError, ValueError):
                    cell_class = ""
            cell_value = "&nbsp;" if value in ("", None) else escape(value)
            cells.append(f"<td class=\"{cell_class}\">{cell_value}</td>")
        body_rows.append(f"<tr class=\"{row_class}\">{''.join(cells)}</tr>")

    css = """
    <style>
      .balance-table { width: 100%; border-collapse: collapse; table-layout: fixed; }
      .balance-table col.line-item { width: 42%; }
      .balance-table col.year { width: 11.6%; }
      .balance-table th, .balance-table td {
        padding: 2px 6px;
        white-space: nowrap;
        line-height: 1.0;
        border: 0;
        font-size: 0.9rem;
      }
      .balance-table th { text-align: right; font-weight: 600; }
      .balance-table th:first-child { text-align: left; }
      .balance-table td { text-align: right; }
      .balance-table td:first-child { text-align: left; }
      .balance-table .section-row td {
        font-weight: 700;
        background: #f9fafb;
      }
      .balance-table .total-row td {
        font-weight: 700;
        background: #f3f4f6;
        border-top: 1px solid #c7c7c7;
      }
      .balance-table td.negative { color: #b45309; }
    </style>
    """
    colgroup = (
        "<colgroup>"
        "<col class=\"line-item\"/>"
        "<col class=\"year\"/><col class=\"year\"/><col class=\"year\"/>"
        "<col class=\"year\"/><col class=\"year\"/>"
        "</colgroup>"
    )
    table_html = (
        f"{css}<table class=\"balance-table\">{colgroup}"
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody></table>"
    )
    st.markdown(table_html, unsafe_allow_html=True)


def _render_custom_table_html(
    statement,
    section_rows,
    bold_rows,
    row_formatters=None,
    year_labels=None,
):
    def escape(text):
        return (
            str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    if year_labels is None:
        year_labels = ["Year 0", "Year 1", "Year 2", "Year 3", "Year 4"]
    columns = ["Line Item"] + year_labels
    header_cells = "".join(f"<th>{escape(col)}</th>" for col in columns)
    row_formatters = row_formatters or {}

    body_rows = []
    for _, row in statement.iterrows():
        label = row["Line Item"]
        row_class = ""
        if label in section_rows:
            row_class = "section-row"
        elif label in bold_rows:
            row_class = "total-row"
        cells = []
        for col in columns:
            value = row[col]
            cell_class = ""
            if col != "Line Item":
                formatter = row_formatters.get(label, format_currency)
                if value is None or value == "" or pd.isna(value):
                    value = ""
                else:
                    value = formatter(value)
                try:
                    if float(row[col]) < 0:
                        cell_class = "negative"
                except (TypeError, ValueError):
                    if label == "Covenant Breach" and value == "YES":
                        cell_class = "negative"
                    else:
                        cell_class = ""
            cell_value = "&nbsp;" if value in ("", None) else escape(value)
            cells.append(f"<td class=\"{cell_class}\">{cell_value}</td>")
        body_rows.append(f"<tr class=\"{row_class}\">{''.join(cells)}</tr>")

    line_item_width = 35
    year_width = (100 - line_item_width) / max(len(year_labels), 1)
    css = f"""
    <style>
      .custom-table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
      .custom-table col.line-item {{ width: {line_item_width}%; }}
      .custom-table col.year {{ width: {year_width:.2f}%; }}
      .custom-table th, .custom-table td {{
        padding: 2px 6px;
        white-space: nowrap;
        line-height: 1.0;
        border: 0;
        font-size: 0.9rem;
      }}
      .custom-table th {{ text-align: right; font-weight: 600; }}
      .custom-table th:first-child {{ text-align: left; }}
      .custom-table td {{ text-align: right; }}
      .custom-table td:first-child {{ text-align: left; }}
      .custom-table .section-row td {{
        font-weight: 700;
        background: #f9fafb;
      }}
      .custom-table .total-row td {{
        font-weight: 700;
        background: #f3f4f6;
        border-top: 1px solid #c7c7c7;
      }}
      .custom-table td.negative {{ color: #b45309; }}
    </style>
    """
    colgroup = "<colgroup><col class=\"line-item\"/>"
    colgroup += "".join("<col class=\"year\"/>" for _ in year_labels)
    colgroup += "</colgroup>"
    table_html = (
        f"{css}<table class=\"custom-table\">{colgroup}"
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody></table>"
    )
    st.markdown(table_html, unsafe_allow_html=True)


def _build_pnl_excel(input_model, pnl_result, cashflow_result, debt_schedule):
    scenario = input_model.scenario_selection["selected_scenario"].value
    scenario_key = scenario.lower()
    cashflow_assumptions = getattr(
        input_model, "cashflow_assumptions", _default_cashflow_assumptions()
    )
    balance_assumptions = getattr(
        input_model,
        "balance_sheet_assumptions",
        _default_balance_sheet_assumptions(input_model),
    )
    valuation_runtime = getattr(
        input_model,
        "valuation_runtime",
        _default_valuation_assumptions(input_model),
    )
    financing_assumptions = getattr(
        input_model,
        "financing_assumptions",
        _default_financing_assumptions(input_model),
    )
    pnl_list = _pnl_dict_to_list(pnl_result)
    equity_assumptions = getattr(
        input_model,
        "equity_assumptions",
        _default_equity_assumptions(input_model),
    )

    assumptions = [
        ("Consulting FTE", input_model.operating_assumptions["consulting_fte_start"].value),
        ("FTE Growth %", input_model.operating_assumptions["consulting_fte_growth_pct"].value),
        ("Workdays per Year", input_model.operating_assumptions["work_days_per_year"].value),
        ("Utilization %", input_model.scenario_parameters["utilization_rate"][scenario_key].value),
        ("Day Rate (EUR)", input_model.scenario_parameters["day_rate_eur"][scenario_key].value),
        ("Day Rate Growth %", input_model.operating_assumptions["day_rate_growth_pct"].value),
        ("Guarantee % Year 1", input_model.operating_assumptions["revenue_guarantee_pct_year_1"].value),
        ("Guarantee % Year 2", input_model.operating_assumptions["revenue_guarantee_pct_year_2"].value),
        ("Guarantee % Year 3", input_model.operating_assumptions["revenue_guarantee_pct_year_3"].value),
        ("Consultant Base Cost (EUR)", input_model.personnel_cost_assumptions["avg_consultant_base_cost_eur_per_year"].value),
        ("Bonus %", input_model.personnel_cost_assumptions["bonus_pct_of_base"].value),
        ("Payroll Burden %", input_model.personnel_cost_assumptions["payroll_burden_pct_of_comp"].value),
        ("Wage Inflation %", input_model.personnel_cost_assumptions["wage_inflation_pct"].value),
        ("Backoffice FTE", input_model.operating_assumptions["backoffice_fte_start"].value),
        ("Backoffice Growth %", input_model.operating_assumptions["backoffice_fte_growth_pct"].value),
        ("Backoffice Salary (EUR)", input_model.operating_assumptions["avg_backoffice_salary_eur_per_year"].value),
        ("Overhead Inflation %", input_model.overhead_and_variable_costs["overhead_inflation_pct"].value),
        ("External Advisors (EUR)", input_model.overhead_and_variable_costs["legal_audit_eur_per_year"].value),
        ("IT (EUR)", input_model.overhead_and_variable_costs["it_and_software_eur_per_year"].value),
        ("Office (EUR)", input_model.overhead_and_variable_costs["rent_eur_per_year"].value),
        ("Insurance (EUR)", input_model.overhead_and_variable_costs["insurance_eur_per_year"].value),
        ("Other Services (EUR)", input_model.overhead_and_variable_costs["other_overhead_eur_per_year"].value),
        ("Depreciation (EUR)", input_model.capex_and_working_capital["depreciation_eur_per_year"].value),
        ("Purchase Price (EUR)", input_model.transaction_and_financing["purchase_price_eur"].value),
        ("Equity Contribution (EUR)", input_model.transaction_and_financing["equity_contribution_eur"].value),
        ("Debt Amount (EUR)", input_model.transaction_and_financing["senior_term_loan_start_eur"].value),
        ("Interest Rate %", input_model.transaction_and_financing["senior_interest_rate_pct"].value),
        ("Tax Rate %", input_model.tax_and_distributions["tax_rate_pct"].value),
        ("Tax Cash Rate (%)", cashflow_assumptions["tax_cash_rate_pct"]),
        ("Tax Payment Lag (Years)", cashflow_assumptions["tax_payment_lag_years"]),
        ("Capex (% of Revenue)", cashflow_assumptions["capex_pct_revenue"]),
        ("Maintenance Capex (% of Revenue)", financing_assumptions["maintenance_capex_pct_revenue"]),
        ("Minimum DSCR", financing_assumptions["minimum_dscr"]),
        ("Amortisation Period (Years)", financing_assumptions["amortization_period_years"]),
        ("Working Capital Adjustment (% of Revenue)", cashflow_assumptions["working_capital_pct_revenue"]),
        ("Opening Cash Balance (EUR)", cashflow_assumptions["opening_cash_balance_eur"]),
        ("Sponsor Equity (EUR)", equity_assumptions["sponsor_equity_eur"]),
        ("Investor Equity (EUR)", equity_assumptions["investor_equity_eur"]),
        ("Exit Year", equity_assumptions["exit_year"]),
        ("Exit Method", equity_assumptions["exit_method"]),
        ("Exit Multiple (x)", equity_assumptions["exit_multiple"]),
    ]
    assumptions_df = pd.DataFrame(assumptions, columns=["Item", "Value"])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        assumptions_df.to_excel(writer, sheet_name="Assumptions", index=False)
        wb = writer.book
        ws_pnl = wb.create_sheet("P&L")
        ws_kpi = wb.create_sheet("KPIs")
        ws_cashflow = wb.create_sheet("Cashflow")
        ws_balance = wb.create_sheet("Balance Sheet")
        ws_valuation = wb.create_sheet("Valuation")
        ws_financing = wb.create_sheet("Financing & Debt")
        ws_equity = wb.create_sheet("Equity Case")

        assumption_cells = {
            name: f"Assumptions!B{idx + 2}" for idx, (name, _) in enumerate(assumptions)
        }

        def year_col(col_index):
            return get_column_letter(col_index)

        year_headers = ["Year 0", "Year 1", "Year 2", "Year 3", "Year 4"]
        ws_pnl["A1"] = "Line Item"
        for idx, header in enumerate(year_headers, start=2):
            ws_pnl.cell(row=1, column=idx, value=header)

        line_items = [
            "Revenue",
            "Total Revenue",
            "Personnel Costs",
            "Consultant Compensation",
            "Backoffice Compensation",
            "Management / MD Compensation",
            "Total Personnel Costs",
            "Operating Expenses",
            "External Consulting / Advisors",
            "IT",
            "Office",
            "Other Services",
            "Total Operating Expenses",
            "EBITDA",
            "Depreciation",
            "EBIT",
            "Interest Expense",
            "EBT",
            "Taxes",
            "Net Income (Jahresüberschuss)",
        ]

        for row_idx, item in enumerate(line_items, start=2):
            ws_pnl.cell(row=row_idx, column=1, value=item)

        reference_volume_cell = "20000000"
        for year_index in range(5):
            col = year_col(2 + year_index)
            fte = f"({assumption_cells['Consulting FTE']}*(1+{assumption_cells['FTE Growth %']})^{year_index})"
            workdays = assumption_cells["Workdays per Year"]
            utilization = assumption_cells["Utilization %"]
            day_rate = f"({assumption_cells['Day Rate (EUR)']}*(1+{assumption_cells['Day Rate Growth %']})^{year_index})"
            guarantee_pct = "0"
            if year_index == 0:
                guarantee_pct = assumption_cells["Guarantee % Year 1"]
            elif year_index == 1:
                guarantee_pct = assumption_cells["Guarantee % Year 2"]
            elif year_index == 2:
                guarantee_pct = assumption_cells["Guarantee % Year 3"]

            total_revenue = f"={fte}*{workdays}*{utilization}*{day_rate}"
            consultant_cost = (
                f"={fte}*{assumption_cells['Consultant Base Cost (EUR)']}*"
                f"(1+{assumption_cells['Bonus %']}+{assumption_cells['Payroll Burden %']})*"
                f"(1+{assumption_cells['Wage Inflation %']})^{year_index}"
            )
            backoffice_fte = f"({assumption_cells['Backoffice FTE']}*(1+{assumption_cells['Backoffice Growth %']})^{year_index})"
            backoffice_cost = (
                f"={backoffice_fte}*{assumption_cells['Backoffice Salary (EUR)']}*"
                f"(1+{assumption_cells['Payroll Burden %']})*"
                f"(1+{assumption_cells['Wage Inflation %']})^{year_index}"
            )
            management_cost = "=0"
            total_personnel = f"={col}5+{col}6+{col}7"

            external_advisors = f"={assumption_cells['External Advisors (EUR)']}*(1+{assumption_cells['Overhead Inflation %']})^{year_index}"
            it_cost = f"={assumption_cells['IT (EUR)']}*(1+{assumption_cells['Overhead Inflation %']})^{year_index}"
            office_cost = f"={assumption_cells['Office (EUR)']}*(1+{assumption_cells['Overhead Inflation %']})^{year_index}"
            other_services = f"=({assumption_cells['Insurance (EUR)']}+{assumption_cells['Other Services (EUR)']})*(1+{assumption_cells['Overhead Inflation %']})^{year_index}"
            total_opex = f"={col}10+{col}11+{col}12+{col}13"

            ebitda = f"={col}3-{col}8-{col}14"
            depreciation = f"={assumption_cells['Depreciation (EUR)']}"
            ebit = f"={col}15-{col}16"
            interest = f"={assumption_cells['Debt Amount (EUR)']}*{assumption_cells['Interest Rate %']}"
            ebt = f"={col}17-{col}18"
            taxes = f"=MAX({col}19,0)*{assumption_cells['Tax Rate %']}"
            net_income = f"={col}19-{col}20"

            ws_pnl[f"{col}3"] = total_revenue
            ws_pnl[f"{col}5"] = consultant_cost
            ws_pnl[f"{col}6"] = backoffice_cost
            ws_pnl[f"{col}7"] = management_cost
            ws_pnl[f"{col}8"] = total_personnel
            ws_pnl[f"{col}10"] = external_advisors
            ws_pnl[f"{col}11"] = it_cost
            ws_pnl[f"{col}12"] = office_cost
            ws_pnl[f"{col}13"] = other_services
            ws_pnl[f"{col}14"] = total_opex
            ws_pnl[f"{col}15"] = ebitda
            ws_pnl[f"{col}16"] = depreciation
            ws_pnl[f"{col}17"] = ebit
            ws_pnl[f"{col}18"] = interest
            ws_pnl[f"{col}19"] = ebt
            ws_pnl[f"{col}20"] = taxes
            ws_pnl[f"{col}21"] = net_income

        ws_kpi["A1"] = "KPI"
        for idx, header in enumerate(year_headers, start=2):
            ws_kpi.cell(row=1, column=idx, value=header)

        kpis = [
            "Revenue per Consultant",
            "EBITDA Margin",
            "EBIT Margin",
            "Personnel Cost Ratio",
            "Opex Ratio",
            "Net Margin",
        ]
        for row_idx, kpi in enumerate(kpis, start=2):
            ws_kpi.cell(row=row_idx, column=1, value=kpi)

        for year_index in range(5):
            col = year_col(2 + year_index)
            fte = f"({assumption_cells['Consulting FTE']}*(1+{assumption_cells['FTE Growth %']})^{year_index})"
            ws_kpi[f"{col}2"] = f"='P&L'!{col}3/{fte}"
            ws_kpi[f"{col}3"] = f"='P&L'!{col}15/'P&L'!{col}3"
            ws_kpi[f"{col}4"] = f"='P&L'!{col}17/'P&L'!{col}3"
            ws_kpi[f"{col}5"] = f"='P&L'!{col}8/'P&L'!{col}3"
            ws_kpi[f"{col}6"] = f"='P&L'!{col}14/'P&L'!{col}3"
            ws_kpi[f"{col}7"] = f"='P&L'!{col}21/'P&L'!{col}3"

        ws_cashflow["A1"] = "Cashflow Assumptions"
        cashflow_assumption_rows = [
            ("Tax Cash Rate (%)", "Tax Cash Rate (%)"),
            ("Tax Payment Lag (Years)", "Tax Payment Lag (Years)"),
            ("Capex (% of Revenue)", "Capex (% of Revenue)"),
            ("Working Capital Adjustment (% of Revenue)", "Working Capital Adjustment (% of Revenue)"),
            ("Opening Cash Balance (EUR)", "Opening Cash Balance (EUR)"),
        ]
        for idx, (label, key) in enumerate(cashflow_assumption_rows, start=2):
            ws_cashflow.cell(row=idx, column=1, value=label)
            ws_cashflow.cell(
                row=idx, column=2, value=f"={assumption_cells[key]}"
            )

        cashflow_table_start = 8
        ws_cashflow.cell(row=cashflow_table_start, column=1, value="Line Item")
        for idx, header in enumerate(year_headers, start=2):
            ws_cashflow.cell(row=cashflow_table_start, column=idx, value=header)

        cashflow_items = [
            "OPERATING CASHFLOW",
            "EBITDA",
            "Taxes Paid",
            "Working Capital Change",
            "Operating Cashflow",
            "INVESTING CASHFLOW",
            "Capex",
            "Free Cashflow",
            "FINANCING CASHFLOW",
            "Debt Drawdown",
            "Interest Paid",
            "Debt Repayment",
            "Net Cashflow",
            "LIQUIDITY",
            "Opening Cash",
            "Net Cashflow",
            "Closing Cash",
        ]
        for row_idx, item in enumerate(
            cashflow_items, start=cashflow_table_start + 1
        ):
            ws_cashflow.cell(row=row_idx, column=1, value=item)

        tax_cash_cell = assumption_cells["Tax Cash Rate (%)"]
        tax_lag_cell = assumption_cells["Tax Payment Lag (Years)"]
        capex_pct_cell = assumption_cells["Capex (% of Revenue)"]
        wc_pct_cell = assumption_cells["Working Capital Adjustment (% of Revenue)"]
        opening_cash_cell = assumption_cells["Opening Cash Balance (EUR)"]
        debt_amount_cell = assumption_cells["Debt Amount (EUR)"]
        interest_rate_cell = assumption_cells["Interest Rate %"]
        amort_period_cell = assumption_cells["Amortisation Period (Years)"]

        cashflow_row_map = {
            "EBITDA": cashflow_table_start + 2,
            "Taxes Paid": cashflow_table_start + 3,
            "Working Capital Change": cashflow_table_start + 4,
            "Operating Cashflow": cashflow_table_start + 5,
            "Capex": cashflow_table_start + 7,
            "Free Cashflow": cashflow_table_start + 8,
            "Debt Drawdown": cashflow_table_start + 10,
            "Interest Paid": cashflow_table_start + 11,
            "Debt Repayment": cashflow_table_start + 12,
            "Net Cashflow (Financing)": cashflow_table_start + 13,
            "Opening Cash": cashflow_table_start + 15,
            "Net Cashflow (Liquidity)": cashflow_table_start + 16,
            "Closing Cash": cashflow_table_start + 17,
        }

        for year_index in range(5):
            col = year_col(2 + year_index)
            prev_col = year_col(1 + year_index) if year_index > 0 else None
            ws_cashflow[f"{col}{cashflow_row_map['EBITDA']}"] = f"='P&L'!{col}17"

            taxes_due = f"MAX('P&L'!{col}21,0)*{tax_cash_cell}"
            if year_index == 0:
                taxes_prev = "0"
            else:
                taxes_prev = f"MAX('P&L'!{prev_col}21,0)*{tax_cash_cell}"
            ws_cashflow[f"{col}{cashflow_row_map['Taxes Paid']}"] = (
                f"=IF({tax_lag_cell}=0,{taxes_due},IF({tax_lag_cell}=1,{taxes_prev},0))"
            )

            ws_cashflow[f"{col}{cashflow_row_map['Working Capital Change']}"] = (
                f"='P&L'!{col}5*{wc_pct_cell}"
            )
            ws_cashflow[f"{col}{cashflow_row_map['Operating Cashflow']}"] = (
                f"={col}{cashflow_row_map['EBITDA']}"
                f"-{col}{cashflow_row_map['Taxes Paid']}"
                f"-{col}{cashflow_row_map['Working Capital Change']}"
            )
            ws_cashflow[f"{col}{cashflow_row_map['Capex']}"] = (
                f"='P&L'!{col}5*{capex_pct_cell}"
            )
            ws_cashflow[f"{col}{cashflow_row_map['Free Cashflow']}"] = (
                f"={col}{cashflow_row_map['Operating Cashflow']}"
                f"-{col}{cashflow_row_map['Capex']}"
            )

            debt_row = debt_schedule[year_index]
            ws_cashflow[f"{col}{cashflow_row_map['Debt Drawdown']}"] = (
                debt_row.get("debt_drawdown", 0.0)
            )
            ws_cashflow[f"{col}{cashflow_row_map['Interest Paid']}"] = (
                debt_row.get("interest_expense", 0.0)
            )
            ws_cashflow[f"{col}{cashflow_row_map['Debt Repayment']}"] = (
                debt_row.get("total_repayment", 0.0)
            )
            net_cashflow_formula = (
                f"={col}{cashflow_row_map['Free Cashflow']}"
                f"+{col}{cashflow_row_map['Debt Drawdown']}"
                f"-{col}{cashflow_row_map['Interest Paid']}"
                f"-{col}{cashflow_row_map['Debt Repayment']}"
            )
            ws_cashflow[
                f"{col}{cashflow_row_map['Net Cashflow (Financing)']}"
            ] = net_cashflow_formula
            ws_cashflow[
                f"{col}{cashflow_row_map['Net Cashflow (Liquidity)']}"
            ] = net_cashflow_formula

            if year_index == 0:
                ws_cashflow[f"{col}{cashflow_row_map['Opening Cash']}"] = (
                    f"={opening_cash_cell}"
                )
            else:
                ws_cashflow[f"{col}{cashflow_row_map['Opening Cash']}"] = (
                    f"={prev_col}{cashflow_row_map['Closing Cash']}"
                )
            ws_cashflow[f"{col}{cashflow_row_map['Closing Cash']}"] = (
                f"={col}{cashflow_row_map['Opening Cash']}"
                f"+{col}{cashflow_row_map['Net Cashflow (Liquidity)']}"
            )

        notes_row = cashflow_table_start + len(cashflow_items) + 2
        ws_cashflow.cell(row=notes_row, column=1, value="Notes")
        ws_cashflow.cell(
            row=notes_row + 1,
            column=1,
            value="Operating CF = EBITDA - Taxes Paid - Working Capital Change.",
        )
        ws_cashflow.cell(
            row=notes_row + 2,
            column=1,
            value="Capex and Working Capital are modeled as % of revenue.",
        )
        ws_cashflow.cell(
            row=notes_row + 3,
            column=1,
            value="Closing Cash = Opening Cash + Net Cashflow.",
        )

        ws_balance["A1"] = "Balance Sheet Assumptions"
        balance_assumption_rows = [
            ("Opening Equity (EUR)", balance_assumptions["opening_equity_eur"]),
            ("Depreciation Rate (%)", balance_assumptions["depreciation_rate_pct"]),
            ("Minimum Cash Balance (EUR)", balance_assumptions["minimum_cash_balance_eur"]),
            ("Opening Fixed Assets (EUR)", 0),
        ]
        for idx, (label, value) in enumerate(balance_assumption_rows, start=2):
            ws_balance.cell(row=idx, column=1, value=label)
            ws_balance.cell(row=idx, column=2, value=value)

        balance_table_start = 8
        ws_balance.cell(row=balance_table_start, column=1, value="Line Item")
        for idx, header in enumerate(year_headers, start=2):
            ws_balance.cell(row=balance_table_start, column=idx, value=header)

        balance_items = [
            "ASSETS",
            "Cash",
            "Fixed Assets (Net)",
            "Total Assets",
            "LIABILITIES",
            "Financial Debt",
            "Total Liabilities",
            "EQUITY",
            "Equity at Start of Year",
            "Net Income",
            "Dividends",
            "Equity Injections",
            "Equity Buybacks / Exit Payouts",
            "Equity at End of Year",
            "CHECK",
            "Total Assets",
            "Total Liabilities + Equity",
        ]
        for row_idx, item in enumerate(
            balance_items, start=balance_table_start + 1
        ):
            ws_balance.cell(row=row_idx, column=1, value=item)

        balance_row_map = {
            "Cash": balance_table_start + 2,
            "Fixed Assets (Net)": balance_table_start + 3,
            "Total Assets (Assets)": balance_table_start + 4,
            "Financial Debt": balance_table_start + 6,
            "Total Liabilities": balance_table_start + 7,
            "Equity at Start of Year": balance_table_start + 9,
            "Net Income": balance_table_start + 10,
            "Dividends": balance_table_start + 11,
            "Equity Injections": balance_table_start + 12,
            "Equity Buybacks / Exit Payouts": balance_table_start + 13,
            "Equity at End of Year": balance_table_start + 14,
            "Total Assets (Check)": balance_table_start + 16,
            "Total Liabilities + Equity": balance_table_start + 17,
        }

        opening_equity_cell = "B2"
        depreciation_rate_cell = "B3"
        opening_fixed_assets_cell = "B5"
        debt_amount_cell = assumption_cells["Debt Amount (EUR)"]
        amort_period_cell = assumption_cells["Amortisation Period (Years)"]

        for year_index in range(5):
            col = year_col(2 + year_index)
            prev_col = year_col(1 + year_index) if year_index > 0 else None
            ws_balance[f"{col}{balance_row_map['Cash']}"] = (
                f"=Cashflow!{col}{cashflow_row_map['Closing Cash']}"
            )

            if year_index == 0:
                ws_balance[f"{col}{balance_row_map['Fixed Assets (Net)']}"] = (
                    f"=MAX({opening_fixed_assets_cell}"
                    f"+Cashflow!{col}{cashflow_row_map['Capex']}"
                    f"-({opening_fixed_assets_cell}"
                    f"+Cashflow!{col}{cashflow_row_map['Capex']})*{depreciation_rate_cell},0)"
                )
                ws_balance[f"{col}{balance_row_map['Equity at Start of Year']}"] = (
                    f"={opening_equity_cell}"
                )
            else:
                ws_balance[f"{col}{balance_row_map['Fixed Assets (Net)']}"] = (
                    f"=MAX({prev_col}{balance_row_map['Fixed Assets (Net)']}"
                    f"+Cashflow!{col}{cashflow_row_map['Capex']}"
                    f"-({prev_col}{balance_row_map['Fixed Assets (Net)']}"
                    f"+Cashflow!{col}{cashflow_row_map['Capex']})*{depreciation_rate_cell},0)"
                )
                ws_balance[f"{col}{balance_row_map['Equity at Start of Year']}"] = (
                    f"={prev_col}{balance_row_map['Equity at End of Year']}"
                )

            ws_balance[f"{col}{balance_row_map['Total Assets (Assets)']}"] = (
                f"={col}{balance_row_map['Cash']}+{col}{balance_row_map['Fixed Assets (Net)']}"
            )
            ws_balance[f"{col}{balance_row_map['Financial Debt']}"] = (
                debt_schedule[year_index].get("closing_debt", 0.0)
            )
            ws_balance[f"{col}{balance_row_map['Total Liabilities']}"] = (
                f"={col}{balance_row_map['Financial Debt']}"
            )
            ws_balance[f"{col}{balance_row_map['Net Income']}"] = (
                f"='P&L'!{col}23"
            )
            ws_balance[f"{col}{balance_row_map['Dividends']}"] = "=0"
            ws_balance[f"{col}{balance_row_map['Equity Injections']}"] = (
                f"=IF({year_index}=0,{assumption_cells['Equity Contribution (EUR)']},0)"
            )
            ws_balance[f"{col}{balance_row_map['Equity Buybacks / Exit Payouts']}"] = "=0"
            ws_balance[f"{col}{balance_row_map['Equity at End of Year']}"] = (
                f"={col}{balance_row_map['Equity at Start of Year']}"
                f"+{col}{balance_row_map['Net Income']}"
                f"-{col}{balance_row_map['Dividends']}"
                f"+{col}{balance_row_map['Equity Injections']}"
                f"-{col}{balance_row_map['Equity Buybacks / Exit Payouts']}"
            )
            ws_balance[f"{col}{balance_row_map['Total Assets (Check)']}"] = (
                f"={col}{balance_row_map['Total Assets (Assets)']}"
            )
            ws_balance[f"{col}{balance_row_map['Total Liabilities + Equity']}"] = (
                f"={col}{balance_row_map['Total Liabilities']}"
                f"+{col}{balance_row_map['Equity at End of Year']}"
            )

        balance_notes_row = balance_table_start + len(balance_items) + 2
        ws_balance.cell(row=balance_notes_row, column=1, value="Notes")
        ws_balance.cell(
            row=balance_notes_row + 1,
            column=1,
            value="Fixed Assets end = prior assets + capex - depreciation.",
        )
        ws_balance.cell(
            row=balance_notes_row + 2,
            column=1,
            value="Equity end = equity start + net income - dividends + injections - buybacks.",
        )
        ws_balance.cell(
            row=balance_notes_row + 3,
            column=1,
            value="Total Assets should equal Total Liabilities + Equity.",
        )

        ws_valuation["A1"] = "Valuation Assumptions"
        valuation_assumption_rows = [
            ("Seller EBITDA Multiple (x)", valuation_runtime["seller_ebit_multiple"]),
            ("Reference Year (0-4)", valuation_runtime["reference_year"]),
            ("Discount Rate (WACC)", valuation_runtime["buyer_discount_rate"]),
            ("Valuation Start Year (0-4)", valuation_runtime["valuation_start_year"]),
            ("Transaction Costs (% of EV)", valuation_runtime["transaction_cost_pct"]),
            ("Include Terminal Value (1=On)", 1 if valuation_runtime["include_terminal_value"] else 0),
        ]
        for idx, (label, value) in enumerate(valuation_assumption_rows, start=2):
            ws_valuation.cell(row=idx, column=1, value=label)
            ws_valuation.cell(row=idx, column=2, value=value)

        seller_table_start = 10
        ws_valuation.cell(row=seller_table_start, column=1, value="Seller Valuation (Multiple-Based)")
        ws_valuation.cell(row=seller_table_start + 1, column=1, value="Line Item")
        for idx, header in enumerate(year_headers, start=2):
            ws_valuation.cell(row=seller_table_start + 1, column=idx, value=header)

        seller_items = [
            "Reference Year EBITDA",
            "Applied EBITDA Multiple",
            "Enterprise Value (EV)",
            "Net Debt at Close",
            "Equity Value (Seller View)",
        ]
        for row_idx, item in enumerate(seller_items, start=seller_table_start + 2):
            ws_valuation.cell(row=row_idx, column=1, value=item)

        seller_row_map = {
            "Reference Year EBITDA": seller_table_start + 2,
            "Applied EBITDA Multiple": seller_table_start + 3,
            "Enterprise Value (EV)": seller_table_start + 4,
            "Net Debt at Close": seller_table_start + 5,
            "Equity Value (Seller View)": seller_table_start + 6,
        }
        seller_multiple_cell = "B2"
        reference_year_cell = "B3"

        for year_index in range(5):
            col = year_col(2 + year_index)
            is_ref_year = f"={reference_year_cell}={year_index}"
            ebitda_cell = f"=INDEX('P&L'!B17:F17,1,{reference_year_cell}+1)"
            ws_valuation[f"{col}{seller_row_map['Reference Year EBITDA']}"] = (
                f"=IF({is_ref_year},{ebitda_cell},\"\")"
            )
            ws_valuation[f"{col}{seller_row_map['Applied EBITDA Multiple']}"] = (
                f"=IF({is_ref_year},{seller_multiple_cell},\"\")"
            )
            ws_valuation[f"{col}{seller_row_map['Enterprise Value (EV)']}"] = (
                f"=IF({is_ref_year},{ebitda_cell}*{seller_multiple_cell},\"\")"
            )
            ws_valuation[f"{col}{seller_row_map['Net Debt at Close']}"] = (
                f"=IF({year_index}=0,'Balance Sheet'!{col}14-'Balance Sheet'!{col}10,\"\")"
            )
            ws_valuation[f"{col}{seller_row_map['Equity Value (Seller View)']}"] = (
                f"=IF({is_ref_year},{col}{seller_row_map['Enterprise Value (EV)']}-'Balance Sheet'!B14+'Balance Sheet'!B10,\"\")"
            )

        buyer_table_start = seller_table_start + len(seller_items) + 4
        ws_valuation.cell(row=buyer_table_start, column=1, value="Buyer Valuation (DCF)")
        ws_valuation.cell(row=buyer_table_start + 1, column=1, value="Line Item")
        for idx, header in enumerate(year_headers, start=2):
            ws_valuation.cell(row=buyer_table_start + 1, column=idx, value=header)

        buyer_items = [
            "Free Cashflow",
            "Discount Factor",
            "Present Value of FCF",
            "Cumulative PV of FCF",
            "Terminal Value",
            "Enterprise Value (DCF)",
            "Net Debt at Close",
            "Transaction Costs",
            "Equity Value (Buyer View)",
        ]
        for row_idx, item in enumerate(buyer_items, start=buyer_table_start + 2):
            ws_valuation.cell(row=row_idx, column=1, value=item)

        buyer_row_map = {
            "Free Cashflow": buyer_table_start + 2,
            "Discount Factor": buyer_table_start + 3,
            "Present Value of FCF": buyer_table_start + 4,
            "Cumulative PV of FCF": buyer_table_start + 5,
            "Terminal Value": buyer_table_start + 6,
            "Enterprise Value (DCF)": buyer_table_start + 7,
            "Net Debt at Close": buyer_table_start + 8,
            "Transaction Costs": buyer_table_start + 9,
            "Equity Value (Buyer View)": buyer_table_start + 10,
        }

        discount_rate_cell = "B4"
        start_year_cell = "B5"
        debt_close_cell = "B6"
        tx_cost_cell = "B7"
        include_terminal_cell = "B8"

        for year_index in range(5):
            col = year_col(2 + year_index)
            prev_col = year_col(1 + year_index) if year_index > 0 else None
            ws_valuation[f"{col}{buyer_row_map['Free Cashflow']}"] = (
                f"=Cashflow!{col}{cashflow_row_map['Free Cashflow']}"
            )
            ws_valuation[f"{col}{buyer_row_map['Discount Factor']}"] = (
                f"=IF({year_index}>={start_year_cell},1/(1+{discount_rate_cell})^({year_index}-{start_year_cell}+1),0)"
            )
            ws_valuation[f"{col}{buyer_row_map['Present Value of FCF']}"] = (
                f"={col}{buyer_row_map['Free Cashflow']}*{col}{buyer_row_map['Discount Factor']}"
            )
            if prev_col:
                ws_valuation[f"{col}{buyer_row_map['Cumulative PV of FCF']}"] = (
                    f"={prev_col}{buyer_row_map['Cumulative PV of FCF']}+{col}{buyer_row_map['Present Value of FCF']}"
                )
            else:
                ws_valuation[f"{col}{buyer_row_map['Cumulative PV of FCF']}"] = (
                    f"={col}{buyer_row_map['Present Value of FCF']}"
                )
            if year_index == 4:
                ws_valuation[f"{col}{buyer_row_map['Terminal Value']}"] = (
                    f"=IF({include_terminal_cell}=1,{col}{buyer_row_map['Free Cashflow']}/{discount_rate_cell},\"\")"
                )
                ws_valuation[f"{col}{buyer_row_map['Enterprise Value (DCF)']}"] = (
                    f"={col}{buyer_row_map['Cumulative PV of FCF']}+IF({include_terminal_cell}=1,{col}{buyer_row_map['Terminal Value']}*{col}{buyer_row_map['Discount Factor']},0)"
                )
                ws_valuation[f"{col}{buyer_row_map['Net Debt at Close']}"] = (
                    f"={debt_close_cell}"
                )
                ws_valuation[f"{col}{buyer_row_map['Transaction Costs']}"] = (
                    f"={col}{buyer_row_map['Enterprise Value (DCF)']}*{tx_cost_cell}"
                )
                ws_valuation[f"{col}{buyer_row_map['Equity Value (Buyer View)']}"] = (
                    f"={col}{buyer_row_map['Enterprise Value (DCF)']}-{col}{buyer_row_map['Net Debt at Close']}-{col}{buyer_row_map['Transaction Costs']}"
                )

        bridge_start = buyer_table_start + len(buyer_items) + 4
        ws_valuation.cell(row=bridge_start, column=1, value="Purchase Price Bridge")
        ws_valuation.cell(row=bridge_start + 1, column=1, value="Line Item")
        ws_valuation.cell(row=bridge_start + 1, column=2, value="Year 0")

        bridge_items = [
            "Seller Equity Value",
            "Buyer Equity Value",
            "Valuation Gap (EUR)",
            "Valuation Gap (%)",
        ]
        for row_idx, item in enumerate(bridge_items, start=bridge_start + 2):
            ws_valuation.cell(row=row_idx, column=1, value=item)

        ws_valuation[f"B{bridge_start + 2}"] = (
            f"=INDEX(B{seller_row_map['Equity Value (Seller View)']}:F{seller_row_map['Equity Value (Seller View)']},1,{reference_year_cell}+1)"
        )
        ws_valuation[f"B{bridge_start + 3}"] = (
            f"=F{buyer_row_map['Equity Value (Buyer View)']}"
        )
        ws_valuation[f"B{bridge_start + 4}"] = (
            f"=B{bridge_start + 3}-B{bridge_start + 2}"
        )
        ws_valuation[f"B{bridge_start + 5}"] = (
            f"=IF(B{bridge_start + 2}=0,0,B{bridge_start + 4}/B{bridge_start + 2})"
        )

        valuation_notes_row = bridge_start + len(bridge_items) + 2
        ws_valuation.cell(row=valuation_notes_row, column=1, value="Notes")
        ws_valuation.cell(
            row=valuation_notes_row + 1,
            column=1,
            value="Seller EV = EBITDA (reference year) × multiple.",
        )
        ws_valuation.cell(
            row=valuation_notes_row + 2,
            column=1,
            value="DCF uses free cashflow discounted at the buyer rate.",
        )
        ws_valuation.cell(
            row=valuation_notes_row + 3,
            column=1,
            value="Equity value = EV - net debt - transaction costs.",
        )

        ws_financing["A1"] = "Financing Assumptions"
        financing_rows = [
            ("Initial Debt Amount (EUR)", f"={assumption_cells['Debt Amount (EUR)']}"),
            ("Interest Rate", f"={assumption_cells['Interest Rate %']}"),
            ("Amortisation Type", "Linear"),
            ("Amortisation Period (Years)", f"={assumption_cells['Amortisation Period (Years)']}"),
            ("Minimum DSCR", f"={assumption_cells['Minimum DSCR']}"),
            ("Maintenance Capex (% of Revenue)", f"={assumption_cells['Maintenance Capex (% of Revenue)']}"),
            ("Working Capital Change (% of Revenue)", "=0"),
            ("Tax Cash Rate (%)", f"={assumption_cells['Tax Cash Rate (%)']}"),
            ("Tax Payment Lag (Years)", f"={assumption_cells['Tax Payment Lag (Years)']}"),
        ]
        for idx, (label, value) in enumerate(financing_rows, start=2):
            ws_financing.cell(row=idx, column=1, value=label)
            ws_financing.cell(row=idx, column=2, value=value)

        debt_table_start = len(financing_rows) + 4
        ws_financing.cell(row=debt_table_start, column=1, value="Debt Schedule")
        ws_financing.cell(row=debt_table_start + 1, column=1, value="Line Item")
        for idx, header in enumerate(year_headers, start=2):
            ws_financing.cell(row=debt_table_start + 1, column=idx, value=header)

        debt_items = [
            "Opening Debt",
            "Debt Drawdown",
            "Scheduled Repayment",
            "Special Repayment",
            "Total Repayment",
            "Closing Debt",
            "Interest Expense",
        ]
        for row_idx, item in enumerate(debt_items, start=debt_table_start + 2):
            ws_financing.cell(row=row_idx, column=1, value=item)

        debt_row_map = {
            "Opening Debt": debt_table_start + 2,
            "Debt Drawdown": debt_table_start + 3,
            "Scheduled Repayment": debt_table_start + 4,
            "Special Repayment": debt_table_start + 5,
            "Total Repayment": debt_table_start + 6,
            "Closing Debt": debt_table_start + 7,
            "Interest Expense": debt_table_start + 8,
        }

        interest_rate_cell = "B3"
        amort_type_cell = "B4"
        amort_period_cell = "B5"
        initial_debt_cell = "B2"

        for year_index in range(5):
            col = year_col(2 + year_index)
            debt_row = debt_schedule[year_index]
            ws_financing[f"{col}{debt_row_map['Opening Debt']}"] = debt_row.get(
                "opening_debt", 0.0
            )
            ws_financing[f"{col}{debt_row_map['Debt Drawdown']}"] = debt_row.get(
                "debt_drawdown", 0.0
            )
            ws_financing[f"{col}{debt_row_map['Scheduled Repayment']}"] = debt_row.get(
                "scheduled_repayment", 0.0
            )
            ws_financing[f"{col}{debt_row_map['Special Repayment']}"] = debt_row.get(
                "special_repayment", 0.0
            )
            ws_financing[f"{col}{debt_row_map['Total Repayment']}"] = debt_row.get(
                "total_repayment", 0.0
            )
            ws_financing[f"{col}{debt_row_map['Closing Debt']}"] = debt_row.get(
                "closing_debt", 0.0
            )
            ws_financing[f"{col}{debt_row_map['Interest Expense']}"] = debt_row.get(
                "interest_expense", 0.0
            )

        bank_table_start = debt_table_start + len(debt_items) + 3
        ws_financing.cell(row=bank_table_start, column=1, value="Bank View")
        ws_financing.cell(row=bank_table_start + 1, column=1, value="Line Item")
        for idx, header in enumerate(year_headers, start=2):
            ws_financing.cell(row=bank_table_start + 1, column=idx, value=header)

        bank_items = [
            "EBITDA",
            "Cash Taxes",
            "Capex (Maintenance)",
            "CFADS",
            "Interest Expense",
            "Scheduled Repayment",
            "Debt Service",
            "DSCR",
            "Minimum Required DSCR",
            "Covenant Breach",
        ]
        for row_idx, item in enumerate(bank_items, start=bank_table_start + 2):
            ws_financing.cell(row=row_idx, column=1, value=item)

        bank_row_map = {
            "EBITDA": bank_table_start + 2,
            "Cash Taxes": bank_table_start + 3,
            "Capex (Maintenance)": bank_table_start + 4,
            "CFADS": bank_table_start + 5,
            "Interest Expense": bank_table_start + 6,
            "Scheduled Repayment": bank_table_start + 7,
            "Debt Service": bank_table_start + 8,
            "DSCR": bank_table_start + 9,
            "Minimum Required DSCR": bank_table_start + 10,
            "Covenant Breach": bank_table_start + 11,
        }

        min_dscr_cell = "B6"
        maintenance_capex_cell = "B7"
        wc_change_cell = "B8"

        for year_index in range(5):
            col = year_col(2 + year_index)
            debt_row = debt_schedule[year_index]
            cashflow_row = cashflow_result[year_index]
            pnl_row = pnl_list[year_index]
            revenue = pnl_row.get("revenue", 0.0)
            maintenance_capex = (
                revenue * financing_assumptions["maintenance_capex_pct_revenue"]
            )
            cfads = (
                pnl_row.get("ebitda", 0.0)
                - cashflow_row.get("taxes_paid", 0.0)
                - maintenance_capex
                + cashflow_row.get("working_capital_change", 0.0)
            )
            debt_service = debt_row.get("interest_expense", 0.0) + debt_row.get(
                "scheduled_repayment", 0.0
            )
            dscr_value = cfads / debt_service if debt_service else 0.0
            ws_financing[f"{col}{bank_row_map['EBITDA']}"] = pnl_row.get(
                "ebitda", 0.0
            )
            ws_financing[f"{col}{bank_row_map['Cash Taxes']}"] = cashflow_row.get(
                "taxes_paid", 0.0
            )
            ws_financing[f"{col}{bank_row_map['Capex (Maintenance)']}"] = (
                maintenance_capex
            )
            ws_financing[f"{col}{bank_row_map['CFADS']}"] = cfads
            ws_financing[f"{col}{bank_row_map['Interest Expense']}"] = debt_row.get(
                "interest_expense", 0.0
            )
            ws_financing[f"{col}{bank_row_map['Scheduled Repayment']}"] = debt_row.get(
                "scheduled_repayment", 0.0
            )
            ws_financing[f"{col}{bank_row_map['Debt Service']}"] = debt_service
            ws_financing[f"{col}{bank_row_map['DSCR']}"] = dscr_value
            ws_financing[f"{col}{bank_row_map['Minimum Required DSCR']}"] = (
                financing_assumptions["minimum_dscr"]
            )
            ws_financing[f"{col}{bank_row_map['Covenant Breach']}"] = (
                "YES" if dscr_value < financing_assumptions["minimum_dscr"] else "NO"
            )

        financing_notes_row = bank_table_start + len(bank_items) + 2
        ws_financing.cell(row=financing_notes_row, column=1, value="Notes")
        ws_financing.cell(
            row=financing_notes_row + 1,
            column=1,
            value="CFADS = EBITDA - Cash Taxes - Maintenance Capex ± Working Capital Change.",
        )
        ws_financing.cell(
            row=financing_notes_row + 2,
            column=1,
            value="Debt Service = Interest Expense + Scheduled Repayment.",
        )
        ws_financing.cell(
            row=financing_notes_row + 3,
            column=1,
            value="DSCR = CFADS / Debt Service.",
        )

        ws_equity["A1"] = "Equity Assumptions"
        equity_rows = [
            ("Sponsor Equity (EUR)", "Sponsor Equity (EUR)"),
            ("Investor Equity (EUR)", "Investor Equity (EUR)"),
            ("Exit Year", "Exit Year"),
            ("Exit Method", "Exit Method"),
            ("Exit Multiple (x)", "Exit Multiple (x)"),
        ]
        for idx, (label, key) in enumerate(equity_rows, start=2):
            ws_equity.cell(row=idx, column=1, value=label)
            ws_equity.cell(
                row=idx, column=2, value=f"={assumption_cells[key]}"
            )

        equity_calc_start = len(equity_rows) + 4
        ws_equity.cell(row=equity_calc_start, column=1, value="Equity Summary")
        ws_equity.cell(row=equity_calc_start + 1, column=1, value="Metric")
        ws_equity.cell(row=equity_calc_start + 1, column=2, value="Value")

        sponsor_cell = "B2"
        investor_cell = "B3"
        exit_year_cell = "B4"
        exit_method_cell = "B5"
        exit_multiple_cell = "B6"

        ws_equity.cell(row=equity_calc_start + 2, column=1, value="Total Equity")
        ws_equity.cell(
            row=equity_calc_start + 2, column=2, value=f"={sponsor_cell}+{investor_cell}"
        )
        ws_equity.cell(row=equity_calc_start + 3, column=1, value="Sponsor %")
        ws_equity.cell(
            row=equity_calc_start + 3,
            column=2,
            value=f"=IF(B{equity_calc_start + 2}=0,0,{sponsor_cell}/B{equity_calc_start + 2})",
        )
        ws_equity.cell(row=equity_calc_start + 4, column=1, value="Investor %")
        ws_equity.cell(
            row=equity_calc_start + 4,
            column=2,
            value=f"=IF(B{equity_calc_start + 2}=0,0,{investor_cell}/B{equity_calc_start + 2})",
        )
        ws_equity.cell(row=equity_calc_start + 5, column=1, value="Exit EBITDA")
        ws_equity.cell(
            row=equity_calc_start + 5,
            column=2,
            value=f"=INDEX('P&L'!B17:F17,1,MIN({exit_year_cell},4)+1)",
        )
        ws_equity.cell(row=equity_calc_start + 6, column=1, value="Net Debt at Exit")
        ws_equity.cell(
            row=equity_calc_start + 6,
            column=2,
            value=f"=INDEX('Balance Sheet'!B14:F14,1,MIN({exit_year_cell},4)+1)-INDEX('Balance Sheet'!B10:F10,1,MIN({exit_year_cell},4)+1)",
        )
        ws_equity.cell(row=equity_calc_start + 7, column=1, value="Enterprise Value Exit")
        ws_equity.cell(
            row=equity_calc_start + 7,
            column=2,
            value=f"={exit_multiple_cell}*B{equity_calc_start + 5}",
        )
        ws_equity.cell(row=equity_calc_start + 8, column=1, value="Equity Value Exit")
        ws_equity.cell(
            row=equity_calc_start + 8,
            column=2,
            value=f"=B{equity_calc_start + 7}-B{equity_calc_start + 6}",
        )

        cashflow_start = equity_calc_start + 11
        ws_equity.cell(row=cashflow_start, column=1, value="Equity Cashflows")
        ws_equity.cell(row=cashflow_start + 1, column=1, value="Line Item")
        equity_year_headers = [
            "Year 0",
            "Year 1",
            "Year 2",
            "Year 3",
            "Year 4",
            "Year 5",
            "Year 6",
            "Year 7",
        ]
        for idx, header in enumerate(equity_year_headers, start=2):
            ws_equity.cell(row=cashflow_start + 1, column=idx, value=header)

        ws_equity.cell(row=cashflow_start + 2, column=1, value="Sponsor Cashflow")
        ws_equity.cell(row=cashflow_start + 3, column=1, value="Investor Cashflow")
        ws_equity.cell(row=cashflow_start + 4, column=1, value="Sponsor Residual Equity Value")

        sponsor_pct_cell = f"B{equity_calc_start + 3}"
        investor_pct_cell = f"B{equity_calc_start + 4}"
        equity_value_exit_cell = f"B{equity_calc_start + 8}"

        for year_index in range(8):
            col = year_col(2 + year_index)
            ws_equity[f"{col}{cashflow_start + 2}"] = (
                f"=IF({year_index}=0,-{sponsor_cell},"
                f"IF({year_index}={exit_year_cell},{equity_value_exit_cell},0))"
            )
            ws_equity[f"{col}{cashflow_start + 3}"] = (
                f"=IF({year_index}=0,-{investor_cell},"
                f"IF({year_index}={exit_year_cell},{equity_value_exit_cell}*{investor_pct_cell},0))"
            )
            ws_equity[f"{col}{cashflow_start + 4}"] = (
                f"=IF({year_index}={exit_year_cell},{equity_value_exit_cell},0)"
            )

        kpi_start = cashflow_start + 6
        ws_equity.cell(row=kpi_start, column=1, value="Equity KPIs")
        ws_equity.cell(row=kpi_start + 1, column=1, value="Investor")
        ws_equity.cell(row=kpi_start + 1, column=2, value="Invested Equity")
        ws_equity.cell(row=kpi_start + 1, column=3, value="Exit Proceeds")
        ws_equity.cell(row=kpi_start + 1, column=4, value="MOIC")
        ws_equity.cell(row=kpi_start + 1, column=5, value="IRR")

        ws_equity.cell(row=kpi_start + 2, column=1, value="Sponsor")
        ws_equity.cell(row=kpi_start + 3, column=1, value="Investor")

        ws_equity.cell(row=kpi_start + 2, column=2, value=f"={sponsor_cell}")
        ws_equity.cell(row=kpi_start + 3, column=2, value=f"={investor_cell}")
        ws_equity.cell(
            row=kpi_start + 2,
            column=3,
            value=f"={equity_value_exit_cell}",
        )
        ws_equity.cell(
            row=kpi_start + 3,
            column=3,
            value=f"={equity_value_exit_cell}*{investor_pct_cell}",
        )
        ws_equity.cell(
            row=kpi_start + 2,
            column=4,
            value="=\"—\"",
        )
        ws_equity.cell(
            row=kpi_start + 3,
            column=4,
            value=f"=IF({investor_cell}=0,0,C{kpi_start + 3}/{investor_cell})",
        )
        ws_equity.cell(
            row=kpi_start + 2,
            column=5,
            value=f"=IRR(B{cashflow_start + 2}:I{cashflow_start + 2})",
        )
        ws_equity.cell(
            row=kpi_start + 3,
            column=5,
            value=f"=IRR(B{cashflow_start + 3}:I{cashflow_start + 3})",
        )

        header_fill = PatternFill("solid", fgColor="E5E7EB")
        total_fill = PatternFill("solid", fgColor="F3F4F6")
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        total_font = Font(bold=True)
        center = Alignment(horizontal="center")
        left = Alignment(horizontal="left")

        def _apply_header(ws, row=1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center if col > 1 else left

        def _set_col_widths(ws, widths):
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

        def _format_currency_range(ws, row_start, row_end, col_start=2, col_end=6):
            for row in range(row_start, row_end + 1):
                for col in range(col_start, col_end + 1):
                    ws.cell(row=row, column=col).number_format = "#,##0"

        def _format_ratio_cells(ws, cells):
            for cell in cells:
                ws[cell].number_format = "0.00"

        def _bold_rows(ws, row_indexes):
            for row in row_indexes:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = total_font
                    cell.fill = total_fill

        # Assumptions sheet formatting.
        ws_assumptions = wb["Assumptions"]
        _apply_header(ws_assumptions, row=1)
        _set_col_widths(ws_assumptions, {"A": 46, "B": 20})
        for row in range(2, ws_assumptions.max_row + 1):
            label = ws_assumptions.cell(row=row, column=1).value or ""
            cell = ws_assumptions.cell(row=row, column=2)
            if "%" in str(label):
                cell.number_format = "0.0%"
            else:
                cell.number_format = "#,##0"

        # P&L formatting.
        _apply_header(ws_pnl, row=1)
        _set_col_widths(ws_pnl, {"A": 40, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
        _format_currency_range(ws_pnl, 2, 21)
        total_rows = {
            "Total Revenue",
            "Total Personnel Costs",
            "Total Operating Expenses",
            "EBITDA",
            "EBIT",
            "Net Income (Jahresüberschuss)",
        }
        total_row_indexes = [
            idx + 1
            for idx, item in enumerate(line_items, start=1)
            if item in total_rows
        ]
        _bold_rows(ws_pnl, total_row_indexes)

        # KPI formatting.
        _apply_header(ws_kpi, row=1)
        _set_col_widths(ws_kpi, {"A": 28, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
        for row in range(2, 8):
            for col in range(2, 7):
                ws_kpi.cell(row=row, column=col).number_format = "0.0%"
        for col in range(2, 7):
            ws_kpi.cell(row=2, column=col).number_format = "#,##0"

        # Cashflow formatting.
        _apply_header(ws_cashflow, row=cashflow_table_start)
        _set_col_widths(ws_cashflow, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
        _format_currency_range(ws_cashflow, cashflow_table_start + 1, cashflow_table_start + 17)
        _bold_rows(
            ws_cashflow,
            [
                cashflow_table_start + 5,
                cashflow_table_start + 8,
                cashflow_table_start + 13,
                cashflow_table_start + 17,
            ],
        )

        # Balance sheet formatting.
        _apply_header(ws_balance, row=balance_table_start)
        _set_col_widths(ws_balance, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
        _format_currency_range(ws_balance, balance_table_start + 1, balance_table_start + 17)
        _bold_rows(
            ws_balance,
            [
                balance_table_start + 4,
                balance_table_start + 7,
                balance_table_start + 14,
                balance_table_start + 17,
            ],
        )

        # Valuation formatting.
        _apply_header(ws_valuation, row=seller_table_start + 1)
        _apply_header(ws_valuation, row=buyer_table_start + 1)
        _apply_header(ws_valuation, row=bridge_start + 1)
        _set_col_widths(ws_valuation, {"A": 40, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
        _format_currency_range(ws_valuation, seller_table_start + 2, seller_table_start + 6)
        _format_currency_range(ws_valuation, buyer_table_start + 2, buyer_table_start + 10)
        _format_currency_range(ws_valuation, bridge_start + 2, bridge_start + 5, col_start=2, col_end=2)
        ws_valuation[f"B{bridge_start + 5}"].number_format = "0.0%"

        # Financing formatting.
        _apply_header(ws_financing, row=debt_table_start + 1)
        _apply_header(ws_financing, row=bank_table_start + 1)
        _set_col_widths(ws_financing, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
        _format_currency_range(ws_financing, debt_table_start + 2, debt_table_start + 8)
        _format_currency_range(ws_financing, bank_table_start + 2, bank_table_start + 8)
        _format_ratio_cells(ws_financing, [f"{year_col(2+i)}{bank_row_map['DSCR']}" for i in range(5)])

        # Equity formatting.
        _apply_header(ws_equity, row=cashflow_start + 1)
        _apply_header(ws_equity, row=kpi_start + 1)
        _set_col_widths(ws_equity, {"A": 34, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16})
        _format_currency_range(ws_equity, cashflow_start + 2, cashflow_start + 4, col_start=2, col_end=9)

        writer.close()

    output.seek(0)
    return output


def run_app(page_override=None):
    st.title("Financial Model")
    st.markdown(
        """
        <style>
          :root, html, body, [data-testid="stAppViewContainer"] {
            color: #111827 !important;
            background-color: #ffffff !important;
            color-scheme: light !important;
          }
          [data-testid="stSidebar"] {
            background-color: #f0f2f6 !important;
          }
          [data-testid="stSidebar"] * {
            color: #111827 !important;
          }
          [data-testid="stMarkdownContainer"] * {
            color: #111827 !important;
          }
          [data-testid="stTable"], [data-testid="stDataFrame"] {
            background: #ffffff !important;
          }
          .stDataFrame, .stTable, table, thead, tbody, tr, th, td {
            color: #111827 !important;
            background-color: #ffffff !important;
          }
          button, .stButton > button {
            background-color: #ffffff !important;
            color: #111827 !important;
            border-color: #d1d5db !important;
          }
          button:hover, .stButton > button:hover {
            background-color: #f3f4f6 !important;
          }
          input, textarea, select {
            background-color: #ffffff !important;
            color: #111827 !important;
            border-color: #d1d5db !important;
          }
          [data-testid="stDataEditor"], .stDataEditor, .stDataEditor [role="grid"] {
            background-color: #ffffff !important;
            color: #111827 !important;
          }
          [data-testid="stDataEditor"] .rdg, .stDataEditor .rdg {
            background-color: #ffffff !important;
            color: #111827 !important;
          }
          [data-testid="stDataEditor"] .rdg-cell, .stDataEditor .rdg-cell {
            background-color: #ffffff !important;
            color: #111827 !important;
          }
          [data-testid="stDataEditor"] .rdg-cell[aria-readonly="true"],
          .stDataEditor .rdg-cell[aria-readonly="true"] {
            background: #f3f4f6 !important;
            color: #6b7280 !important;
          }
          [data-testid="stDataEditor"] .rdg-header-row,
          [data-testid="stDataEditor"] .rdg-header-row .rdg-cell,
          .stDataEditor .rdg-header-row,
          .stDataEditor .rdg-header-row .rdg-cell {
            background-color: #f9fafb !important;
            color: #111827 !important;
          }
        </style>
        """,
        unsafe_allow_html=True,
    )

    base_model = create_demo_input_model()
    if not st.session_state.get("defaults_initialized"):
        _seed_session_defaults(base_model)
        st.session_state["defaults_initialized"] = True

    def _clamp_pct(value):
        if value is None or pd.isna(value):
            return 0.0
        return max(0.0, min(float(value), 1.0))

    def _non_negative(value):
        if value is None or pd.isna(value):
            return 0.0
        return max(0.0, float(value))

    def _seed_assumptions_state():
        scenario_labels = ["Base", "Best", "Worst"]
        revenue_model = {
            "reference_revenue_eur": {
                scenario: base_model.revenue_model["reference_revenue_eur"].value
                for scenario in scenario_labels
            },
            "guarantee_pct_by_year": {
                scenario: [
                    base_model.revenue_model[
                        f"guarantee_pct_year_{year_index}"
                    ].value
                    for year_index in range(5)
                ]
                for scenario in scenario_labels
            },
            "workdays_per_year": {
                scenario: [
                    value
                    for value in (
                        [205, 205, 205, 205, 205]
                        if scenario == "Worst"
                        else [215, 215, 215, 215, 215]
                        if scenario == "Best"
                        else [210, 210, 210, 210, 210]
                    )
                ]
                for scenario in scenario_labels
            },
            "utilization_rate": {
                scenario: [
                    value
                    for value in [0.68, 0.69, 0.70, 0.71, 0.71]
                ]
                for scenario in scenario_labels
            },
            "group_day_rate_eur": {
                scenario: [
                    value
                    for value in (
                        [2400, 2400, 2400, 2400, 2400]
                        if scenario == "Worst"
                        else [2600, 2600, 2600, 2600, 2600]
                        if scenario == "Best"
                        else [2500, 2500, 2500, 2500, 2500]
                    )
                ]
                for scenario in scenario_labels
            },
            "external_day_rate_eur": {
                scenario: [
                    value
                    for value in (
                        [2700, 2700, 2700, 2700, 2700]
                        if scenario == "Worst"
                        else [3300, 3300, 3300, 3300, 3300]
                        if scenario == "Best"
                        else [3000, 3000, 3000, 3000, 3000]
                    )
                ]
                for scenario in scenario_labels
            },
            "day_rate_growth_pct": {
                scenario: [
                    value
                    for value in [0.00, 0.01, 0.015, 0.02, 0.02]
                ]
                for scenario in scenario_labels
            },
            "revenue_growth_pct": {
                scenario: [0.0, 0.0, 0.0, 0.0, 0.0]
                for scenario in scenario_labels
            },
            "group_capacity_share_pct": {
                scenario: [0.80, 0.75, 0.70, 0.65, 0.60]
                for scenario in scenario_labels
            },
            "external_capacity_share_pct": {
                scenario: [0.20, 0.25, 0.30, 0.35, 0.40]
                for scenario in scenario_labels
            },
        }

        cost_personnel_rows = []
        cost_fixed_rows = []
        cost_variable_rows = []
        consultant_fte_base = [63, 61, 60, 60, 60]
        backoffice_fte_base = [18, 18, 17, 17, 17]
        office_rent_base = [1730000, 1400000, 1200000, 1200000, 1200000]
        it_cost_base = [440000, 448800, 448800, 448800, 448800]
        for year_index in range(5):
            year_label = f"Year {year_index}"
            cost_personnel_rows.append(
                {
                    "Year": year_label,
                    "Consultant FTE": consultant_fte_base[year_index],
                    "Consultant Loaded Cost (EUR)": base_model.cost_model[
                        f"consultant_base_cost_eur_year_{year_index}"
                    ].value,
                    "Backoffice FTE": backoffice_fte_base[year_index],
                    "Backoffice Loaded Cost (EUR)": base_model.cost_model[
                        f"backoffice_base_cost_eur_year_{year_index}"
                    ].value,
                    "Management Cost (EUR)": 1200000,
                }
            )
            cost_fixed_rows.append(
                {
                    "Year": year_label,
                    "Advisory": base_model.cost_model[
                        f"fixed_overhead_advisory_year_{year_index}"
                    ].value,
                    "Legal": base_model.cost_model[
                        f"fixed_overhead_legal_year_{year_index}"
                    ].value,
                    "IT & Software": it_cost_base[year_index],
                    "Office Rent": office_rent_base[year_index],
                    "Services": base_model.cost_model[
                        f"fixed_overhead_services_year_{year_index}"
                    ].value,
                    "Other Services": 0.0,
                }
            )
            cost_variable_rows.append(
                {
                    "Year": year_label,
                    "Training Type": "EUR",
                    "Training Value": consultant_fte_base[year_index] * 2500,
                    "Travel Type": "EUR",
                    "Travel Value": consultant_fte_base[year_index] * 5500,
                    "Communication Type": "EUR",
                    "Communication Value": 0.0,
                }
            )

        return {
            "revenue_model": revenue_model,
            "cost_model": {
                "personnel": cost_personnel_rows,
                "fixed_overhead": cost_fixed_rows,
                "variable_costs": cost_variable_rows,
                "inflation": {
                    "apply": False,
                    "rate_pct": base_model.personnel_cost_assumptions[
                        "wage_inflation_pct"
                    ].value,
                },
            },
            "personnel_costs": [
                {"Role": "Consultant Base Salary", "Cost Type": "Fixed", "Base Value (EUR)": base_model.personnel_cost_assumptions["avg_consultant_base_cost_eur_per_year"].value, "Growth (%)": base_model.personnel_cost_assumptions["wage_inflation_pct"].value, "Notes": "Base salary per consultant."},
                {"Role": "Consultant Variable (% Revenue)", "Cost Type": "Percent of Base", "Base Value (EUR)": base_model.personnel_cost_assumptions["bonus_pct_of_base"].value, "Growth (%)": "", "Notes": "Bonus as % of base salary."},
                {"Role": "Backoffice Cost per FTE", "Cost Type": "Fixed", "Base Value (EUR)": base_model.operating_assumptions["avg_backoffice_salary_eur_per_year"].value, "Growth (%)": base_model.personnel_cost_assumptions["wage_inflation_pct"].value, "Notes": "Average backoffice salary."},
                {"Role": "Management / MD Cost", "Cost Type": "Fixed", "Base Value (EUR)": 1200000, "Growth (%)": 0.02, "Notes": "Fixed management cost (base case)."},
            ],
            "opex": [
                {"Category": "External Consulting", "Cost Type": "Fixed", "Value": base_model.overhead_and_variable_costs["legal_audit_eur_per_year"].value, "Unit": "EUR", "Notes": "External advisors."},
                {"Category": "IT", "Cost Type": "Fixed", "Value": base_model.overhead_and_variable_costs["it_and_software_eur_per_year"].value, "Unit": "EUR", "Notes": "IT and software."},
                {"Category": "Office", "Cost Type": "Fixed", "Value": base_model.overhead_and_variable_costs["rent_eur_per_year"].value, "Unit": "EUR", "Notes": "Office rent."},
                {"Category": "Other Services", "Cost Type": "Fixed", "Value": base_model.overhead_and_variable_costs["other_overhead_eur_per_year"].value, "Unit": "EUR", "Notes": "Other services (excludes insurance)."},
            ],
            "financing": [
                {"Parameter": "Senior Debt Amount", "Value": base_model.transaction_and_financing["senior_term_loan_start_eur"].value, "Unit": "EUR", "Notes": "Opening senior term loan."},
                {"Parameter": "Interest Rate", "Value": base_model.transaction_and_financing["senior_interest_rate_pct"].value, "Unit": "%", "Notes": "Fixed interest rate."},
                {"Parameter": "Amortisation Years", "Value": _default_financing_assumptions(base_model)["amortization_period_years"], "Unit": "Years", "Notes": "Linear amortisation period."},
                {"Parameter": "Transaction Fees (%)", "Value": _default_valuation_assumptions(base_model)["transaction_cost_pct"], "Unit": "%", "Notes": "Fees as % of EV."},
            ],
            "equity": [
                {"Parameter": "Sponsor Equity Contribution", "Value": _default_equity_assumptions(base_model)["sponsor_equity_eur"], "Unit": "EUR", "Notes": "Management equity contribution."},
                {"Parameter": "Investor Equity Contribution", "Value": _default_equity_assumptions(base_model)["investor_equity_eur"], "Unit": "EUR", "Notes": "External investor contribution."},
                {"Parameter": "Investor Exit Year", "Value": _default_equity_assumptions(base_model)["exit_year"], "Unit": "Year", "Notes": "Exit year for investor."},
                {"Parameter": "Exit Multiple (x EBITDA)", "Value": _default_equity_assumptions(base_model)["exit_multiple"], "Unit": "x", "Notes": "Exit multiple on EBITDA."},
                {"Parameter": "Distribution Rule", "Value": "Pro-rata", "Unit": "", "Notes": "Fixed distribution rule."},
            ],
            "cashflow": [
                {"Parameter": "Tax Cash Rate", "Value": _default_cashflow_assumptions()["tax_cash_rate_pct"], "Unit": "%", "Notes": "Cash tax rate on EBT."},
                {"Parameter": "Tax Payment Lag", "Value": _default_cashflow_assumptions()["tax_payment_lag_years"], "Unit": "Years", "Notes": "Timing lag for cash taxes."},
                {"Parameter": "Capex (% of Revenue)", "Value": _default_cashflow_assumptions()["capex_pct_revenue"], "Unit": "%", "Notes": "Capex as % of revenue."},
                {"Parameter": "Working Capital (% of Revenue)", "Value": _default_cashflow_assumptions()["working_capital_pct_revenue"], "Unit": "%", "Notes": "Working capital adjustment."},
                {"Parameter": "Opening Cash Balance", "Value": _default_cashflow_assumptions()["opening_cash_balance_eur"], "Unit": "EUR", "Notes": "Opening cash balance."},
            ],
            "balance_sheet": [
                {"Parameter": "Opening Equity", "Value": _default_balance_sheet_assumptions(base_model)["opening_equity_eur"], "Unit": "EUR", "Notes": "Opening equity value."},
                {"Parameter": "Depreciation Rate", "Value": _default_balance_sheet_assumptions(base_model)["depreciation_rate_pct"], "Unit": "%", "Notes": "Fixed asset depreciation rate."},
                {"Parameter": "Minimum Cash Balance", "Value": _default_balance_sheet_assumptions(base_model)["minimum_cash_balance_eur"], "Unit": "EUR", "Notes": "Minimum cash balance."},
            ],
            "valuation": [
                {"Parameter": "Seller EBITDA Multiple", "Value": _default_valuation_assumptions(base_model)["seller_ebit_multiple"], "Unit": "x", "Notes": "EBITDA multiple for seller view."},
                {"Parameter": "Reference Year", "Value": _default_valuation_assumptions(base_model)["reference_year"], "Unit": "Year", "Notes": "Reference year for multiple."},
                {"Parameter": "Discount Rate (WACC)", "Value": _default_valuation_assumptions(base_model)["buyer_discount_rate"], "Unit": "%", "Notes": "DCF discount rate."},
                {"Parameter": "Valuation Start Year", "Value": _default_valuation_assumptions(base_model)["valuation_start_year"], "Unit": "Year", "Notes": "DCF start year."},
                {"Parameter": "Transaction Costs (%)", "Value": _default_valuation_assumptions(base_model)["transaction_cost_pct"], "Unit": "%", "Notes": "Fees as % of EV."},
            ],
        }

    def _ensure_assumptions_schema():
        defaults = _seed_assumptions_state()
        assumptions = st.session_state.get("assumptions", {})
        if isinstance(assumptions.get("revenue_model"), dict):
            if "reference" in assumptions["revenue_model"]:
                assumptions["revenue_model"] = defaults["revenue_model"]
        for key, value in defaults.items():
            if key not in assumptions:
                assumptions[key] = value
            elif isinstance(value, dict) and isinstance(assumptions.get(key), dict):
                for sub_key, sub_val in value.items():
                    if sub_key not in assumptions[key]:
                        assumptions[key][sub_key] = sub_val
        valuation_rows = assumptions.get("valuation", [])
        normalized_rows = []
        for row in valuation_rows:
            if row.get("Parameter") == "Debt at Close":
                continue
            if row.get("Parameter") == "Seller EBIT Multiple":
                row = dict(row)
                row["Parameter"] = "Seller EBITDA Multiple"
            normalized_rows.append(row)
        assumptions["valuation"] = normalized_rows
        st.session_state["assumptions"] = assumptions

    st.session_state.setdefault("assumptions", _seed_assumptions_state())
    _ensure_assumptions_schema()
    st.session_state.setdefault("assumptions.auto_sync", True)

    def _apply_assumptions_state():
        state = st.session_state["assumptions"]
        active_scenario = st.session_state.get("assumptions.scenario", "Base")
        scenario_col = active_scenario

        revenue_state = state.get("revenue_model", {})
        if revenue_state:
            st.session_state["revenue_model.reference_revenue_eur"] = _non_negative(
                revenue_state["reference_revenue_eur"].get(scenario_col, 0.0)
            )
            for year_index in range(5):
                st.session_state[
                    f"revenue_model.guarantee_pct_year_{year_index}"
                ] = _clamp_pct(
                    revenue_state["guarantee_pct_by_year"][scenario_col][
                        year_index
                    ]
                )
                st.session_state[
                    f"revenue_model.workdays_year_{year_index}"
                ] = _non_negative(
                    revenue_state["workdays_per_year"][scenario_col][year_index]
                )
                st.session_state[
                    f"revenue_model.utilization_rate_year_{year_index}"
                ] = _clamp_pct(
                    revenue_state["utilization_rate"][scenario_col][year_index]
                )
                st.session_state[
                    f"revenue_model.group_day_rate_eur_year_{year_index}"
                ] = _non_negative(
                    revenue_state["group_day_rate_eur"][scenario_col][year_index]
                )
                st.session_state[
                    f"revenue_model.external_day_rate_eur_year_{year_index}"
                ] = _non_negative(
                    revenue_state["external_day_rate_eur"][scenario_col][year_index]
                )
                st.session_state[
                    f"revenue_model.day_rate_growth_pct_year_{year_index}"
                ] = _clamp_pct(
                    revenue_state["day_rate_growth_pct"][scenario_col][year_index]
                )
                st.session_state[
                    f"revenue_model.revenue_growth_pct_year_{year_index}"
                ] = _clamp_pct(
                    revenue_state["revenue_growth_pct"][scenario_col][year_index]
                )
                st.session_state[
                    f"revenue_model.group_capacity_share_pct_year_{year_index}"
                ] = _clamp_pct(
                    revenue_state["group_capacity_share_pct"][scenario_col][
                        year_index
                    ]
                )
                st.session_state[
                    f"revenue_model.external_capacity_share_pct_year_{year_index}"
                ] = _clamp_pct(
                    revenue_state["external_capacity_share_pct"][scenario_col][
                        year_index
                    ]
                )

        cost_state = state.get("cost_model", {})
        if "inflation" in cost_state:
            st.session_state["cost_model.apply_inflation"] = bool(
                cost_state["inflation"].get("apply", False)
            )
            st.session_state["cost_model.inflation_rate_pct"] = _clamp_pct(
                cost_state["inflation"].get("rate_pct", 0.0)
            )
        for row in cost_state.get("personnel", []):
            year_index = int(row["Year"].split()[-1])
            st.session_state[
                f"cost_model.consultant_fte_year_{year_index}"
            ] = _non_negative(row["Consultant FTE"])
            st.session_state[
                f"cost_model.consultant_base_cost_eur_year_{year_index}"
            ] = _non_negative(row["Consultant Loaded Cost (EUR)"])
            st.session_state[
                f"cost_model.backoffice_fte_year_{year_index}"
            ] = _non_negative(row["Backoffice FTE"])
            st.session_state[
                f"cost_model.backoffice_base_cost_eur_year_{year_index}"
            ] = _non_negative(row["Backoffice Loaded Cost (EUR)"])
            st.session_state[
                f"cost_model.management_cost_eur_year_{year_index}"
            ] = _non_negative(row["Management Cost (EUR)"])

        for row in cost_state.get("fixed_overhead", []):
            year_index = int(row["Year"].split()[-1])
            st.session_state[
                f"cost_model.fixed_overhead_advisory_year_{year_index}"
            ] = _non_negative(row["Advisory"])
            st.session_state[
                f"cost_model.fixed_overhead_legal_year_{year_index}"
            ] = _non_negative(row["Legal"])
            st.session_state[
                f"cost_model.fixed_overhead_it_year_{year_index}"
            ] = _non_negative(row["IT & Software"])
            st.session_state[
                f"cost_model.fixed_overhead_office_year_{year_index}"
            ] = _non_negative(row["Office Rent"])
            st.session_state[
                f"cost_model.fixed_overhead_services_year_{year_index}"
            ] = _non_negative(row["Services"])
            st.session_state[
                f"cost_model.fixed_overhead_other_year_{year_index}"
            ] = _non_negative(row["Other Services"])

        for row in cost_state.get("variable_costs", []):
            year_index = int(row["Year"].split()[-1])
            training_value = _non_negative(row["Training Value"])
            travel_value = _non_negative(row["Travel Value"])
            communication_value = _non_negative(row["Communication Value"])

            st.session_state[
                f"cost_model.variable_training_pct_year_{year_index}"
            ] = training_value if row["Training Type"] == "%" else 0.0
            st.session_state[
                f"cost_model.variable_training_eur_year_{year_index}"
            ] = training_value if row["Training Type"] == "EUR" else 0.0
            st.session_state[
                f"cost_model.variable_travel_pct_year_{year_index}"
            ] = travel_value if row["Travel Type"] == "%" else 0.0
            st.session_state[
                f"cost_model.variable_travel_eur_year_{year_index}"
            ] = travel_value if row["Travel Type"] == "EUR" else 0.0
            st.session_state[
                f"cost_model.variable_communication_pct_year_{year_index}"
            ] = communication_value if row["Communication Type"] == "%" else 0.0
            st.session_state[
                f"cost_model.variable_communication_eur_year_{year_index}"
            ] = communication_value if row["Communication Type"] == "EUR" else 0.0

        for row in state["personnel_costs"]:
            role = row["Role"]
            if role == "Consultant Base Salary":
                st.session_state[
                    "personnel_cost_assumptions.avg_consultant_base_cost_eur_per_year"
                ] = _non_negative(row["Base Value (EUR)"])
                st.session_state["personnel_cost_assumptions.wage_inflation_pct"] = _clamp_pct(row["Growth (%)"])
            elif role == "Consultant Variable (% Revenue)":
                st.session_state["personnel_cost_assumptions.bonus_pct_of_base"] = _clamp_pct(row["Base Value (EUR)"])
            elif role == "Backoffice Cost per FTE":
                st.session_state["operating_assumptions.avg_backoffice_salary_eur_per_year"] = _non_negative(row["Base Value (EUR)"])
            elif role == "Management / MD Cost":
                # Wire management fixed cost and growth into session state.
                st.session_state["personnel_costs.management_md_cost_eur"] = _non_negative(row["Base Value (EUR)"])
                st.session_state["personnel_costs.management_md_growth_pct"] = _clamp_pct(row["Growth (%)"])

        for row in state["opex"]:
            category = row["Category"]
            if category == "External Consulting":
                st.session_state["overhead_and_variable_costs.legal_audit_eur_per_year"] = _non_negative(row["Value"])
            elif category == "IT":
                st.session_state["overhead_and_variable_costs.it_and_software_eur_per_year"] = _non_negative(row["Value"])
            elif category == "Office":
                st.session_state["overhead_and_variable_costs.rent_eur_per_year"] = _non_negative(row["Value"])
            elif category == "Other Services":
                st.session_state["overhead_and_variable_costs.other_overhead_eur_per_year"] = _non_negative(row["Value"])

        for row in state["financing"]:
            param = row["Parameter"]
            if param == "Senior Debt Amount":
                st.session_state["transaction_and_financing.senior_term_loan_start_eur"] = _non_negative(row["Value"])
            elif param == "Interest Rate":
                st.session_state["transaction_and_financing.senior_interest_rate_pct"] = _clamp_pct(row["Value"])
            elif param == "Amortisation Years":
                st.session_state["financing.amortization_period_years"] = _non_negative(row["Value"])
            elif param == "Transaction Fees (%)":
                st.session_state["valuation.transaction_cost_pct"] = _clamp_pct(row["Value"])

        for row in state["equity"]:
            param = row["Parameter"]
            if param == "Sponsor Equity Contribution":
                st.session_state["equity.sponsor_equity_eur"] = _non_negative(row["Value"])
            elif param == "Investor Equity Contribution":
                st.session_state["equity.investor_equity_eur"] = _non_negative(row["Value"])
            elif param == "Investor Exit Year":
                try:
                    exit_year = int(float(row["Value"]))
                except (TypeError, ValueError):
                    exit_year = _default_equity_assumptions(base_model)["exit_year"]
                st.session_state["equity.exit_year"] = int(max(3, min(7, exit_year)))
            elif param == "Exit Multiple (x EBITDA)":
                st.session_state["equity.exit_multiple"] = _non_negative(row["Value"])

        for row in state["cashflow"]:
            param = row["Parameter"]
            if param == "Tax Cash Rate":
                st.session_state["cashflow.tax_cash_rate_pct"] = _clamp_pct(row["Value"])
            elif param == "Tax Payment Lag":
                st.session_state["cashflow.tax_payment_lag_years"] = int(max(0, min(1, row["Value"])))
            elif param == "Capex (% of Revenue)":
                st.session_state["cashflow.capex_pct_revenue"] = _clamp_pct(row["Value"])
            elif param == "Working Capital (% of Revenue)":
                st.session_state["cashflow.working_capital_pct_revenue"] = _clamp_pct(row["Value"])
            elif param == "Opening Cash Balance":
                st.session_state["cashflow.opening_cash_balance_eur"] = _non_negative(row["Value"])

        for row in state["balance_sheet"]:
            param = row["Parameter"]
            if param == "Opening Equity":
                st.session_state["balance_sheet.opening_equity_eur"] = _non_negative(row["Value"])
            elif param == "Depreciation Rate":
                st.session_state["balance_sheet.depreciation_rate_pct"] = _clamp_pct(row["Value"])
            elif param == "Minimum Cash Balance":
                st.session_state["balance_sheet.minimum_cash_balance_eur"] = _non_negative(row["Value"])

        for row in state["valuation"]:
            param = row["Parameter"]
            if param == "Seller EBITDA Multiple":
                st.session_state["valuation.seller_ebit_multiple"] = _non_negative(row["Value"])
            elif param == "Reference Year":
                st.session_state["valuation.reference_year"] = int(max(0, row["Value"]))
            elif param == "Discount Rate (WACC)":
                st.session_state["valuation.buyer_discount_rate"] = _clamp_pct(row["Value"])
            elif param == "Valuation Start Year":
                st.session_state["valuation.valuation_start_year"] = int(max(0, row["Value"]))
            elif param == "Transaction Costs (%)":
                st.session_state["valuation.transaction_cost_pct"] = _clamp_pct(row["Value"])

    _apply_assumptions_state()

    def _hash_payload(payload):
        payload_str = json.dumps(payload, sort_keys=True, default=str)
        return hashlib.sha256(payload_str.encode("utf-8")).hexdigest()

    def _render_input_scenario_selector():
        if st.session_state.get("page_key") not in {
            "Revenue Model",
            "Cost Model",
        }:
            return
        st.radio(
            label="",
            options=["Worst", "Base", "Best"],
            horizontal=True,
            key="assumptions.scenario",
        )

    def _render_output_scenario_selector():
        st.radio(
            "Scenario (View Only)",
            options=["Worst", "Base", "Best"],
            horizontal=True,
            key="output_scenario",
        )

    st.session_state.setdefault("assumptions.scenario", "Base")
    st.session_state.setdefault("output_scenario", "Base")
    input_scenario = st.session_state["assumptions.scenario"]
    st.session_state["scenario_selection.selected_scenario"] = input_scenario

    # Build input model and collect editable values from the assumptions page.
    selected_scenario = input_scenario
    scenario_key = selected_scenario.lower()
    input_model = create_demo_input_model()
    for section_key, section_value in input_model.__dict__.items():
        if isinstance(section_value, dict):
            edited_values = _collect_values_from_session(
                section_value, section_key
            )
            _apply_section_values(section_value, edited_values)

    input_model.scenario_selection["selected_scenario"].value = selected_scenario
    input_model.utilization_by_year = st.session_state.get(
        "utilization_by_year", []
    )
    input_model.management_md_cost_eur_per_year = st.session_state.get(
        "personnel_costs.management_md_cost_eur", 0.0
    )
    input_model.management_md_cost_growth_pct = st.session_state.get(
        "personnel_costs.management_md_growth_pct", 0.0
    )

    input_model.cashflow_assumptions = _default_cashflow_assumptions()
    for key, default_value in input_model.cashflow_assumptions.items():
        input_model.cashflow_assumptions[key] = st.session_state.get(
            f"cashflow.{key}", default_value
        )
    input_model.balance_sheet_assumptions = _default_balance_sheet_assumptions(
        input_model
    )
    for key, default_value in input_model.balance_sheet_assumptions.items():
        input_model.balance_sheet_assumptions[key] = st.session_state.get(
            f"balance_sheet.{key}", default_value
        )
    input_model.financing_assumptions = _default_financing_assumptions(
        input_model
    )
    for key, default_value in input_model.financing_assumptions.items():
        input_model.financing_assumptions[key] = st.session_state.get(
            f"financing.{key}", default_value
        )
    senior_debt_amount = st.session_state.get("financing.senior_debt_amount")
    if senior_debt_amount is None:
        raise ValueError("Senior Debt Amount missing from financing assumptions.")
    input_model.financing_assumptions["senior_debt_amount"] = senior_debt_amount
    input_model.financing_assumptions["initial_debt_eur"] = senior_debt_amount
    input_model.transaction_and_financing[
        "senior_term_loan_start_eur"
    ].value = senior_debt_amount
    input_model.valuation_runtime = _default_valuation_assumptions(input_model)
    for key, default_value in input_model.valuation_runtime.items():
        input_model.valuation_runtime[key] = st.session_state.get(
            f"valuation.{key}", default_value
        )
    input_model.equity_assumptions = _default_equity_assumptions(input_model)
    for key, default_value in input_model.equity_assumptions.items():
        input_model.equity_assumptions[key] = st.session_state.get(
            f"equity.{key}", default_value
        )

    page = page_override or st.session_state.get(
        "page_key", "Operating Model (P&L)"
    )
    output_pages = {
        "Overview",
        "Operating Model (P&L)",
        "Cashflow & Liquidity",
        "Balance Sheet",
        "Financing & Debt",
        "Equity Case",
        "Valuation & Purchase Price",
        "Model Settings",
    }
    output_selector_pages = {
        "Operating Model (P&L)",
        "Cashflow & Liquidity",
        "Balance Sheet",
    }
    if page in output_pages:
        if page in output_selector_pages:
            _render_output_scenario_selector()
        output_scenario = st.session_state["output_scenario"]

        revenue_state = st.session_state["assumptions"]["revenue_model"]
        cost_state = st.session_state["assumptions"]["cost_model"]
        financing_state = st.session_state["assumptions"]["financing"]
        cashflow_state = st.session_state["assumptions"]["cashflow"]
        balance_state = st.session_state["assumptions"]["balance_sheet"]
        valuation_state = st.session_state["assumptions"]["valuation"]

        revenue_hash = _hash_payload(
            {"scenario": output_scenario, "revenue": revenue_state}
        )
        if st.session_state.get("cache.revenue_hash") != revenue_hash:
            revenue_final_by_year, revenue_components_by_year = build_revenue_model_outputs(
                st.session_state["assumptions"], output_scenario
            )
            st.session_state["cache.revenue_hash"] = revenue_hash
            st.session_state["cache.revenue_outputs"] = (
                revenue_final_by_year,
                revenue_components_by_year,
            )
        else:
            revenue_final_by_year, revenue_components_by_year = st.session_state[
                "cache.revenue_outputs"
            ]

        cost_hash = _hash_payload(
            {"cost": cost_state, "revenue_final": revenue_final_by_year}
        )
        if st.session_state.get("cache.cost_hash") != cost_hash:
            cost_model_totals = build_cost_model_outputs(
                st.session_state["assumptions"], revenue_final_by_year
            )
            st.session_state["cache.cost_hash"] = cost_hash
            st.session_state["cache.cost_outputs"] = cost_model_totals
        else:
            cost_model_totals = st.session_state["cache.cost_outputs"]

        input_model.revenue_final_by_year = revenue_final_by_year
        input_model.revenue_components_by_year = revenue_components_by_year
        input_model.cost_model_totals_by_year = cost_model_totals

        debt_inputs = {
            "senior_debt_amount": input_model.financing_assumptions[
                "senior_debt_amount"
            ],
            "interest_rate_pct": input_model.financing_assumptions[
                "interest_rate_pct"
            ],
            "amortization_type": input_model.financing_assumptions[
                "amortization_type"
            ],
            "amortization_period_years": input_model.financing_assumptions[
                "amortization_period_years"
            ],
            "grace_period_years": input_model.financing_assumptions[
                "grace_period_years"
            ],
            "special_repayment_year": input_model.financing_assumptions[
                "special_repayment_year"
            ],
            "special_repayment_amount_eur": input_model.financing_assumptions[
                "special_repayment_amount_eur"
            ],
            "minimum_dscr": input_model.financing_assumptions["minimum_dscr"],
        }
        debt_hash = _hash_payload({"financing": financing_state, "debt": debt_inputs})
        if st.session_state.get("cache.debt_hash") != debt_hash:
            debt_schedule = calculate_debt_schedule(input_model)
            st.session_state["cache.debt_hash"] = debt_hash
            st.session_state["cache.debt_schedule"] = debt_schedule
        else:
            debt_schedule = st.session_state["cache.debt_schedule"]

        pnl_hash = _hash_payload(
            {
                "revenue_final": revenue_final_by_year,
                "cost_totals": cost_model_totals,
                "debt": debt_schedule,
            }
        )
        if st.session_state.get("cache.pnl_hash") != pnl_hash:
            pnl_list = calculate_pnl(
                input_model,
                revenue_final_by_year=revenue_final_by_year,
                cost_totals_by_year=cost_model_totals,
                debt_schedule=debt_schedule,
            )
            st.session_state["cache.pnl_hash"] = pnl_hash
            st.session_state["cache.pnl_list"] = pnl_list
        else:
            pnl_list = st.session_state["cache.pnl_list"]
        pnl_result = {f"Year {row['year']}": row for row in pnl_list}

        cashflow_hash = _hash_payload(
            {
                "pnl": pnl_list,
                "debt": debt_schedule,
                "cashflow": cashflow_state,
            }
        )
        if st.session_state.get("cache.cashflow_hash") != cashflow_hash:
            cashflow_result = calculate_cashflow(
                input_model, pnl_list, debt_schedule
            )
            st.session_state["cache.cashflow_hash"] = cashflow_hash
            st.session_state["cache.cashflow_result"] = cashflow_result
        else:
            cashflow_result = st.session_state["cache.cashflow_result"]

        balance_hash = _hash_payload(
            {
                "cashflow": cashflow_result,
                "debt": debt_schedule,
                "pnl": pnl_list,
                "balance": balance_state,
            }
        )
        if st.session_state.get("cache.balance_hash") != balance_hash:
            balance_sheet = calculate_balance_sheet(
                input_model, cashflow_result, debt_schedule, pnl_list
            )
            st.session_state["cache.balance_hash"] = balance_hash
            st.session_state["cache.balance_sheet"] = balance_sheet
        else:
            balance_sheet = st.session_state["cache.balance_sheet"]

        investment_hash = _hash_payload(
            {
                "cashflow": cashflow_result,
                "pnl": pnl_list,
                "balance": balance_sheet,
                "valuation": valuation_state,
            }
        )
        if st.session_state.get("cache.investment_hash") != investment_hash:
            investment_result = calculate_investment(
                input_model, cashflow_result, pnl_list, balance_sheet
            )
            st.session_state["cache.investment_hash"] = investment_hash
            st.session_state["cache.investment_result"] = investment_result
        else:
            investment_result = st.session_state["cache.investment_result"]
    editor_css = """
    <style>
      .rdg-cell[aria-readonly="true"] {
        background: #f3f4f6;
        color: #6b7280;
        font-style: italic;
      }
      .rdg-cell[aria-readonly="false"] {
        background: #eff6ff;
        border: 1px solid #93c5fd;
      }
    </style>
    """
    st.markdown(editor_css, unsafe_allow_html=True)

    assumptions_state = st.session_state["assumptions"]

    def _sidebar_editor(title, key, df, column_config):
        st.markdown(f"### {title}")
        display_df = _apply_unit_display(df)
        config = dict(column_config)
        for col in display_df.columns:
            if col not in config:
                config[col] = st.column_config.TextColumn()
        edited = st.data_editor(
            display_df,
            hide_index=True,
            key=key,
            column_config=config,
            use_container_width=True,
        )
        return _restore_unit_values(edited)

    if page == "Cashflow & Liquidity":
        cashflow_df = pd.DataFrame(assumptions_state["cashflow"])
        edited_cashflow = _sidebar_editor(
            "Cashflow Assumptions",
            "sidebar.cashflow",
            cashflow_df,
            {
                "Parameter": st.column_config.TextColumn(disabled=True),
                "Unit": st.column_config.TextColumn(disabled=True),
                "Notes": st.column_config.TextColumn(disabled=True),
            },
        )
        assumptions_state["cashflow"] = edited_cashflow.to_dict("records")
        _apply_assumptions_state()

    if page == "Balance Sheet":
        balance_df = pd.DataFrame(assumptions_state["balance_sheet"])
        edited_balance = _sidebar_editor(
            "Balance Sheet Assumptions",
            "sidebar.balance_sheet",
            balance_df,
            {
                "Parameter": st.column_config.TextColumn(disabled=True),
                "Unit": st.column_config.TextColumn(disabled=True),
                "Notes": st.column_config.TextColumn(disabled=True),
            },
        )
        assumptions_state["balance_sheet"] = edited_balance.to_dict("records")
        _apply_assumptions_state()

    if page == "Financing & Debt":
        with st.expander("Financing Assumptions", expanded=False):
            financing_df = pd.DataFrame(assumptions_state["financing"])
            edited_financing = _sidebar_editor(
                "Financing Assumptions",
                "sidebar.financing",
                financing_df,
                {
                    "Parameter": st.column_config.TextColumn(disabled=True),
                    "Unit": st.column_config.TextColumn(disabled=True),
                    "Notes": st.column_config.TextColumn(disabled=True),
                },
            )
            assumptions_state["financing"] = edited_financing.to_dict("records")
            _apply_assumptions_state()

    if page == "Valuation & Purchase Price":
        with st.expander("Valuation Assumptions", expanded=False):
            valuation_df = pd.DataFrame(assumptions_state["valuation"])
            edited_valuation = _sidebar_editor(
                "Valuation Assumptions",
                "sidebar.valuation",
                valuation_df,
                {
                    "Parameter": st.column_config.TextColumn(disabled=True),
                    "Unit": st.column_config.TextColumn(disabled=True),
                    "Notes": st.column_config.TextColumn(disabled=True),
                },
            )
            assumptions_state["valuation"] = edited_valuation.to_dict("records")
            _apply_assumptions_state()

    if page == "Equity Case":
        with st.expander("Equity Assumptions", expanded=False):
            equity_df = pd.DataFrame(assumptions_state["equity"])
            edited_equity = _sidebar_editor(
                "Equity Assumptions",
                "sidebar.equity",
                equity_df,
                {
                    "Parameter": st.column_config.TextColumn(disabled=True),
                    "Unit": st.column_config.TextColumn(disabled=True),
                    "Notes": st.column_config.TextColumn(disabled=True),
                },
            )
            assumptions_state["equity"] = edited_equity.to_dict("records")
            _apply_assumptions_state()

    if page == "Revenue Model":
        st.title("Revenue Model")
        _render_input_scenario_selector()
        render_revenue_model_assumptions(input_model, show_header=False)
        _apply_assumptions_state()
        return

    if page == "Cost Model":
        st.title("Cost Model")
        _render_input_scenario_selector()
        render_cost_model_assumptions(input_model, show_header=False)
        _apply_assumptions_state()
        return

    if page == "Other Assumptions":
        st.title("Other Assumptions")
        st.write("Master input sheet – all remaining assumptions.")
        # Other Assumptions are intentionally scenario-agnostic.
        render_advanced_assumptions(input_model, show_header=False)
        return

    if page == "Overview":
        st.title("Deal Summary (IC View)")
        st.caption(
            "Conservative decision view based on current inputs and selected output scenario."
        )
        scenario_label = st.session_state.get("output_scenario", "Base")
        st.caption(f"Scenario being viewed: {scenario_label}")
        st.divider()

        def _safe_div(numerator, denominator):
            return numerator / denominator if denominator else 0.0

        def _min_with_year(values, years):
            if not values:
                return 0.0, None
            min_value = min(values)
            min_index = values.index(min_value)
            return min_value, years[min_index]

        def _max_value(values):
            return max(values) if values else 0.0

        def _format_multiple(value):
            if value is None:
                return "—"
            return f"{value:.2f}x"

        def _status_from_thresholds(value, warn_at, fail_at, higher_is_better=True):
            if value is None:
                return "OK"
            if higher_is_better:
                if value < fail_at:
                    return "FAIL"
                if value < warn_at:
                    return "WATCH"
                return "OK"
            if value > fail_at:
                return "FAIL"
            if value > warn_at:
                return "WATCH"
            return "OK"

        year_indices = [row.get("year", 0) for row in cashflow_result]
        year_indices = year_indices or [0]
        year1_index = 1 if 1 in year_indices else year_indices[0]
        last_year_index = max(year_indices)
        cashflow_by_year = {row["year"]: row for row in cashflow_result}
        debt_by_year = {row["year"]: row for row in debt_schedule}
        balance_by_year = {row["year"]: row for row in balance_sheet}

        pnl_year1 = pnl_result.get(f"Year {year1_index}", {})
        pnl_last = pnl_result.get(f"Year {last_year_index}", {})
        revenue_y1 = pnl_year1.get("revenue", 0.0)
        ebitda_y1 = pnl_year1.get("ebitda", 0.0)
        margin_y1 = _safe_div(ebitda_y1, revenue_y1)

        cashflow_year1 = cashflow_by_year.get(year1_index, {})
        fcf_y1 = cashflow_year1.get("free_cashflow")
        net_income_y1 = pnl_year1.get("net_income", 0.0)

        cash_balances = [
            row.get("cash_balance", 0.0) for row in cashflow_result
        ]
        cash_years = [row.get("year", 0) for row in cashflow_result]
        min_cash_balance, min_cash_year = _min_with_year(
            cash_balances, cash_years
        )

        purchase_price = input_model.transaction_and_financing[
            "purchase_price_eur"
        ].value
        equity_assumptions = input_model.equity_assumptions
        sponsor_equity = equity_assumptions.get("sponsor_equity_eur", 0.0)
        investor_equity = equity_assumptions.get("investor_equity_eur", 0.0)
        required_equity = sponsor_equity + investor_equity
        if required_equity == 0:
            required_equity = input_model.transaction_and_financing[
                "equity_contribution_eur"
            ].value

        debt_at_close = debt_schedule[0]["debt_drawdown"] if debt_schedule else 0.0
        peak_debt = _max_value(
            [
                row.get("opening_debt", 0.0) + row.get("debt_drawdown", 0.0)
                for row in debt_schedule
            ]
        )
        entry_multiple = _safe_div(debt_at_close, ebitda_y1)
        entry_label = "Debt/EBITDA (Entry)"
        exit_year = equity_assumptions.get("exit_year")
        exit_multiple = equity_assumptions.get("exit_multiple")

        debt_year1 = debt_by_year.get(year1_index, {})
        closing_debt_y1 = debt_year1.get("closing_debt", 0.0)
        debt_to_ebitda_y1 = _safe_div(closing_debt_y1, ebitda_y1)

        dscr_values = []
        for year, cf_row in cashflow_by_year.items():
            debt_row = debt_by_year.get(year, {})
            debt_service = debt_row.get("interest_expense", 0.0) + debt_row.get(
                "total_repayment", 0.0
            )
            if debt_service <= 0:
                continue
            cfads = (
                cf_row.get("ebitda", 0.0)
                - cf_row.get("taxes_paid", 0.0)
                - cf_row.get("capex", 0.0)
                - cf_row.get("working_capital_change", 0.0)
            )
            dscr_values.append(_safe_div(cfads, debt_service))
        min_dscr = min(dscr_values) if dscr_values else None

        st.markdown(
            "### A. Deal Snapshot (What are we buying and how is it funded?)"
        )
        snapshot_cols = st.columns(6)
        snapshot_cols[0].metric(
            "Purchase Price", format_currency(purchase_price)
        )
        snapshot_cols[1].metric(
            "Debt at Close", format_currency(debt_at_close)
        )
        snapshot_cols[2].metric(
            "Equity at Close", format_currency(required_equity)
        )
        snapshot_cols[3].metric(entry_label, _format_multiple(entry_multiple))
        snapshot_cols[4].metric(
            "Exit Year",
            f"Year {exit_year}" if exit_year is not None else "—",
        )
        snapshot_cols[5].metric(
            "Exit Multiple",
            _format_multiple(exit_multiple)
            if exit_multiple is not None
            else "—",
        )
        st.markdown("**Interpretation**")
        st.markdown(
            "- Debt at Close shows the initial borrowing that must be serviced from operating cash."
        )
        st.markdown(
            "- Equity at Close is the cash contributed by management and investors before debt service begins."
        )
        st.markdown(
            "- Debt/EBITDA (Entry) indicates leverage at entry; higher leverage reduces flexibility in an MBO."
        )
        st.divider()

        st.markdown(
            "### B. Operating Strength (Does the business generate cash reliably?)"
        )
        operating_cols = st.columns(4)
        operating_cols[0].metric(
            f"Revenue (Year {year1_index})", format_currency(revenue_y1)
        )
        operating_cols[1].metric(
            f"EBITDA (Year {year1_index})", format_currency(ebitda_y1)
        )
        operating_cols[2].metric(
            f"EBITDA Margin (Year {year1_index})", format_pct(margin_y1)
        )
        if fcf_y1 is not None:
            operating_cols[3].metric(
                f"FCF (Year {year1_index})", format_currency(fcf_y1)
            )
        else:
            operating_cols[3].metric(
                f"Net Income (Year {year1_index})",
                format_currency(net_income_y1),
            )
        st.info(
            "What this tells us:\n"
            "- EBITDA margin drives debt capacity and headroom in downside cases.\n"
            "- Cash conversion matters more than accounting profit for lenders.\n"
            "- This view uses modeled five-year cash generation only; no additional upside is assumed."
        )
        st.divider()

        st.markdown(
            "### C. Bankability & Liquidity (Can we carry the deal without running out of cash?)"
        )
        bank_cols = st.columns(5)
        bank_cols[0].metric(
            "Minimum Cash", format_currency(min_cash_balance)
        )
        bank_cols[1].metric(
            "Year of Min Cash",
            f"Year {min_cash_year}" if min_cash_year is not None else "—",
        )
        bank_cols[2].metric("Peak Debt", format_currency(peak_debt))
        bank_cols[3].metric(
            "Min DSCR", _format_multiple(min_dscr) if min_dscr is not None else "—"
        )
        bank_cols[4].metric(
            f"Debt/EBITDA (Year {year1_index})",
            _format_multiple(debt_to_ebitda_y1),
        )

        st.markdown("**Bank view interpretation**")
        st.markdown(
            "- Minimum cash shows the lowest liquidity point; persistent negatives indicate funding risk."
        )
        st.markdown(
            "- Peak debt reflects the highest balance the business must carry through the plan."
        )
        if min_dscr is not None:
            st.markdown(
                "- DSCR measures cash available for debt service versus debt payments in each year."
            )
        st.markdown(
            "- Conservative guidelines: Min cash > 0; Debt/EBITDA < 3.0x"
            + ("; DSCR > 1.3x." if min_dscr is not None else ".")
        )

        st.markdown("**Stress signals**")
        st.markdown(
            f"- Minimum cash: {format_currency(min_cash_balance)} (guideline > 0)"
        )
        st.markdown(
            f"- Peak debt: {format_currency(peak_debt)} (track deleveraging pace)"
        )
        st.markdown(
            f"- Debt/EBITDA: {_format_multiple(debt_to_ebitda_y1)} (guideline < 3.0x)"
        )
        if min_dscr is not None:
            st.markdown(
                f"- Min DSCR: {_format_multiple(min_dscr)} (guideline > 1.3x)"
            )

        if min_cash_balance < 0:
            st.error(
                "Minimum cash is below zero, indicating a funding shortfall in the plan."
            )
        elif min_cash_balance < 500_000:
            st.warning(
                "Minimum cash remains positive but thin; liquidity headroom is limited."
            )
        else:
            st.success(
                "Minimum cash remains positive with a basic liquidity buffer."
            )

        if debt_to_ebitda_y1 >= 3.0:
            st.error(
                "Debt/EBITDA exceeds 3.0x, indicating aggressive leverage at entry."
            )
        elif debt_to_ebitda_y1 >= 2.5:
            st.warning(
                "Debt/EBITDA is approaching 3.0x; leverage is tight for this plan."
            )
        else:
            st.success(
                "Debt/EBITDA is within a conservative leverage range."
            )

        if min_dscr is not None:
            if min_dscr < 1.3:
                st.error(
                    "Minimum DSCR is below 1.3x, indicating limited debt service headroom."
                )
            elif min_dscr < 1.5:
                st.warning(
                    "Minimum DSCR is above 1.3x but remains tight."
                )
            else:
                st.success(
                    "Minimum DSCR is above 1.5x, indicating comfortable coverage."
                )

        with st.expander("Show details"):
            st.dataframe(
                pd.DataFrame(pnl_list), use_container_width=True
            )
            st.dataframe(
                pd.DataFrame(cashflow_result), use_container_width=True
            )
            st.dataframe(
                pd.DataFrame(debt_schedule), use_container_width=True
            )
        st.divider()

        st.markdown("### D. Deal Breakers (What kills the deal first?)")
        negative_cash_years = [
            row["year"]
            for row in cashflow_result
            if row.get("cash_balance", 0.0) < 0
        ]
        exit_year_index = (
            exit_year
            if exit_year in debt_by_year
            else last_year_index
        )
        closing_debt_exit = debt_by_year.get(exit_year_index, {}).get(
            "closing_debt", 0.0
        )
        taxes_payable = [
            row.get("tax_payable", 0.0) for row in balance_sheet
        ]
        taxes_payable_rising = False
        if len(taxes_payable) >= 3:
            increases = sum(
                1
                for i in range(1, len(taxes_payable))
                if taxes_payable[i] > taxes_payable[i - 1]
            )
            taxes_payable_rising = (
                taxes_payable[-1] - taxes_payable[0] > 100_000
                and increases >= len(taxes_payable) - 1
            )

        floor_coverage_years = [
            input_model.operating_assumptions[
                "revenue_guarantee_pct_year_1"
            ].value,
            input_model.operating_assumptions[
                "revenue_guarantee_pct_year_2"
            ].value,
            input_model.operating_assumptions[
                "revenue_guarantee_pct_year_3"
            ].value,
        ]
        floor_years_active = [
            i + 1 for i, value in enumerate(floor_coverage_years) if value > 0
        ]
        last_floor_year = max(floor_years_active) if floor_years_active else None
        debt_after_floor = False
        if last_floor_year is not None:
            debt_after_floor = any(
                row["year"] > last_floor_year
                and row.get("closing_debt", 0.0) > 0
                for row in debt_schedule
            )

        margin_last = _safe_div(
            pnl_last.get("ebitda", 0.0), pnl_last.get("revenue", 0.0)
        )
        margin_drop = margin_y1 - margin_last

        flags = [
            {
                "title": "Negative cash in any year",
                "status": "FAIL" if negative_cash_years else "OK",
                "impact": "Impact: Cash balance falls below zero, creating an immediate funding gap.",
                "fix": "Typical fix: reduce purchase price, add equity, or slow capex.",
            },
            {
                "title": "Minimum cash below zero",
                "status": "FAIL" if min_cash_balance < 0 else "OK",
                "impact": f"Impact: Minimum cash reaches {format_currency(min_cash_balance)}.",
                "fix": "Typical fix: increase liquidity buffer or reduce leverage at close.",
            },
            {
                "title": "Debt not fully repaid by exit",
                "status": "FAIL" if closing_debt_exit > 0 else "OK",
                "impact": "Impact: Residual debt reduces equity value at exit.",
                "fix": "Typical fix: lower initial debt or extend the repayment period assumptions.",
            },
            {
                "title": "Taxes payable growing structurally",
                "status": "WATCH" if taxes_payable_rising else "OK",
                "impact": "Impact: Tax payable build-up can strain near-term liquidity.",
                "fix": "Typical fix: validate tax timing and working capital assumptions.",
            },
            {
                "title": "EBITDA margin deterioration",
                "status": "FAIL" if margin_drop > 0.05 else "WATCH" if margin_drop > 0.02 else "OK",
                "impact": "Impact: Margin erosion reduces cash available for debt service.",
                "fix": "Typical fix: validate pricing, utilization, and cost controls.",
            },
        ]
        if last_floor_year is not None:
            flags.append(
                {
                    "title": "Floor coverage ends before deleveraging",
                    "status": "WATCH" if debt_after_floor else "OK",
                    "impact": "Impact: Debt remains after floor coverage ends, increasing downside exposure.",
                    "fix": "Typical fix: stress test without floor coverage and add equity headroom.",
                }
            )
        if min_dscr is not None:
            flags.append(
                {
                    "title": "Minimum DSCR below guideline",
                    "status": _status_from_thresholds(
                        min_dscr, 1.5, 1.3, higher_is_better=True
                    ),
                    "impact": "Impact: Lower coverage raises covenant and refinancing risk.",
                    "fix": "Typical fix: reduce leverage or increase cash generation assumptions.",
                }
            )

        for flag in flags:
            left_col, right_col = st.columns([1, 2])
            with left_col:
                st.markdown(f"**{flag['status']}** — {flag['title']}")
            with right_col:
                st.write(f"{flag['impact']} {flag['fix']}")
        st.divider()

        st.markdown("### E. Management Takeaway (So what?)")
        negative_cash = bool(negative_cash_years) or min_cash_balance < 0
        deleveraging_slow = closing_debt_exit > 0
        if negative_cash:
            verdict = "NO-GO"
        elif min_cash_balance < 500_000 or deleveraging_slow:
            verdict = "GO WITH CONDITIONS"
        else:
            verdict = "GO"
        st.markdown(f"**Verdict:** {verdict}")

        st.markdown("**What to negotiate**")
        st.markdown("- Purchase price adjustment to protect entry leverage.")
        st.markdown("- Higher equity contribution to improve liquidity headroom.")
        st.markdown("- Debt amount reduction to lower leverage at entry.")

        st.markdown("**What to validate operationally**")
        st.markdown("- Utilization stability versus plan.")
        st.markdown("- Day rate resilience under client pressure.")
        st.markdown("- Client retention and pipeline visibility.")

        st.markdown("**Next model checks**")
        st.markdown("- Run the worst case output scenario.")
        st.markdown("- Remove floor coverage to test downside cash.")
        st.markdown("- Reduce utilization and confirm debt service headroom.")

    if page == "Model Settings":
        st.title("Model Settings")
        st.caption("Model transparency, export, and technical controls")

        st.markdown("### Model Snapshot / Export")
        st.markdown(
            """
            <div style="background:#f3f4f6;padding:12px 14px;border-radius:6px;">
              <strong>Model Snapshot</strong><br/>
              <span style="color:#6b7280;">For internal / AI-assisted analysis only</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button(
            "Generate Model Snapshot for GPT",
            key="generate_model_snapshot",
        ):
            payload = _build_model_snapshot_payload(
                input_model,
                st.session_state["assumptions"],
                pnl_result,
                cashflow_result,
                balance_sheet,
                debt_schedule,
                investment_result,
            )
            st.session_state["model_snapshot_payload"] = payload
            st.session_state["model_snapshot_text"] = json.dumps(
                payload, indent=2
            )
            st.session_state["model_snapshot_zip"] = _build_model_snapshot_zip(
                payload
            )

        if st.session_state.get("model_snapshot_payload"):
            st.download_button(
                "Download Snapshot ZIP",
                data=st.session_state["model_snapshot_zip"],
                file_name="model_snapshot.zip",
                mime="application/zip",
            )
            if st.button(
                "Copy GPT Prompt",
                key="copy_gpt_prompt",
            ):
                st.session_state["model_snapshot_prompt"] = (
                    "Here is the full model snapshot. "
                    "Use this as ground truth."
                )
            if st.session_state.get("model_snapshot_prompt"):
                prompt_text = st.session_state["model_snapshot_prompt"]
                st.text_area(
                    "GPT Prompt",
                    value=prompt_text,
                    height=80,
                )
                st.components.v1.html(
                    f"""
                    <script>
                      navigator.clipboard.writeText({prompt_text!r});
                    </script>
                    """,
                    height=0,
                )
            st.text_area(
                "Snapshot (JSON)",
                value=st.session_state.get("model_snapshot_text", ""),
                height=420,
            )

        st.markdown("### Model Transparency")
        st.markdown("**Revenue Logic**")
        st.write(
            "Revenue follows the Revenue Model bridge, with the final revenue "
            "equal to the maximum of the guaranteed floor and modeled revenue."
        )
        st.markdown("**Cost Logic**")
        st.write(
            "Operating costs are sourced from the Cost Model and aggregated into "
            "personnel, fixed overhead, and variable costs."
        )
        st.markdown("**Financing Logic**")
        st.write(
            "Debt service equals interest on opening debt plus scheduled repayment "
            "over the amortisation period."
        )
        st.markdown("**Equity Logic**")
        st.write(
            "Investor exits in the selected year at the exit multiple. Sponsor "
            "retains residual equity thereafter."
        )

        st.markdown("### Model State")
        st.write(
            "Inputs are stored in Streamlit session_state during the session."
        )
        st.write("Values do not persist across a full browser refresh.")
        st.write("Revenue and cost planning live in dedicated model pages.")

        st.markdown("### Export Status")
        st.write("Excel export: Enabled (Beta)")
        st.write(
            "Sheets included: Assumptions, Revenue Model, Cost Model, P&L, "
            "Cashflow, Balance Sheet, Valuation, Financing, Equity"
        )

        with st.expander("Model Controls", expanded=False):
            reset_confirm = st.checkbox(
                "I understand this will reset model state",
                key="reset_confirm",
            )
            if st.button("Reset Model", disabled=not reset_confirm):
                st.session_state.clear()
                st.rerun()
            if st.button("Reset Scenario", disabled=not reset_confirm):
                st.session_state["scenario_selection.selected_scenario"] = (
                    base_model.scenario_selection["selected_scenario"].value
                )
                st.rerun()
            if st.button("Clear Session State", disabled=not reset_confirm):
                st.session_state.clear()
                st.rerun()

    if page == "Valuation & Purchase Price":
        st.title("Valuation & Purchase Price")
        st.write(
            "This page compares seller expectations with a conservative buyer view. "
            "The buyer view focuses on cash generation, financing constraints and downside risk."
        )

        valuation_assumptions = _default_valuation_assumptions(input_model)
        seller_multiple = st.session_state.get(
            "valuation.seller_ebit_multiple",
            valuation_assumptions["seller_ebit_multiple"],
        )
        reference_year = st.session_state.get(
            "valuation.reference_year",
            valuation_assumptions["reference_year"],
        )
        buyer_discount_rate = st.session_state.get(
            "valuation.buyer_discount_rate",
            valuation_assumptions["buyer_discount_rate"],
        )
        valuation_start_year = st.session_state.get(
            "valuation.valuation_start_year",
            valuation_assumptions["valuation_start_year"],
        )
        transaction_cost_pct = st.session_state.get(
            "valuation.transaction_cost_pct",
            valuation_assumptions["transaction_cost_pct"],
        )
        include_terminal_value = st.session_state.get(
            "valuation.include_terminal_value",
            valuation_assumptions["include_terminal_value"],
        )

        pnl_table = pd.DataFrame.from_dict(pnl_result, orient="index")
        ebitda_ref = pnl_table.loc[f"Year {reference_year}", "ebitda"]
        seller_ev = ebitda_ref * seller_multiple
        if not debt_schedule:
            st.error("Debt schedule is missing. Cannot reconcile valuation.")
            st.stop()
        debt_year0 = debt_schedule[0]
        net_debt_close = debt_year0.get("closing_debt", 0.0)
        opening_debt_y0 = debt_year0.get("opening_debt", 0.0)
        drawdown_y0 = debt_year0.get("debt_drawdown", 0.0)
        repayment_y0 = debt_year0.get("total_repayment", 0.0)

        senior_debt_amount = input_model.financing_assumptions["senior_debt_amount"]
        amort_years = input_model.financing_assumptions.get(
            "amortization_period_years", 5
        )
        amort_type = input_model.financing_assumptions.get(
            "amortization_type", "Linear"
        )
        grace_years = input_model.financing_assumptions.get(
            "grace_period_years", 0
        )
        expected_repayment = None
        first_repay_row = debt_year0
        if amort_type == "Linear" and grace_years == 0:
            first_repay_row = next(
                (row for row in debt_schedule if row.get("opening_debt", 0.0) > 0),
                debt_year0,
            )
            expected_repayment = (
                first_repay_row.get("opening_debt", 0.0) / amort_years
                if amort_years
                else 0.0
            )
        debt_errors = []
        if abs(opening_debt_y0) > 1e-6:
            debt_errors.append("Opening Debt in Year 0 must be 0.")
        if abs(drawdown_y0 - senior_debt_amount) > 1e-6:
            debt_errors.append(
                "Debt drawdown in Year 0 must equal the Senior Debt Amount."
            )
        actual_repayment = (
            first_repay_row.get("scheduled_repayment", 0.0)
            if expected_repayment is not None
            else repayment_y0
        )
        if (
            expected_repayment is not None
            and abs(actual_repayment - expected_repayment) > 1e-6
        ):
            debt_errors.append(
                "Linear amortisation repayment does not match Senior Debt / Amortisation Years."
            )
        if debt_errors:
            st.error(
                "Debt schedule is not consistent with financing assumptions: "
                + " ".join(debt_errors)
            )
            st.stop()
        debt_service_values = [
            row.get("interest_expense", 0.0) + row.get("total_repayment", 0.0)
            for row in debt_schedule
        ]
        if senior_debt_amount > 0 and all(
            abs(value) < 1e-6 for value in debt_service_values
        ):
            st.error(
                "Debt service is zero across the schedule. Financing is not linked to Senior Debt."
            )
            st.stop()
        dscr_values = []
        for row in cashflow_result:
            year = row.get("year")
            debt_row = next(
                (item for item in debt_schedule if item.get("year") == year), None
            )
            if not debt_row:
                continue
            debt_service = debt_row.get("interest_expense", 0.0) + debt_row.get(
                "total_repayment", 0.0
            )
            if debt_service <= 0:
                continue
            cfads = (
                row.get("ebitda", 0.0)
                - row.get("taxes_paid", 0.0)
                - row.get("capex", 0.0)
                - row.get("working_capital_change", 0.0)
            )
            dscr_values.append(cfads / debt_service if debt_service else 0.0)
        if senior_debt_amount > 0 and not dscr_values:
            st.error(
                "DSCR cannot be derived from current cashflow and debt schedule."
            )
            st.stop()
        seller_equity_value = seller_ev - net_debt_close

        cashflow_table = pd.DataFrame(cashflow_result)
        free_cashflows = cashflow_table["free_cashflow"].tolist()
        cumulative_pv = 0.0
        for year_index, fcf in enumerate(free_cashflows):
            if year_index >= valuation_start_year:
                exponent = year_index - valuation_start_year + 1
                discount_factor = (
                    1 / ((1 + buyer_discount_rate) ** exponent)
                    if buyer_discount_rate
                    else 1.0
                )
            else:
                discount_factor = 0.0
            cumulative_pv += fcf * discount_factor

        terminal_value = 0.0
        terminal_pv = 0.0
        if include_terminal_value and buyer_discount_rate:
            terminal_value = free_cashflows[-1] / buyer_discount_rate
            last_exponent = max(1, len(free_cashflows) - valuation_start_year)
            terminal_pv = terminal_value / (
                (1 + buyer_discount_rate) ** last_exponent
            )

        enterprise_value_dcf = cumulative_pv + terminal_pv
        transaction_costs = enterprise_value_dcf * transaction_cost_pct
        buyer_equity_value = (
            enterprise_value_dcf - net_debt_close - transaction_costs
        )

        valuation_gap = buyer_equity_value - seller_equity_value
        valuation_gap_pct = (
            valuation_gap / seller_equity_value
            if seller_equity_value
            else 0
        )

        metric_cols = st.columns(3)
        metric_cols[0].metric(
            "Seller Equity Value", format_currency(seller_equity_value)
        )
        metric_cols[1].metric(
            "Buyer Affordability (Equity Value after financing)",
            format_currency(buyer_equity_value),
        )
        metric_cols[2].metric(
            "Gap (EUR / %)",
            f"{format_currency(valuation_gap)} | {format_pct(valuation_gap_pct)}",
        )
        st.info(
            "Net Debt at Close is taken from the Debt Schedule (Year 0): "
            "Closing Debt."
        )
        net_debt_table = pd.DataFrame(
            [
                {"Metric": "Opening Debt (Year 0)", "Value": format_currency(opening_debt_y0)},
                {"Metric": "Drawdown (Year 0)", "Value": format_currency(drawdown_y0)},
                {"Metric": "Repayment (Year 0)", "Value": format_currency(repayment_y0)},
                {"Metric": "Closing Debt (Year 0)", "Value": format_currency(net_debt_close)},
                {
                    "Metric": "Net Debt at Close",
                    "Value": format_currency(net_debt_close),
                },
            ]
        )
        st.dataframe(net_debt_table, use_container_width=True)
        if net_debt_close < 0:
            st.caption("Net Debt at Close is negative (Net Cash Position).")

        with st.expander("Seller Valuation (Multiple-Based)", expanded=False):
            st.write("Seller expectation based on EBITDA multiple.")
            seller_summary_table = pd.DataFrame(
                [
                    {
                        "Metric": "Reference EBITDA (Year)",
                        "Value": f"Year {reference_year}",
                    },
                    {
                        "Metric": "Applied EBITDA Multiple",
                        "Value": f"{seller_multiple:.2f}x",
                    },
                    {
                        "Metric": "Enterprise Value (EV)",
                        "Value": format_currency(seller_ev),
                    },
                    {
                        "Metric": "Net Debt at Close",
                        "Value": format_currency(net_debt_close),
                    },
                    {
                        "Metric": "Equity Value (Seller View)",
                        "Value": format_currency(seller_equity_value),
                    },
                ]
            )
            st.dataframe(seller_summary_table, use_container_width=True)
            with st.expander("Sensitivity (optional): +/- 1.0x multiple", expanded=False):
                seller_multiple_low = max(seller_multiple - 1.0, 0)
                seller_multiple_high = seller_multiple + 1.0
                seller_ev_low = ebitda_ref * seller_multiple_low
                seller_ev_high = ebitda_ref * seller_multiple_high
                seller_equity_low = seller_ev_low - net_debt_close
                seller_equity_high = seller_ev_high - net_debt_close
                sensitivity_table = pd.DataFrame(
                    [
                        {
                            "Multiple": f"{seller_multiple_low:.2f}x",
                            "Equity Value": format_currency(seller_equity_low),
                        },
                        {
                            "Multiple": f"{seller_multiple:.2f}x",
                            "Equity Value": format_currency(seller_equity_value),
                        },
                        {
                            "Multiple": f"{seller_multiple_high:.2f}x",
                            "Equity Value": format_currency(seller_equity_high),
                        },
                    ]
                )
                st.dataframe(sensitivity_table, use_container_width=True)

        with st.expander("Buyer Valuation (Cash-Based)", expanded=True):
            st.write(
                "Buyer view is based on discounted free cashflow and explicitly "
                "allows for negative equity values if pricing is too high."
            )
            st.caption("Note: No terminal value included (conservative downside view).")

            dcf_rows = {
                "Free Cashflow": {},
                "Discount Factor": {},
                "Present Value of FCF": {},
                "Cumulative PV of FCF": {},
                "Terminal Value": {},
                "PV of 5Y FCF (no terminal)": {},
                "Net Debt at Close": {},
                "Transaction Costs": {},
                "Equity Value (Buyer View)": {},
            }
            cumulative_pv = 0.0
            for year_index, fcf in enumerate(free_cashflows):
                year_label = f"Year {year_index}"
                if year_index >= valuation_start_year:
                    exponent = year_index - valuation_start_year + 1
                    discount_factor = (
                        1 / ((1 + buyer_discount_rate) ** exponent)
                        if buyer_discount_rate
                        else 1.0
                    )
                else:
                    discount_factor = 0.0
                pv_fcf = fcf * discount_factor
                cumulative_pv += pv_fcf

                dcf_rows["Free Cashflow"][year_label] = fcf
                dcf_rows["Discount Factor"][year_label] = discount_factor
                dcf_rows["Present Value of FCF"][year_label] = pv_fcf
                dcf_rows["Cumulative PV of FCF"][year_label] = cumulative_pv

            if include_terminal_value and buyer_discount_rate:
                dcf_rows["Terminal Value"]["Year 4"] = terminal_value

            for year_index in range(5):
                year_label = f"Year {year_index}"
                dcf_rows["PV of 5Y FCF (no terminal)"][year_label] = (
                    enterprise_value_dcf if year_index == 4 else ""
                )
                dcf_rows["Net Debt at Close"][year_label] = (
                    net_debt_close if year_index == 4 else ""
                )
                dcf_rows["Transaction Costs"][year_label] = (
                    transaction_costs if year_index == 4 else ""
                )
                dcf_rows["Equity Value (Buyer View)"][year_label] = (
                    buyer_equity_value if year_index == 4 else ""
                )

            dcf_table = pd.DataFrame.from_dict(dcf_rows, orient="index")
            dcf_table = dcf_table[[f"Year {i}" for i in range(5)]].reset_index()
            dcf_table.rename(columns={"index": "Line Item"}, inplace=True)
            dcf_total_rows = {
                "PV of 5Y FCF (no terminal)",
                "Equity Value (Buyer View)",
            }
            dcf_formatters = {
                "Discount Factor": lambda value: f"{value:.2f}"
                if value not in ("", None)
                else "",
            }
            _render_custom_table_html(
                dcf_table, set(), dcf_total_rows, dcf_formatters
            )

        st.markdown("### Purchase Price Logic")
        st.write(
            "Affordability (Equity Value after financing) is the equity value that does not "
            "break liquidity or financing constraints."
        )
        st.caption("If < 0: price must be reduced to 0 and/or seller support required.")
        max_price_table = pd.DataFrame(
            [
                {
                    "Metric": "Affordability (Equity Value after financing)",
                    "Value": format_currency(buyer_equity_value),
                }
            ]
        )
        st.dataframe(max_price_table, use_container_width=True)

        st.markdown("### Purchase Price Bridge")
        bridge_rows = {
            "Seller Equity Value": {"Year 0": seller_equity_value},
            "Buyer Equity Value": {"Year 0": buyer_equity_value},
            "Net Debt at Close": {"Year 0": net_debt_close},
            "Valuation Gap (EUR)": {"Year 0": valuation_gap},
            "Valuation Gap (%)": {"Year 0": valuation_gap_pct},
        }
        for year_index in range(1, 5):
            year_label = f"Year {year_index}"
            for key in bridge_rows:
                bridge_rows[key][year_label] = ""
        bridge_table = pd.DataFrame.from_dict(bridge_rows, orient="index")
        bridge_table = bridge_table[[f"Year {i}" for i in range(5)]].reset_index()
        bridge_table.rename(columns={"index": "Line Item"}, inplace=True)
        bridge_formatters = {
            "Valuation Gap (%)": format_pct,
        }
        _render_custom_table_html(
            bridge_table, set(), {"Valuation Gap (EUR)"}, bridge_formatters
        )
        gap_label = "Buyer > Seller" if valuation_gap >= 0 else "Buyer < Seller"
        st.caption(f"Valuation gap indicator: {gap_label}.")

        st.markdown("### Buyer View (Active Scenario)")
        buyer_view_table = pd.DataFrame(
            [
                {
                    "Metric": "Buyer Equity Value (DCF)",
                    "Value": format_currency(buyer_equity_value),
                }
            ]
        )
        st.dataframe(buyer_view_table, use_container_width=True)
        with st.expander(
            "Sensitivity (optional): discount rate +/- 1.0%", expanded=False
        ):
            buyer_discount_low = max(buyer_discount_rate - 0.01, 0)
            buyer_discount_high = buyer_discount_rate + 0.01
            sensitivity_rows = []
            for rate in [buyer_discount_low, buyer_discount_rate, buyer_discount_high]:
                cumulative_pv = 0.0
                for year_index, fcf in enumerate(free_cashflows):
                    if year_index >= valuation_start_year:
                        exponent = year_index - valuation_start_year + 1
                        discount_factor = (
                            1 / ((1 + rate) ** exponent)
                            if rate
                            else 1.0
                        )
                    else:
                        discount_factor = 0.0
                    cumulative_pv += fcf * discount_factor
                terminal_value = 0.0
                terminal_pv = 0.0
                if include_terminal_value and rate:
                    terminal_value = free_cashflows[-1] / rate
                    last_exponent = max(
                        1, len(free_cashflows) - valuation_start_year
                    )
                    terminal_pv = terminal_value / ((1 + rate) ** last_exponent)
                enterprise_value = cumulative_pv + terminal_pv
                transaction_costs_sens = enterprise_value * transaction_cost_pct
                equity_value = (
                    enterprise_value - net_debt_close - transaction_costs_sens
                )
                sensitivity_rows.append(
                    {
                        "Discount Rate": format_pct(rate),
                        "Equity Value": format_currency(equity_value),
                    }
                )
            st.dataframe(pd.DataFrame(sensitivity_rows), use_container_width=True)

        st.markdown("### Decision KPIs")
        purchase_price = input_model.transaction_and_financing[
            "purchase_price_eur"
        ].value
        year0_revenue = pnl_table.loc["Year 0", "revenue"]
        implied_ev_multiple = (
            seller_ev / ebitda_ref if ebitda_ref else 0
        )
        purchase_price_pct_revenue = (
            purchase_price / year0_revenue if year0_revenue else 0
        )
        kpi_table = pd.DataFrame(
            [
                {
                    "KPI": "Seller EV / EBITDA",
                    "Value": f"{implied_ev_multiple:.2f}x",
                },
                {
                    "KPI": "Buyer IRR at Seller Price",
                    "Value": format_pct(investment_result["irr"]),
                },
                {
                "KPI": "Affordability (Equity Value after financing)",
                "Value": format_currency(buyer_equity_value),
                },
                {
                    "KPI": "Headroom vs Seller Ask",
                    "Value": format_currency(valuation_gap),
                },
            ]
        )
        st.dataframe(kpi_table, use_container_width=True)

        explain_valuation = st.toggle("Explain valuation logic")
        if explain_valuation:
            st.write(
                "Seller and buyer values differ because the seller anchors on market "
                "multiples while the buyer focuses on cash generation and financing "
                "constraints."
            )
            st.write(
                "The buyer valuation is conservative by design to avoid overstretching "
                "liquidity or debt service capacity."
            )
            st.write(
                "This gap supports disciplined purchase price negotiation and shows "
                "where the buyer can rationally walk away."
            )
            st.write(
                "Not modeled: growth premiums, market synergies, or optimistic exit "
                "assumptions."
            )
            st.caption(
                f"Reference Year EBITDA (Year {reference_year}) = {format_currency(ebitda_ref)}."
            )
            st.caption(
                f"Enterprise Value = EBITDA × Multiple = {format_currency(ebitda_ref)} "
                f"× {seller_multiple:.2f}x = {format_currency(seller_ev)}."
            )
            st.caption(
                f"Equity Value = EV - Net Debt = {format_currency(seller_ev)} "
                f"- {format_currency(net_debt_close)} = {format_currency(seller_equity_value)}."
            )

            st.markdown("### Buyer Perspective")
            st.write(
                "The buyer view discounts free cashflows from the valuation start year "
                "at the required return."
            )
            st.caption(
                f"Discount Rate = {format_pct(buyer_discount_rate)}; "
                f"Valuation Start Year = Year {valuation_start_year}."
            )
            st.caption(
                "Present Value of FCF = FCF × Discount Factor; "
                "PV of 5Y FCF (no terminal) is the sum of PVs."
            )
            if include_terminal_value:
                st.caption(
                    f"Terminal Value = Final Year FCF / Discount Rate = "
                    f"{format_currency(free_cashflows[-1])} / "
                    f"{format_pct(buyer_discount_rate)}."
                )
            st.caption(
                f"Equity Value (Buyer) = EV - Net Debt at Close - Transaction Costs "
                f"= {format_currency(enterprise_value_dcf)} - "
                f"{format_currency(net_debt_close)} - "
                f"{format_currency(transaction_costs)}."
            )
            st.markdown("### Buyer vs. Seller Gap")
            st.write(
                "The valuation gap highlights the difference between seller expectations "
                "and buyer affordability after financing and transaction costs."
            )

    if page == "Operating Model (P&L)":
        st.title("Operating Model (P&L)")
        selected_scenario = st.session_state["output_scenario"]
        scenario_key = selected_scenario.lower()
        utilization_by_year = getattr(input_model, "utilization_by_year", None)
        if not isinstance(utilization_by_year, list) or len(utilization_by_year) < 5:
            base_util = input_model.scenario_parameters["utilization_rate"][
                scenario_key
            ].value
            utilization_by_year = [base_util] * 5

        utilization_field = _get_field_by_path(
            input_model.__dict__,
            ["scenario_parameters", "utilization_rate", scenario_key],
        )
        day_rate_field = _get_field_by_path(
            input_model.__dict__,
            ["scenario_parameters", "day_rate_eur", scenario_key],
        )
        fte_field = _get_field_by_path(
            input_model.__dict__,
            ["operating_assumptions", "consulting_fte_start"],
        )
        fte_growth_field = _get_field_by_path(
            input_model.__dict__,
            ["operating_assumptions", "consulting_fte_growth_pct"],
        )
        work_days_field = _get_field_by_path(
            input_model.__dict__,
            ["operating_assumptions", "work_days_per_year"],
        )
        day_rate_growth_field = _get_field_by_path(
            input_model.__dict__,
            ["operating_assumptions", "day_rate_growth_pct"],
        )
        guarantee_y1_field = _get_field_by_path(
            input_model.__dict__,
            ["operating_assumptions", "revenue_guarantee_pct_year_1"],
        )
        guarantee_y2_field = _get_field_by_path(
            input_model.__dict__,
            ["operating_assumptions", "revenue_guarantee_pct_year_2"],
        )
        guarantee_y3_field = _get_field_by_path(
            input_model.__dict__,
            ["operating_assumptions", "revenue_guarantee_pct_year_3"],
        )

        st.markdown("### P&L (GuV)")
        pnl_table = pd.DataFrame.from_dict(pnl_result, orient="index")
        year_indexes = list(range(len(pnl_table)))

        wage_inflation = input_model.personnel_cost_assumptions[
            "wage_inflation_pct"
        ].value
        consultant_base_cost = input_model.personnel_cost_assumptions[
            "avg_consultant_base_cost_eur_per_year"
        ].value
        bonus_pct = input_model.personnel_cost_assumptions[
            "bonus_pct_of_base"
        ].value
        payroll_pct = input_model.personnel_cost_assumptions[
            "payroll_burden_pct_of_comp"
        ].value
        management_cost = getattr(
            input_model, "management_md_cost_eur_per_year", 0.0
        )
        management_growth = getattr(
            input_model, "management_md_cost_growth_pct", 0.0
        )
        backoffice_fte_start = input_model.operating_assumptions[
            "backoffice_fte_start"
        ].value
        backoffice_growth = input_model.operating_assumptions[
            "backoffice_fte_growth_pct"
        ].value
        backoffice_salary = input_model.operating_assumptions[
            "avg_backoffice_salary_eur_per_year"
        ].value
        overhead_inflation = input_model.overhead_and_variable_costs[
            "overhead_inflation_pct"
        ].value
        depreciation = input_model.capex_and_working_capital[
            "depreciation_eur_per_year"
        ].value

        interest_by_year = {
            row["year"]: row["interest_expense"]
            for row in debt_schedule
        }

        line_items = {}
        def _set_line_value(name, year_label, value):
            if name not in line_items:
                line_items[name] = {
                    "Line Item": name,
                    "Year 0": "",
                    "Year 1": "",
                    "Year 2": "",
                    "Year 3": "",
                    "Year 4": "",
                }
            line_items[name][year_label] = value

        for year_index in year_indexes:
            year_label = f"Year {year_index}"
            total_revenue = revenue_final_by_year[year_index]
            year_costs = cost_model_totals[year_index]
            consultant_comp = year_costs["consultant_costs"]
            backoffice_comp = year_costs["backoffice_costs"]
            management_comp = year_costs["management_costs"]
            total_personnel = year_costs["personnel_costs"]

            cost_state = st.session_state["assumptions"]["cost_model"]
            fixed_row = cost_state["fixed_overhead"][year_index]
            variable_row = cost_state["variable_costs"][year_index]
            external_advisors = _non_negative(fixed_row["Advisory"])
            it_cost = _non_negative(fixed_row["IT & Software"])
            office_cost = _non_negative(fixed_row["Office Rent"])
            other_services = (
                _non_negative(fixed_row["Legal"])
                + _non_negative(fixed_row["Services"])
                + _non_negative(fixed_row["Other Services"])
            )
            variable_total = 0.0
            for prefix in ["Training", "Travel", "Communication"]:
                cost_type = variable_row[f"{prefix} Type"]
                value = _non_negative(variable_row[f"{prefix} Value"])
                if cost_type == "%":
                    variable_total += total_revenue * value
                else:
                    variable_total += value
            other_services += variable_total

            total_operating = year_costs["overhead_and_variable_costs"]
            ebitda = pnl_table.iloc[year_index]["ebitda"]
            ebit = pnl_table.iloc[year_index]["ebit"]
            interest = interest_by_year.get(year_index, 0)
            ebt = ebit - interest
            taxes = pnl_table.iloc[year_index]["taxes"]
            net_income = pnl_table.iloc[year_index]["net_income"]

            _set_line_value("Total Revenue", year_label, total_revenue)
            _set_line_value(
                "Consultant Compensation", year_label, consultant_comp
            )
            _set_line_value(
                "Backoffice Compensation", year_label, backoffice_comp
            )
            _set_line_value(
                "Management / MD Compensation", year_label, management_comp
            )
            _set_line_value(
                "Total Personnel Costs", year_label, total_personnel
            )
            _set_line_value(
                "External Consulting / Advisors",
                year_label,
                external_advisors,
            )
            _set_line_value("IT", year_label, it_cost)
            _set_line_value("Office", year_label, office_cost)
            _set_line_value("Other Services", year_label, other_services)
            _set_line_value(
                "Total Operating Expenses", year_label, total_operating
            )
            _set_line_value("EBITDA", year_label, ebitda)
            _set_line_value("Depreciation", year_label, depreciation)
            _set_line_value("EBIT", year_label, ebit)
            _set_line_value("Interest Expense", year_label, interest)
            _set_line_value("EBT", year_label, ebt)
            _set_line_value("Taxes", year_label, taxes)
            _set_line_value("Net Income (Jahresueberschuss)", year_label, net_income)

            consultant_fte_kpi = _non_negative(
                cost_state["personnel"][year_index]["Consultant FTE"]
            )
            revenue_per_consultant = (
                total_revenue / consultant_fte_kpi if consultant_fte_kpi else 0
            )
            ebitda_margin = ebitda / total_revenue if total_revenue else 0
            ebit_margin = ebit / total_revenue if total_revenue else 0
            personnel_cost_ratio = (
                total_personnel / total_revenue if total_revenue else 0
            )
            guaranteed_pct = revenue_components_by_year[year_index][
                "share_guaranteed"
            ]
            net_margin = net_income / total_revenue if total_revenue else 0
            opex_ratio = total_operating / total_revenue if total_revenue else 0

            _set_line_value(
                "Revenue per Consultant",
                year_label,
                revenue_per_consultant,
            )
            _set_line_value("EBITDA Margin", year_label, ebitda_margin)
            _set_line_value("EBIT Margin", year_label, ebit_margin)
            _set_line_value("Personnel Cost Ratio", year_label, personnel_cost_ratio)
            _set_line_value("Guaranteed Revenue %", year_label, guaranteed_pct)
            _set_line_value("Net Margin", year_label, net_margin)
            _set_line_value("Opex Ratio", year_label, opex_ratio)

        row_order = [
            "Revenue",
            "Total Revenue",
            "Personnel Costs",
            "Consultant Compensation",
            "Backoffice Compensation",
            "Management / MD Compensation",
            "Total Personnel Costs",
            "Operating Expenses",
            "External Consulting / Advisors",
            "IT",
            "Office",
            "Other Services",
            "Total Operating Expenses",
            "EBITDA",
            "Depreciation",
            "EBIT",
            "Interest Expense",
            "EBT",
            "Taxes",
            "Net Income (Jahresueberschuss)",
            "KPI",
            "Revenue per Consultant",
            "EBITDA Margin",
            "EBIT Margin",
            "Personnel Cost Ratio",
            "Guaranteed Revenue %",
            "Net Margin",
            "Opex Ratio",
        ]

        label_rows = []
        for label in row_order:
            if label in ("Revenue", "Personnel Costs", "Operating Expenses", "KPI"):
                label_rows.append(
                    {
                        "Line Item": label,
                        "Year 0": "",
                        "Year 1": "",
                        "Year 2": "",
                        "Year 3": "",
                        "Year 4": "",
                    }
                )
            else:
                row = line_items.get(label)
                if row:
                    label_rows.append(row)
                else:
                    label_rows.append(
                        {
                            "Line Item": label,
                            "Year 0": "",
                            "Year 1": "",
                            "Year 2": "",
                            "Year 3": "",
                            "Year 4": "",
                        }
                    )

        pnl_statement = pd.DataFrame(label_rows)
        format_map = {
            "Year 0": format_currency,
            "Year 1": format_currency,
            "Year 2": format_currency,
            "Year 3": format_currency,
            "Year 4": format_currency,
        }
        bold_rows = {
            "Total Revenue",
            "Total Personnel Costs",
            "Total Operating Expenses",
            "EBITDA",
            "EBIT",
            "EBT",
            "Net Income (Jahresueberschuss)",
        }
        section_rows = {"Revenue", "Personnel Costs", "Operating Expenses", "KPI"}

        def _style_pnl(df):
            styles = pd.DataFrame("", index=df.index, columns=df.columns)
            for idx, row in df.iterrows():
                for col in df.columns:
                    align = "text-align: left;" if col == "Line Item" else "text-align: right;"
                    styles.at[idx, col] += align
                if row["Line Item"] in section_rows:
                    styles.loc[idx, :] += "font-weight: 700; font-size: 1.05rem; background-color: #f9fafb;"
                if row["Line Item"] in bold_rows:
                    styles.loc[idx, :] += "font-weight: 700; background-color: #f3f4f6; border-top: 1px solid #c7c7c7;"
            return styles

        _render_pnl_html(pnl_statement, section_rows, bold_rows)

        # Table rendered via HTML for full-width layout.

        explain_pnl = st.toggle("Explain P&L logic")
        if explain_pnl:
            def _format_currency_expl(value):
                formatted = format_currency(value)
                return formatted if formatted else "—"

            def _format_pct_expl(value):
                formatted = format_pct(value)
                return formatted if formatted else "—"

            def _format_int_expl(value):
                formatted = format_int(value)
                return formatted if formatted else "—"

            def _safe_calc(values, func):
                if any(value is None for value in values):
                    return None
                try:
                    return func(*values)
                except Exception:
                    return None

            year_labels = [f"Year {year_index}" for year_index in year_indexes]

            st.markdown("### Revenue Logic")
            st.write(
                "Revenue is sourced from the Revenue Model. The Group Revenue Floor "
                "is compared to modeled group revenue to determine group revenue "
                "after the floor check and the final revenue used in the P&L."
            )
            revenue_metrics = {
                "Group Revenue Floor": {},
                "Modeled Total Revenue": {},
                "Group Revenue (after Floor Check)": {},
                "Final Revenue": {},
            }
            for year_index, year_label in enumerate(year_labels):
                components = revenue_components_by_year[year_index]
                revenue_metrics["Group Revenue Floor"][year_label] = components[
                    "guaranteed_floor"
                ]
                revenue_metrics["Modeled Total Revenue"][year_label] = components[
                    "modeled_total_revenue"
                ]
                revenue_metrics["Group Revenue (after Floor Check)"][year_label] = components[
                    "guaranteed_group_revenue"
                ]
                revenue_metrics["Final Revenue"][year_label] = components[
                    "final_total"
                ]
            revenue_table = pd.DataFrame.from_dict(
                revenue_metrics, orient="index"
            )
            revenue_table = revenue_table[year_labels].applymap(
                _format_currency_expl
            )
            st.dataframe(revenue_table, use_container_width=True)

            st.markdown("### Personnel Costs Logic")
            st.write(
                "Consultant compensation uses the all-in cost per consultant with "
                "wage inflation applied annually. Backoffice costs follow the "
                "same inflation logic. Management cost is a fixed annual amount "
                "grown by the management growth rate."
            )

            personnel_metrics = {}
            missing_personnel_inputs = [
                name
                for name, value in [
                    ("Consulting FTE", fte_field.value),
                    ("FTE Growth %", fte_growth_field.value),
                    ("Consultant Base Cost", consultant_base_cost),
                    ("Wage Inflation %", wage_inflation),
                    ("Backoffice FTE", backoffice_fte_start),
                    ("Backoffice Growth %", backoffice_growth),
                    ("Backoffice Salary", backoffice_salary),
                    ("Management / MD Cost", management_cost),
                    ("Management / MD Growth", management_growth),
                ]
                if value is None
            ]

            for year_index, year_label in enumerate(year_labels):
                consultants_fte = _safe_calc(
                    [fte_field.value, fte_growth_field.value],
                    lambda fte, growth: fte * ((1 + growth) ** year_index),
                )
                consultant_cost_per_fte = _safe_calc(
                    [consultant_base_cost, wage_inflation],
                    lambda base, inflation: base
                    * ((1 + inflation) ** year_index),
                )
                consultant_comp = _safe_calc(
                    [consultants_fte, consultant_cost_per_fte],
                    lambda fte, cost: fte * cost,
                )
                backoffice_fte = _safe_calc(
                    [backoffice_fte_start, backoffice_growth],
                    lambda fte, growth: fte * ((1 + growth) ** year_index),
                )
                backoffice_cost_per_fte = _safe_calc(
                    [backoffice_salary, wage_inflation],
                    lambda salary, inflation: salary
                    * ((1 + inflation) ** year_index),
                )
                backoffice_comp = _safe_calc(
                    [backoffice_fte, backoffice_cost_per_fte],
                    lambda fte, cost: fte * cost,
                )
                management_comp = _safe_calc(
                    [management_cost, management_growth],
                    lambda cost, growth: cost * ((1 + growth) ** year_index),
                )
                total_personnel = _safe_calc(
                    [consultant_comp, backoffice_comp, management_comp],
                    lambda a, b, c: a + b + c,
                )

                for metric, value in (
                    ("Consultant Compensation", consultant_comp),
                    ("Backoffice Compensation", backoffice_comp),
                    ("Management / MD Compensation", management_comp),
                    ("Total Personnel Costs", total_personnel),
                ):
                    if metric not in personnel_metrics:
                        personnel_metrics[metric] = {
                            year: None for year in year_labels
                        }
                    personnel_metrics[metric][year_label] = value

            personnel_table = pd.DataFrame.from_dict(
                personnel_metrics, orient="index"
            )
            personnel_table = personnel_table[year_labels].applymap(
                _format_currency_expl
            )
            st.dataframe(personnel_table, use_container_width=True)
            personnel_ratio_table = pd.DataFrame.from_dict(
                {"Personnel Cost Ratio": line_items["Personnel Cost Ratio"]},
                orient="index",
            )
            personnel_ratio_table = personnel_ratio_table[year_labels]
            personnel_ratio_table = personnel_ratio_table.applymap(
                _format_pct_expl
            )
            st.dataframe(personnel_ratio_table, use_container_width=True)
            if missing_personnel_inputs:
                st.caption(
                    f"Missing inputs: {', '.join(missing_personnel_inputs)}."
                )

            st.markdown("### Operating Expenses Logic")
            st.write(
                "Operating expenses are built from fixed annual costs inflated by "
                "the overhead inflation assumption."
            )

            opex_metrics = {}
            missing_opex_inputs = [
                name
                for name, value in [
                    ("External Advisors", input_model.overhead_and_variable_costs["legal_audit_eur_per_year"].value),
                    ("IT", input_model.overhead_and_variable_costs["it_and_software_eur_per_year"].value),
                    ("Office", input_model.overhead_and_variable_costs["rent_eur_per_year"].value),
                    ("Insurance", input_model.overhead_and_variable_costs["insurance_eur_per_year"].value),
                    ("Other Services", input_model.overhead_and_variable_costs["other_overhead_eur_per_year"].value),
                    ("Overhead Inflation %", overhead_inflation),
                ]
                if value is None
            ]

            for year_index, year_label in enumerate(year_labels):
                external_advisors = _safe_calc(
                    [
                        input_model.overhead_and_variable_costs[
                            "legal_audit_eur_per_year"
                        ].value,
                        overhead_inflation,
                    ],
                    lambda base, inflation: base * ((1 + inflation) ** year_index),
                )
                it_cost = _safe_calc(
                    [
                        input_model.overhead_and_variable_costs[
                            "it_and_software_eur_per_year"
                        ].value,
                        overhead_inflation,
                    ],
                    lambda base, inflation: base * ((1 + inflation) ** year_index),
                )
                office_cost = _safe_calc(
                    [
                        input_model.overhead_and_variable_costs[
                            "rent_eur_per_year"
                        ].value,
                        overhead_inflation,
                    ],
                    lambda base, inflation: base * ((1 + inflation) ** year_index),
                )
                other_services = _safe_calc(
                    [
                        input_model.overhead_and_variable_costs[
                            "insurance_eur_per_year"
                        ].value,
                        input_model.overhead_and_variable_costs[
                            "other_overhead_eur_per_year"
                        ].value,
                        overhead_inflation,
                    ],
                    lambda insurance, other, inflation: (insurance + other)
                    * ((1 + inflation) ** year_index),
                )
                total_opex = _safe_calc(
                    [external_advisors, it_cost, office_cost, other_services],
                    lambda a, b, c, d: a + b + c + d,
                )

                for metric, value in (
                    ("External Consulting / Advisors", external_advisors),
                    ("IT", it_cost),
                    ("Office", office_cost),
                    ("Other Services", other_services),
                    ("Total Operating Expenses", total_opex),
                ):
                    if metric not in opex_metrics:
                        opex_metrics[metric] = {
                            year: None for year in year_labels
                        }
                    opex_metrics[metric][year_label] = value

            opex_table = pd.DataFrame.from_dict(opex_metrics, orient="index")
            opex_table = opex_table[year_labels].applymap(
                _format_currency_expl
            )
            st.dataframe(opex_table, use_container_width=True)
            opex_ratio_table = pd.DataFrame.from_dict(
                {"Opex Ratio": line_items["Opex Ratio"]},
                orient="index",
            )
            opex_ratio_table = opex_ratio_table[year_labels]
            opex_ratio_table = opex_ratio_table.applymap(_format_pct_expl)
            st.dataframe(opex_ratio_table, use_container_width=True)
            if missing_opex_inputs:
                st.caption(
                    f"Missing inputs: {', '.join(missing_opex_inputs)}."
                )

            st.markdown("### Earnings Bridge")
            st.write(
                "EBITDA bridges revenue to operating costs. EBIT subtracts "
                "depreciation, then interest and taxes produce net income."
            )

            earnings_metrics = {}
            for year_index, year_label in enumerate(year_labels):
                revenue = line_items["Total Revenue"][year_label]
                total_personnel = line_items["Total Personnel Costs"][year_label]
                total_operating = line_items["Total Operating Expenses"][year_label]
                ebitda = line_items["EBITDA"][year_label]
                interest = line_items["Interest Expense"][year_label]
                taxes = line_items["Taxes"][year_label]
                net_income = line_items["Net Income (Jahresueberschuss)"][
                    year_label
                ]
                ebit = line_items["EBIT"][year_label]
                ebt = line_items["EBT"][year_label]

                for metric, value in (
                    ("Total Revenue", revenue),
                    ("Total Personnel Costs", total_personnel),
                    ("Total Operating Expenses", total_operating),
                    ("EBITDA", ebitda),
                    ("Depreciation", depreciation),
                    ("EBIT", ebit),
                    ("Interest Expense", interest),
                    ("EBT", ebt),
                    ("Taxes", taxes),
                    ("Net Income (Jahresueberschuss)", net_income),
                ):
                    if metric not in earnings_metrics:
                        earnings_metrics[metric] = {
                            year: None for year in year_labels
                        }
                    earnings_metrics[metric][year_label] = value

            earnings_table = pd.DataFrame.from_dict(
                earnings_metrics, orient="index"
            )
            earnings_table = earnings_table[year_labels].applymap(
                _format_currency_expl
            )
            st.dataframe(earnings_table, use_container_width=True)

            st.markdown("### KPI Definitions")
            st.write(
                "KPIs summarize profitability and efficiency using the P&L "
                "line items for each year."
            )

            kpi_metrics = {
                "Revenue per Consultant": {},
                "EBITDA Margin": {},
                "EBIT Margin": {},
                "Personnel Cost Ratio": {},
                "Guaranteed Revenue %": {},
                "Net Margin": {},
                "Opex Ratio": {},
            }
            for year_label in year_labels:
                kpi_metrics["Revenue per Consultant"][year_label] = line_items[
                    "Revenue per Consultant"
                ][year_label]
                kpi_metrics["EBITDA Margin"][year_label] = line_items[
                    "EBITDA Margin"
                ][year_label]
                kpi_metrics["EBIT Margin"][year_label] = line_items[
                    "EBIT Margin"
                ][year_label]
                kpi_metrics["Personnel Cost Ratio"][year_label] = line_items[
                    "Personnel Cost Ratio"
                ][year_label]
                kpi_metrics["Guaranteed Revenue %"][year_label] = line_items[
                    "Guaranteed Revenue %"
                ][year_label]
                kpi_metrics["Net Margin"][year_label] = line_items[
                    "Net Margin"
                ][year_label]
                kpi_metrics["Opex Ratio"][year_label] = line_items[
                    "Opex Ratio"
                ][year_label]

            kpi_table = pd.DataFrame.from_dict(kpi_metrics, orient="index")
            kpi_table = kpi_table[year_labels]
            percent_kpis = {
                "EBITDA Margin",
                "EBIT Margin",
                "Personnel Cost Ratio",
                "Guaranteed Revenue %",
                "Net Margin",
                "Opex Ratio",
            }
            for metric in kpi_table.index:
                formatter = (
                    _format_pct_expl
                    if metric in percent_kpis
                    else _format_currency_expl
                )
                kpi_table.loc[metric] = kpi_table.loc[metric].apply(formatter)
            st.dataframe(kpi_table, use_container_width=True)

        pnl_excel = _build_pnl_excel(
            input_model, pnl_result, cashflow_result, debt_schedule
        )
        st.download_button(
            "Download P&L as Excel",
            data=pnl_excel.getvalue(),
            file_name="Financial_Model_PnL.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if page == "Cashflow & Liquidity":
        st.title("Cashflow & Liquidity")
        st.write("Consolidated cashflow statement (5-year plan)")
        cashflow_line_items = {}

        def _set_cashflow_value(name, year_label, value):
            if name not in cashflow_line_items:
                cashflow_line_items[name] = {
                    "Line Item": name,
                    "Year 0": "",
                    "Year 1": "",
                    "Year 2": "",
                    "Year 3": "",
                    "Year 4": "",
                }
            cashflow_line_items[name][year_label] = value

        for row in cashflow_result:
            year_label = f"Year {row['year']}"
            _set_cashflow_value("EBITDA", year_label, row["ebitda"])
            _set_cashflow_value("Taxes Paid", year_label, row["taxes_paid"])
            _set_cashflow_value(
                "Working Capital Change",
                year_label,
                row["working_capital_change"],
            )
            _set_cashflow_value(
                "Operating Cashflow", year_label, row["operating_cf"]
            )
            _set_cashflow_value("Capex", year_label, row["capex"])
            _set_cashflow_value(
                "Free Cashflow", year_label, row["free_cashflow"]
            )
            _set_cashflow_value(
                "Debt Drawdown", year_label, row["debt_drawdown"]
            )
            _set_cashflow_value(
                "Interest Paid", year_label, row["interest_paid"]
            )
            _set_cashflow_value(
                "Debt Repayment", year_label, row["debt_repayment"]
            )
            _set_cashflow_value(
                "Net Cashflow", year_label, row["net_cashflow"]
            )
            _set_cashflow_value(
                "Opening Cash", year_label, row["opening_cash"]
            )
            _set_cashflow_value(
                "Closing Cash", year_label, row["cash_balance"]
            )

        cashflow_row_order = [
            "OPERATING CASHFLOW",
            "EBITDA",
            "Taxes Paid",
            "Working Capital Change",
            "Operating Cashflow",
            "INVESTING CASHFLOW",
            "Capex",
            "Free Cashflow",
            "FINANCING CASHFLOW",
            "Debt Drawdown",
            "Interest Paid",
            "Debt Repayment",
            "Net Cashflow",
            "LIQUIDITY",
            "Opening Cash",
            "Net Cashflow",
            "Closing Cash",
        ]

        cashflow_rows = []
        for label in cashflow_row_order:
            if label in (
                "OPERATING CASHFLOW",
                "INVESTING CASHFLOW",
                "FINANCING CASHFLOW",
                "LIQUIDITY",
            ):
                cashflow_rows.append(
                    {
                        "Line Item": label,
                        "Year 0": "",
                        "Year 1": "",
                        "Year 2": "",
                        "Year 3": "",
                        "Year 4": "",
                    }
                )
            else:
                cashflow_rows.append(cashflow_line_items.get(label))

        cashflow_statement = pd.DataFrame(cashflow_rows)
        cashflow_section_rows = {
            "OPERATING CASHFLOW",
            "INVESTING CASHFLOW",
            "FINANCING CASHFLOW",
            "LIQUIDITY",
        }
        cashflow_total_rows = {
            "Operating Cashflow",
            "Free Cashflow",
            "Net Cashflow",
            "Closing Cash",
        }
        _render_cashflow_html(
            cashflow_statement, cashflow_section_rows, cashflow_total_rows
        )

        cashflow_years = [f"Year {i}" for i in range(5)]
        kpi_metrics = {
            "Operating Cashflow": {},
            "Free Cashflow": {},
            "Cash Conversion": {},
        }
        for year_label in cashflow_years:
            ebitda = cashflow_line_items["EBITDA"][year_label]
            free_cf = cashflow_line_items["Free Cashflow"][year_label]
            operating_cf = cashflow_line_items["Operating Cashflow"][year_label]
            kpi_metrics["Operating Cashflow"][year_label] = operating_cf
            kpi_metrics["Free Cashflow"][year_label] = free_cf
            kpi_metrics["Cash Conversion"][year_label] = (
                free_cf / ebitda if ebitda else 0
            )

        kpi_table = pd.DataFrame.from_dict(kpi_metrics, orient="index")
        kpi_table = kpi_table[cashflow_years]
        for metric in kpi_table.index:
            formatter = (
                format_pct if metric == "Cash Conversion" else format_currency
            )
            kpi_table.loc[metric] = kpi_table.loc[metric].apply(formatter)
        st.markdown("### KPI Summary")
        st.dataframe(kpi_table, use_container_width=True)

        cash_balances = [
            row["cash_balance"] for row in cashflow_result
        ]
        min_cash = min(cash_balances) if cash_balances else 0
        max_cash = max(cash_balances) if cash_balances else 0
        negative_years = [
            f"Year {row['year']}"
            for row in cashflow_result
            if row["cash_balance"] < 0
        ]
        summary_table = pd.DataFrame(
            [
                {
                    "Metric": "Minimum Cash Balance",
                    "Value": format_currency(min_cash),
                },
                {
                    "Metric": "Cash Volatility (Max - Min)",
                    "Value": format_currency(max_cash - min_cash),
                },
                {
                    "Metric": "Years with Negative Cash",
                    "Value": ", ".join(negative_years) if negative_years else "None",
                },
            ]
        )
        st.dataframe(summary_table, use_container_width=True)

        explain_cashflow = st.toggle("Explain Cashflow Logic")
        if explain_cashflow:
            cashflow_assumptions = input_model.cashflow_assumptions
            st.markdown("### Operating Cashflow Logic")
            st.write(
                "Operating cashflow starts from EBITDA and adjusts for cash "
                "taxes and the working capital timing proxy."
            )
            st.write(
                "Cash taxes differ from tax expense because they are based on "
                "EBT and can be paid with a lag."
            )
            operating_table = pd.DataFrame.from_dict(
                {
                    "EBITDA": cashflow_line_items["EBITDA"],
                    "Taxes Paid": cashflow_line_items["Taxes Paid"],
                    "Working Capital Change": cashflow_line_items[
                        "Working Capital Change"
                    ],
                    "Operating Cashflow": cashflow_line_items[
                        "Operating Cashflow"
                    ],
                },
                orient="index",
            )
            operating_table = operating_table[cashflow_years].applymap(
                format_currency
            )
            st.dataframe(operating_table, use_container_width=True)
            st.caption(
                "Operating Cashflow = EBITDA - Taxes Paid - Working Capital Change."
            )
            st.caption(
                "Taxes Paid = max(EBT, 0) × "
                f"{format_pct(cashflow_assumptions['tax_cash_rate_pct'])} "
                f"with a {cashflow_assumptions['tax_payment_lag_years']}-year lag."
            )
            st.caption(
                "Working Capital Change = Revenue × "
                f"{format_pct(cashflow_assumptions['working_capital_pct_revenue'])}."
            )

            st.markdown("### Investing Cashflow Logic")
            st.write(
                "Capex is modeled as a stable percentage of revenue, which "
                "is typical for consulting businesses with limited fixed assets."
            )
            investing_table = pd.DataFrame.from_dict(
                {
                    "Capex": cashflow_line_items["Capex"],
                    "Free Cashflow": cashflow_line_items["Free Cashflow"],
                },
                orient="index",
            )
            investing_table = investing_table[cashflow_years].applymap(
                format_currency
            )
            st.dataframe(investing_table, use_container_width=True)
            st.caption("Free Cashflow = Operating Cashflow - Capex.")
            st.caption(
                "Capex = Revenue × "
                f"{format_pct(cashflow_assumptions['capex_pct_revenue'])}."
            )

            st.markdown("### Financing Cashflow Logic")
            st.write(
                "Financing cashflow reflects initial debt funding and annual "
                "debt service. Interest paid is based on outstanding principal."
            )
            financing_table = pd.DataFrame.from_dict(
                {
                    "Debt Drawdown": cashflow_line_items["Debt Drawdown"],
                    "Interest Paid": cashflow_line_items["Interest Paid"],
                    "Debt Repayment": cashflow_line_items["Debt Repayment"],
                    "Net Cashflow": cashflow_line_items["Net Cashflow"],
                },
                orient="index",
            )
            financing_table = financing_table[cashflow_years].applymap(
                format_currency
            )
            st.dataframe(financing_table, use_container_width=True)
            st.caption(
                "Net Cashflow = Free Cashflow + Financing Cashflow."
            )

            st.markdown("### Liquidity Logic")
            st.write(
                "Closing cash is the opening balance plus net cashflow, "
                "highlighting years with potential liquidity pressure."
            )
            liquidity_table = pd.DataFrame.from_dict(
                {
                    "Opening Cash": cashflow_line_items["Opening Cash"],
                    "Net Cashflow": cashflow_line_items["Net Cashflow"],
                    "Closing Cash": cashflow_line_items["Closing Cash"],
                },
                orient="index",
            )
            liquidity_table = liquidity_table[cashflow_years].applymap(
                format_currency
            )
            st.dataframe(liquidity_table, use_container_width=True)

    if page == "Balance Sheet":
        st.title("Balance Sheet")
        st.write("Simplified balance sheet (5-year plan)")
        balance_line_items = {}

        def _set_balance_value(name, year_label, value):
            if name not in balance_line_items:
                balance_line_items[name] = {
                    "Line Item": name,
                    "Year 0": "",
                    "Year 1": "",
                    "Year 2": "",
                    "Year 3": "",
                    "Year 4": "",
                }
            balance_line_items[name][year_label] = value

        for row in balance_sheet:
            year_label = f"Year {row['year']}"
            _set_balance_value("Cash", year_label, row["cash"])
            _set_balance_value(
                "Fixed Assets (Net)", year_label, row["fixed_assets"]
            )
            _set_balance_value(
                "Total Assets", year_label, row["total_assets"]
            )
            _set_balance_value(
                "Financial Debt", year_label, row["financial_debt"]
            )
            _set_balance_value(
                "Total Liabilities", year_label, row["total_liabilities"]
            )
            _set_balance_value(
                "Equity at Start of Year", year_label, row["equity_start"]
            )
            _set_balance_value("Net Income", year_label, row["net_income"])
            _set_balance_value("Dividends", year_label, row["dividends"])
            _set_balance_value(
                "Equity Injections", year_label, row.get("equity_injection", 0.0)
            )
            _set_balance_value(
                "Equity Buybacks / Exit Payouts",
                year_label,
                row.get("equity_buyback", 0.0),
            )
            _set_balance_value(
                "Equity at End of Year", year_label, row["equity_end"]
            )
            _set_balance_value(
                "Total Liabilities + Equity",
                year_label,
                row["total_liabilities_equity"],
            )
            _set_balance_value(
                "Total Assets",
                year_label,
                row["total_assets"],
            )

        balance_row_order = [
            "ASSETS",
            "Cash",
            "Fixed Assets (Net)",
            "Total Assets",
            "LIABILITIES",
            "Financial Debt",
            "Total Liabilities",
            "EQUITY",
            "Equity at Start of Year",
            "Net Income",
            "Dividends",
            "Equity Injections",
            "Equity Buybacks / Exit Payouts",
            "Equity at End of Year",
            "CHECK",
            "Total Assets",
            "Total Liabilities + Equity",
        ]

        balance_rows = []
        for label in balance_row_order:
            if label in ("ASSETS", "LIABILITIES", "EQUITY", "CHECK"):
                balance_rows.append(
                    {
                        "Line Item": label,
                        "Year 0": "",
                        "Year 1": "",
                        "Year 2": "",
                        "Year 3": "",
                        "Year 4": "",
                    }
                )
            else:
                row = balance_line_items.get(label)
                if row:
                    balance_rows.append(row)
                else:
                    balance_rows.append(
                        {
                            "Line Item": label,
                            "Year 0": "",
                            "Year 1": "",
                            "Year 2": "",
                            "Year 3": "",
                            "Year 4": "",
                        }
                    )

        balance_statement = pd.DataFrame(balance_rows)
        balance_section_rows = {"ASSETS", "LIABILITIES", "EQUITY", "CHECK"}
        balance_total_rows = {
            "Total Assets",
            "Total Liabilities",
            "Equity at End of Year",
            "Total Liabilities + Equity",
        }
        _render_balance_sheet_html(
            balance_statement, balance_section_rows, balance_total_rows
        )

        reconciliation_issues = [
            f"Year {row['year']}"
            for row in balance_sheet
            if abs(row["balance_check"]) > 1e-2
        ]
        if reconciliation_issues:
            st.caption(
                "Balance check (Assets - Liabilities - Equity) is not zero in "
                f"{', '.join(reconciliation_issues)}."
            )

        cashflow_years = [f"Year {i}" for i in range(5)]
        ebitda_by_year = pd.DataFrame.from_dict(
            pnl_result, orient="index"
        )["ebitda"].to_dict()
        kpi_metrics = {
            "Net Debt": {},
            "Equity Ratio": {},
            "Net Debt / EBITDA": {},
            "Minimum Cash Headroom": {},
        }
        min_cash = input_model.balance_sheet_assumptions[
            "minimum_cash_balance_eur"
        ]

        for row in balance_sheet:
            year_label = f"Year {row['year']}"
            net_debt = row["financial_debt"] - row["cash"]
            equity_ratio = (
                row["equity_end"] / row["total_assets"]
                if row["total_assets"]
                else 0
            )
            ebitda = ebitda_by_year.get(f"Year {row['year']}", 0)
            net_debt_to_ebitda = (
                net_debt / ebitda if ebitda else 0
            )
            cash_headroom = row["cash"] - min_cash

            kpi_metrics["Net Debt"][year_label] = net_debt
            kpi_metrics["Equity Ratio"][year_label] = equity_ratio
            kpi_metrics["Net Debt / EBITDA"][year_label] = net_debt_to_ebitda
            kpi_metrics["Minimum Cash Headroom"][year_label] = cash_headroom

        kpi_table = pd.DataFrame.from_dict(kpi_metrics, orient="index")
        kpi_table = kpi_table[cashflow_years]
        for metric in kpi_table.index:
            if metric in {"Equity Ratio"}:
                kpi_table.loc[metric] = kpi_table.loc[metric].apply(
                    format_pct
                )
            elif metric == "Net Debt / EBITDA":
                kpi_table.loc[metric] = kpi_table.loc[metric].apply(
                    lambda value: f"{value:.2f}x"
                )
            else:
                kpi_table.loc[metric] = kpi_table.loc[metric].apply(
                    format_currency
                )
        st.markdown("### KPI Summary")
        st.dataframe(kpi_table, use_container_width=True)

        explain_balance = st.toggle("Explain Balance Sheet Logic")
        if explain_balance:
            balance_assumptions = input_model.balance_sheet_assumptions
            st.markdown("### Balance Sheet Scope")
            st.write(
                "This balance sheet is intentionally simplified for a consulting "
                "carve-out. It focuses on cash, fixed assets, debt, and equity "
                "to support decision-making and bank discussions."
            )
            st.write(
                "Receivables, payables, and inventory are excluded because the "
                "model uses a working-capital proxy in cashflow."
            )

            st.markdown("### Asset Logic")
            asset_table = pd.DataFrame.from_dict(
                {
                    "Cash": balance_line_items["Cash"],
                    "Fixed Assets (Net)": balance_line_items[
                        "Fixed Assets (Net)"
                    ],
                    "Total Assets": balance_line_items["Total Assets"],
                },
                orient="index",
            )
            asset_table = asset_table[cashflow_years].applymap(
                format_currency
            )
            st.dataframe(asset_table, use_container_width=True)
            st.caption(
                "Cash is taken directly from the cashflow closing cash balance."
            )
            st.caption(
                "Fixed Assets end = Prior Fixed Assets + Capex - Depreciation."
            )
            st.caption(
                "Depreciation = (Fixed Assets + Capex) × "
                f"{format_pct(balance_assumptions['depreciation_rate_pct'])}."
            )

            st.markdown("### Debt Logic")
            debt_table = pd.DataFrame.from_dict(
                {
                    "Financial Debt": balance_line_items["Financial Debt"],
                    "Total Liabilities": balance_line_items["Total Liabilities"],
                },
                orient="index",
            )
            debt_table = debt_table[cashflow_years].applymap(format_currency)
            st.dataframe(debt_table, use_container_width=True)
            st.caption(
                "Financial debt follows the debt schedule from the financing model."
            )

            st.markdown("### Equity Logic")
            equity_table = pd.DataFrame.from_dict(
                {
                    "Equity at Start of Year": balance_line_items[
                        "Equity at Start of Year"
                    ],
                    "Net Income": balance_line_items["Net Income"],
                    "Dividends": balance_line_items["Dividends"],
                    "Equity Injections": balance_line_items["Equity Injections"],
                    "Equity Buybacks / Exit Payouts": balance_line_items[
                        "Equity Buybacks / Exit Payouts"
                    ],
                    "Equity at End of Year": balance_line_items[
                        "Equity at End of Year"
                    ],
                },
                orient="index",
            )
            equity_table = equity_table[cashflow_years].applymap(
                format_currency
            )
            st.dataframe(equity_table, use_container_width=True)
            st.caption(
                "Equity end = Equity start + Net Income - Dividends "
                "+ Equity Injections - Equity Buybacks."
            )

            st.markdown("### Reconciliation Check")
            check_table = pd.DataFrame.from_dict(
                {
                    "Total Assets": balance_line_items["Total Assets"],
                    "Total Liabilities + Equity": balance_line_items[
                        "Total Liabilities + Equity"
                    ],
                },
                orient="index",
            )
            check_table = check_table[cashflow_years].applymap(
                format_currency
            )
            st.dataframe(check_table, use_container_width=True)

    if page == "Financing & Debt":
        st.title("Financing & Debt")
        st.write("Debt structure, service and bankability (5-year plan)")
        financing_assumptions = input_model.financing_assumptions
        senior_debt_amount = financing_assumptions["senior_debt_amount"]
        amort_years = financing_assumptions["amortization_period_years"]
        amort_type = financing_assumptions.get("amortization_type", "Linear")
        grace_years = financing_assumptions.get("grace_period_years", 0)
        peak_debt = max(
            row["opening_debt"] + row["debt_drawdown"] for row in debt_schedule
        )
        if abs(peak_debt - senior_debt_amount) > 1.0:
            raise ValueError("Debt input not reflected in debt schedule.")
        if amort_type == "Linear" and grace_years == 0:
            first_repay_row = next(
                (row for row in debt_schedule if row.get("opening_debt", 0.0) > 0),
                debt_schedule[0],
            )
            scheduled_repayment = first_repay_row.get("scheduled_repayment", 0.0)
            if abs(scheduled_repayment * amort_years - first_repay_row.get("opening_debt", 0.0)) > 1.0:
                raise ValueError("Amortisation inconsistency.")
        cashflow_by_year = {row["year"]: row for row in cashflow_result}
        maintenance_capex_pct = financing_assumptions[
            "maintenance_capex_pct_revenue"
        ]

        bank_rows = []
        for year_index in range(5):
            year_label = f"Year {year_index}"
            ebitda = pnl_result[year_label]["ebitda"]
            cash_taxes = cashflow_by_year[year_index]["taxes_paid"]
            revenue = pnl_result[year_label]["revenue"]
            maintenance_capex = revenue * maintenance_capex_pct
            working_capital = cashflow_by_year[year_index][
                "working_capital_change"
            ]
            cfads = ebitda - cash_taxes - maintenance_capex + working_capital
            interest = debt_schedule[year_index]["interest_expense"]
            scheduled_repayment = debt_schedule[year_index][
                "scheduled_repayment"
            ]
            debt_service = interest + scheduled_repayment
            dscr = cfads / debt_service if debt_service else 0
            min_dscr = financing_assumptions["minimum_dscr"]
            dscr_display = -abs(dscr) if dscr < min_dscr else dscr
            breach = "YES" if dscr < min_dscr else "NO"

            bank_rows.append({"Line Item": "EBITDA", year_label: ebitda})
            bank_rows.append(
                {"Line Item": "Cash Taxes", year_label: cash_taxes}
            )
            bank_rows.append(
                {"Line Item": "Capex (Maintenance)", year_label: maintenance_capex}
            )
            bank_rows.append({"Line Item": "CFADS", year_label: cfads})
            bank_rows.append(
                {"Line Item": "Interest Expense", year_label: interest}
            )
            bank_rows.append(
                {
                    "Line Item": "Scheduled Repayment",
                    year_label: scheduled_repayment,
                }
            )
            bank_rows.append(
                {"Line Item": "Debt Service", year_label: debt_service}
            )
            bank_rows.append({"Line Item": "DSCR", year_label: dscr_display})
            bank_rows.append(
                {
                    "Line Item": "Minimum Required DSCR",
                    year_label: min_dscr,
                }
            )
            bank_rows.append(
                {"Line Item": "Covenant Breach", year_label: breach}
            )

        bank_metrics = {}
        for entry in bank_rows:
            label = entry["Line Item"]
            bank_metrics.setdefault(label, {})
            for year_index in range(5):
                year_label = f"Year {year_index}"
                if year_label in entry:
                    bank_metrics[label][year_label] = entry[year_label]

        bank_table = pd.DataFrame.from_dict(bank_metrics, orient="index")
        bank_table = bank_table[
            [f"Year {i}" for i in range(5)]
        ].reset_index()
        bank_table.rename(columns={"index": "Line Item"}, inplace=True)
        bank_formatters = {
            "DSCR": lambda value: f"{abs(value):.2f}x"
            if value not in ("", None)
            else "",
            "Minimum Required DSCR": lambda value: f"{value:.2f}x"
            if value not in ("", None)
            else "",
            "Covenant Breach": lambda value: value if value else "",
        }
        st.markdown("### Bank View")
        _render_custom_table_html(
            bank_table,
            set(),
            {"CFADS", "Debt Service"},
            bank_formatters,
        )
        st.caption(
            "CFADS = EBITDA - Cash Taxes - Maintenance Capex ± Working Capital Change."
        )
        st.caption(
            "DSCR = CFADS / (Interest Expense + Scheduled Repayment)."
        )
        st.caption(
            "Peak Debt may differ from initial drawdown when repayments occur within Year 0."
        )

        dscr_values = [
            abs(value)
            for value in bank_metrics.get("DSCR", {}).values()
            if isinstance(value, (int, float))
        ]
        avg_dscr = sum(dscr_values) / len(dscr_values) if dscr_values else 0
        min_dscr_value = min(dscr_values) if dscr_values else 0
        if any(value > 50 for value in dscr_values):
            st.warning(
                "DSCR exceeds 50x in at least one year. This may indicate broken debt logic."
            )
        debt_at_close = debt_schedule[0]["debt_drawdown"]
        ebitda_year0 = pnl_result["Year 0"]["ebitda"]
        debt_to_ebitda = debt_at_close / ebitda_year0 if ebitda_year0 else 0

        st.markdown("### KPIs")
        kpi_table = pd.DataFrame(
            [
                {"KPI": "Average DSCR", "Value": f"{avg_dscr:.2f}x"},
                {"KPI": "Minimum DSCR", "Value": f"{min_dscr_value:.2f}x"},
                {"KPI": "Peak Debt", "Value": format_currency(peak_debt)},
                {
                    "KPI": "Debt / EBITDA (at close)",
                    "Value": f"{debt_to_ebitda:.2f}x",
                },
            ]
        )
        st.dataframe(kpi_table, use_container_width=True)

        explain_financing = st.toggle("Explain Financing Logic")
        if explain_financing:
            st.markdown("### Explanation")
            st.write(
                "CFADS is the cash the business generates to service debt "
                "after cash taxes and maintenance capex."
            )
            critical_years = []
            for year_index in range(5):
                year_label = f"Year {year_index}"
                cfads = bank_metrics["CFADS"].get(year_label, 0)
                debt_service = bank_metrics["Debt Service"].get(year_label, 0)
                dscr_value = abs(bank_metrics["DSCR"].get(year_label, 0))
                min_dscr = bank_metrics["Minimum Required DSCR"].get(
                    year_label, 0
                )
                status = (
                    "above"
                    if dscr_value >= min_dscr
                    else "below"
                )
                st.write(
                    f"In {year_label}, the business generates "
                    f"{format_currency(cfads)} of cash available for debt service. "
                    f"Required debt service is {format_currency(debt_service)}, "
                    f"resulting in a DSCR of {dscr_value:.2f}x. "
                    f"This is {status} the required covenant of {min_dscr:.2f}x."
                )
                if dscr_value < min_dscr:
                    critical_years.append(year_label)

            if critical_years:
                st.write(
                    f"Critical years: {', '.join(critical_years)}. "
                    "The structure would not be acceptable to a senior bank "
                    "without changes."
                )
            else:
                st.write(
                    "All years meet the covenant threshold. "
                    "The structure is bankable under current assumptions."
                )

    if page == "Equity Case":
        st.title("Equity Case")
        st.write(
            "Management Buy-Out with an external minority investor. "
            "Holding period defined by the investor exit year. "
            "Exit mechanism: management buys out the investor."
        )

        equity_defaults = _default_equity_assumptions(input_model)
        sponsor_equity = st.session_state.get(
            "equity.sponsor_equity_eur",
            equity_defaults["sponsor_equity_eur"],
        )
        investor_equity = st.session_state.get(
            "equity.investor_equity_eur",
            equity_defaults["investor_equity_eur"],
        )
        total_equity = sponsor_equity + investor_equity
        sponsor_pct = sponsor_equity / total_equity if total_equity else 0.0
        investor_pct = investor_equity / total_equity if total_equity else 0.0
        exit_year = st.session_state.get(
            "equity.exit_year",
            equity_defaults["exit_year"],
        )
        exit_multiple = st.session_state.get(
            "equity.exit_multiple",
            equity_defaults["exit_multiple"],
        )
        exit_year = min(max(exit_year, 3), 7)
        exit_year_index = min(exit_year, 4)
        ebitda_exit = pnl_result[f"Year {exit_year_index}"]["ebitda"]
        net_debt_exit = (
            balance_sheet[exit_year_index]["financial_debt"]
            - balance_sheet[exit_year_index]["cash"]
        )
        enterprise_value_exit = ebitda_exit * exit_multiple
        equity_value_exit = enterprise_value_exit - net_debt_exit

        st.markdown("### Capital at Risk (Entry View)")
        entry_table = pd.DataFrame(
            [
                {
                    "Line Item": "Management (Sponsor) Equity",
                    "Equity (EUR)": format_currency(sponsor_equity),
                    "Ownership (%)": format_pct(sponsor_pct),
                },
                {
                    "Line Item": "External Investor Equity",
                    "Equity (EUR)": format_currency(investor_equity),
                    "Ownership (%)": format_pct(investor_pct),
                },
                {
                    "Line Item": "Total Equity",
                    "Equity (EUR)": format_currency(total_equity),
                    "Ownership (%)": format_pct(1.0),
                },
            ]
        )
        _render_custom_table_html(
            entry_table,
            set(),
            {"Total Equity"},
            {},
            year_labels=["Equity (EUR)", "Ownership (%)"],
        )

        sponsor_cashflows = []
        investor_cashflows = []
        sponsor_residual_value = equity_value_exit
        investor_exit_proceeds = equity_value_exit * investor_pct
        for year_index in range(8):
            if year_index == 0:
                sponsor_cf = -sponsor_equity
                investor_cf = -investor_equity
            elif year_index == exit_year:
                sponsor_cf = sponsor_residual_value
                investor_cf = investor_exit_proceeds
            else:
                sponsor_cf = 0.0
                investor_cf = 0.0
            sponsor_cashflows.append(sponsor_cf)
            investor_cashflows.append(investor_cf)

        sponsor_irr = _calculate_irr(sponsor_cashflows)
        investor_irr = _calculate_irr(investor_cashflows)
        st.session_state["equity.sponsor_irr"] = sponsor_irr
        st.session_state["equity.investor_irr"] = investor_irr

        st.markdown("### Headline Outcomes")
        investor_cols = st.columns(4)
        investor_cols[0].metric("External Investor – Invested Equity", format_currency(investor_equity))
        investor_cols[1].metric("External Investor – Exit Proceeds", format_currency(investor_exit_proceeds))
        investor_cols[2].metric(
            "External Investor – MOIC",
            f"{(investor_exit_proceeds / investor_equity) if investor_equity else 0:.2f}x",
        )
        investor_cols[3].metric("External Investor – IRR", format_pct(investor_irr))

        sponsor_cols = st.columns(4)
        sponsor_cols[0].metric("Management – Invested Equity", format_currency(sponsor_equity))
        sponsor_cols[1].metric("Management – Cash Proceeds at Exit", format_currency(sponsor_residual_value))
        sponsor_cols[2].metric("Management – IRR", format_pct(sponsor_irr))
        sponsor_cols[3].metric("Management – Ownership After Exit", "100%")

        st.markdown("### Exit Equity Bridge (Exit Year)")
        exit_year_label = f"Year {exit_year_index}"
        exit_bridge_rows = [
            {
                "Line Item": "Enterprise Value at Exit (EBIT × Multiple)",
                exit_year_label: enterprise_value_exit,
            },
            {
                "Line Item": "Net Debt at Exit",
                exit_year_label: net_debt_exit,
            },
            {
                "Line Item": "Excess Cash at Exit",
                exit_year_label: 0,
            },
            {
                "Line Item": "Total Equity Value at Exit",
                exit_year_label: equity_value_exit,
            },
            {
                "Line Item": "Investor Exit Proceeds",
                exit_year_label: investor_exit_proceeds,
            },
            {
                "Line Item": "Management Residual Equity Value",
                exit_year_label: sponsor_residual_value,
            },
        ]
        exit_bridge = pd.DataFrame(exit_bridge_rows)
        _render_custom_table_html(
            exit_bridge,
            set(),
            {"Total Equity Value at Exit", "Management Residual Equity Value"},
            {},
            year_labels=[exit_year_label],
        )

        with st.expander("Equity Cashflows", expanded=False):
            investor_cashflow_rows = [
                {
                    "Line Item": "Investor Cashflow",
                    **{f"Year {i}": investor_cashflows[i] for i in range(8)},
                },
            ]
            investor_cashflow_table = pd.DataFrame(investor_cashflow_rows)
            year_labels = [f"Year {i}" for i in range(8)]
            investor_cashflow_table = investor_cashflow_table[
                ["Line Item"] + year_labels
            ]
            _render_custom_table_html(
                investor_cashflow_table,
                set(),
                set(),
                {},
                year_labels=year_labels,
            )

            sponsor_cashflow_rows = [
                {
                    "Line Item": "Management Cashflow",
                    **{f"Year {i}": sponsor_cashflows[i] for i in range(8)},
                },
            ]
            sponsor_cashflow_table = pd.DataFrame(sponsor_cashflow_rows)
            sponsor_cashflow_table = sponsor_cashflow_table[
                ["Line Item"] + year_labels
            ]
            _render_custom_table_html(
                sponsor_cashflow_table,
                set(),
                set(),
                {},
                year_labels=year_labels,
            )


def _render_sidebar():
    nav_index = (
        NAV_OPTIONS.index(st.session_state["page_key"])
        if st.session_state["page_key"] in NAV_OPTIONS
        else 0
    )
    with st.sidebar:
        st.markdown(
            """
            <style>
              [data-testid="stSidebar"],
              [data-testid="stSidebarContent"] {
                background: #f0f2f6;
              }
              [data-testid="stSidebar"] > div {
                padding: 1rem 0.85rem;
              }
              [data-testid="stSidebar"] {
                min-width: 280px;
                max-width: 280px;
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label {
                display: flex;
                flex-direction: column;
                align-items: flex-start;
                padding: 0.35rem 0.6rem 0.35rem 0.8rem;
                border-radius: 6px;
                margin-bottom: 0.15rem;
                color: #111827;
                font-weight: 400;
                position: relative;
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:hover {
                background: #f1f3f5;
              }
              [data-testid="stSidebar"] div[role="radiogroup"] input,
              [data-testid="stSidebar"] div[role="radiogroup"] svg,
              [data-testid="stSidebar"] div[role="radiogroup"] > label > div:first-child {
                display: none;
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:has(input:checked) {
                background: #eef2f5;
                border-left: 3px solid #2563eb;
                color: #111827;
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(1)::before,
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(2)::before,
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(5)::before,
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(8)::before,
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(10)::before,
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(11)::before {
                display: block;
                font-size: 0.7rem;
                letter-spacing: 0.14em;
                text-transform: uppercase;
                color: #6b7280;
                margin: 0.9rem 0 0.35rem;
                width: 100%;
                pointer-events: none;
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(1)::before {
                content: "OVERVIEW";
                margin-top: 0;
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(2)::before {
                content: "OPERATING MODEL";
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(5)::before {
                content: "PLANNING";
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(8)::before {
                content: "FINANCING";
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(10)::before {
                content: "VALUATION";
              }
              [data-testid="stSidebar"] div[role="radiogroup"] > label:nth-child(11)::before {
                content: "SETTINGS";
              }
            </style>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("**MBO Financial Model**")
        selection = st.radio(
            "Navigation",
            NAV_OPTIONS,
            index=nav_index,
            key="main_navigation",
            label_visibility="collapsed",
        )
    return selection


def main():
    st.set_page_config(layout="wide")
    st.session_state.setdefault("page_key", "Overview")
    selection = _render_sidebar()
    st.session_state["page_key"] = selection
    run_app(st.session_state["page_key"])


if __name__ == "__main__":
    main()
